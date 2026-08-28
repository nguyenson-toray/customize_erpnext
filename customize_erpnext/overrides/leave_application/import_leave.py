# Copyright (c) 2026, IT Team - TIQN
# License: MIT

"""Import sổ nghỉ phép của HR (`AL_data.xlsx`) thành Leave Application **draft**.

Kế hoạch + số liệu đã kiểm chứng: `PLAN_IMPORT_AL_2026.md`.
Quy tắc nghiệp vụ: `QUY_DINH_NGHI_PHEP_2025.md`.


===============================================================================
QUY TRÌNH 4 BƯỚC
===============================================================================

    B1  run(dry_run=1)     đọc file, báo cáo, KHÔNG ghi gì          an toàn
    B2  run(dry_run=0)     tạo Leave Application draft              xoá lại được
    B3  verify()           đối chiếu số ngày ERP ↔ file HR          chỉ đọc
    --- HR rà soát 
    B4  submit_imported()  submit draft → sinh Attendance + sổ phép  🔴 MỘT CHIỀU

Chạy B1 → B2 → B3 liên tiếp được. **B4 phải đợi HR duyệt xong**, và sau B4 còn
một bước bắt buộc nữa (xem mục "SAU KHI SUBMIT").


===============================================================================
LỆNH CỤ THỂ
===============================================================================

B1 — thử trước, không ghi gì (LUÔN chạy bước này trước):

    bench --site erp.tiqn.local execute \
        customize_erpnext.overrides.leave_application.import_leave.run \
        --kwargs "{'dry_run': 1}"

    In ra: số dòng đọc được, số dòng bị loại, số đơn sẽ tạo, phân bố theo mã.

B2 — tạo draft thật:

    bench --site erp.tiqn.local execute \
        customize_erpnext.overrides.leave_application.import_leave.run \
        --kwargs "{'dry_run': 0}"

    Muốn thử một lô nhỏ trước cho chắc thì thêm `limit`:
        --kwargs "{'dry_run': 0, 'limit': 20}"

B3 — đối chiếu tổng số ngày nghỉ theo từng mã:

    bench --site erp.tiqn.local execute \
        customize_erpnext.overrides.leave_application.import_leave.verify

    Lệch phải giải thích được bằng đúng danh sách dòng lỗi của B2. Lệch không
    rõ nguồn ⇒ dừng, đừng submit.

B4 — submit (🔴 chỉ khi HR đã rà soát xong, KHÔNG rollback được):

    bench --site erp.tiqn.local execute \
        customize_erpnext.overrides.leave_application.import_leave.submit_imported \
        --kwargs "{'to_date': '2026-08-25'}"

    `to_date` giới hạn theo `to_date` của đơn (không phải `from_date`) — đơn vắt
    qua mốc sẽ sinh Attendance cho cả những ngày sau mốc, nên phải cắt ở đầu sau.


===============================================================================
THAM SỐ
===============================================================================

run(dry_run=1, path=None, batch_size=200, limit=0)
    dry_run     1 = chỉ báo cáo · 0 = ghi thật
    path        mặc định `AL_data.xlsx` cùng thư mục này
    batch_size  số đơn mỗi lần `db.commit()`
    limit       0 = tất cả · >0 = chỉ xử lý N đơn đầu (để thử)

submit_imported(batch_size=200, limit=0, to_date=None)
verify(path=None)


===============================================================================
CHẠY LẠI NHIỀU LẦN CÓ AN TOÀN KHÔNG? — CÓ
===============================================================================

`_already_there()` bỏ qua đơn đã tồn tại theo khoá:

    (employee, leave_type, from_date, to_date, docstatus < 2)

nên **cứ chạy lại trên toàn bộ file mỗi khi HR bổ sung dòng mới** — script tự bỏ
qua phần cũ và chỉ tạo phần mới. Không cần và không có tham số "chỉ import từ
dòng N".

⚠ Mặt trái: nếu HR **sửa** một dòng cũ (đổi ngày, đổi mã), khoá trên không khớp
nữa ⇒ sinh đơn MỚI trong khi đơn cũ vẫn còn ⇒ đếm trùng. Sau B2 nên kiểm:

    select name, employee, leave_type, from_date, to_date
    from `tabLeave Application`
    where docstatus = 0 and description like '[Import AL_data]%'
      and from_date < '<ngày đầu của khối dữ liệu mới>'

    Ra rỗng là sạch. Có dòng ⇒ đối chiếu tay với file HR trước khi submit.

Mỗi đơn có savepoint riêng, một dòng lỗi không cuốn theo cả lô.


===============================================================================
YÊU CẦU VỀ FILE EXCEL — 2 BẪY ĐÃ DÍNH
===============================================================================

Header 2 dòng, dữ liệu từ dòng 3, mỗi dòng = một nhân viên một ngày.
Cột dùng tới (0-based): 0 STT · 1 Mã NV · 30 Ngày · 31 Chấm công · 33 Ghi chú.

Mỗi lần HR mở và lưu lại file, Excel đổi hai thứ — cả hai đều từng làm script
chết, nên `_read()` và `_date()` giờ chịu được:

1. **Tên sheet đổi** (`Sheet1` → `Sheet2` → ...). Không neo theo tên nữa, lấy
   sheet đầu tiên. Neo cứng tên ⇒ `KeyError`.

2. **Cột ngày mất định dạng, về serial number của Excel** (số ngày kể từ
   30/12/1899, vd `46240`) thay vì `datetime`. `getdate(46240)` trả `None`
   ⇒ `TypeError: NoneType - NoneType` ở `_split_if_daycount_mismatch()`.
   `_date()` đã có nhánh quy đổi serial.

Nếu B1 nổ lỗi lạ, kiểm hai thứ này trước tiên.


===============================================================================
DÒNG BỊ LOẠI VÀ DÒNG LỖI
===============================================================================

"Dòng/đơn bị loại" — hai nhóm, **cố ý bỏ qua, không phải lỗi**:
  - mã ở cột `Chấm công` là **số** (`0.5` `0.9` `0.8`) = đi trễ / về sớm
  - đơn rơi **trọn** vào ngày nghỉ (CN/lễ) với loại nghỉ `include_holiday = 0`
    → HRMS tính `total_leave_days = 0`; tạo được cũng chỉ ra một đơn 0 ngày,
      không ghi sổ phép, không sinh Attendance. Đoạn nhiều ngày chỉ **vắt qua**
      Chủ Nhật thì KHÔNG bị loại.

"LỖI" — HRMS `validate()` chặn. Ba nhóm thường gặp, đều phải HR xử lý tay:
  - `applicable after N calendar days` — nhân viên mới chưa đủ thời gian
  - `has already applied for ... between` — đụng đơn đã có trong ERP
  - `the day(s) ... are holidays` — ngày xin nghỉ rơi vào ngày lễ


===============================================================================
XOÁ LẠI (chỉ áp dụng cho draft, TRƯỚC khi submit)
===============================================================================

    bench --site erp.tiqn.local console

    # Viết trên MỘT dòng: console tách block nhiều dòng thành từng cell riêng
    [frappe.delete_doc("Leave Application", n, force=1) for n in frappe.get_all("Leave Application", {"docstatus": 0, "description": ["like", "[Import AL_data]%"]}, pluck="name")]; frappe.db.commit()

Đơn đã submit (`docstatus = 1`) thì cancel chứ không xoá, và phải tính lại
Attendance sau đó.


===============================================================================
SAU KHI SUBMIT (B4) — BƯỚC BẮT BUỘC
===============================================================================

Submit sinh Attendance qua luồng Leave Application. Engine tính công
(`shift_type_optimized.py`) là luồng ghi độc lập thứ hai vào cùng các field đó và
luôn ghi sau cùng. Phải chạy lại engine cho đúng khoảng ngày vừa submit để hai
luồng đồng thuận:

    bench --site erp.tiqn.local console
    from customize_erpnext.overrides.shift_type.shift_type_optimized import bulk_update_attendance_optimized
    bulk_update_attendance_optimized("2026-07-27", "2026-08-25", force_sync=1)


===============================================================================
VÌ SAO TẠO DRAFT (`docstatus = 0`)
===============================================================================

  - chưa sinh Attendance (HRMS chỉ chạy `update_attendance()` ở `on_submit`)
  - chưa ghi `Leave Ledger Entry` ⇒ số dư phép không đổi
  - HR rà soát trên UI rồi mới duyệt và submit

`validate()` **vẫn chạy** khi insert, nên các dòng sai (rơi vào ngày nghỉ, ngoài
khoảng làm việc) vẫn bị bắt ngay ở bước này thay vì lúc submit hàng loạt.

`status = "Open"` chứ không phải `"Approved"`:
  - đúng nghĩa "chưa duyệt" của một bản nháp
  - tránh `notify_approval_status()` (`hrms/mixins/pwa_notifications.py:10`) sinh
    ~6.900 bản PWA Notification rác — hàm đó chỉ chạy khi status đổi sang
    Approved/Rejected và **không** bị chặn bởi `HR Settings.send_leave_notification`


===============================================================================
LỊCH SỬ CHẠY
===============================================================================

10/08/2026  7.097 đơn (dữ liệu tới 25/08/2026), sau đó đã submit hết
27/08/2026  file lên 10.614 dòng · 7.722 đơn suy ra · tạo mới 616 draft
            (27/07→25/08/2026) · bỏ qua 7.095 · 388 dòng đi trễ/về sớm · 11 lỗi
            verify lệch −16 ngày, giải thích trọn vẹn bằng 11 dòng lỗi đó
"""

import collections
import os
from datetime import date as _date_cls, timedelta

import frappe
from frappe.utils import getdate

# Cột (0-based) trong AL_data.xlsx — header 2 dòng, dữ liệu từ dòng 3
COL_STT, COL_EMP, COL_DATE, COL_CODE, COL_NOTE = 0, 1, 30, 31, 33

# Mã "Chấm công" của quy định → abbreviation của Leave Type
CODE_TO_ABBR = {
	"P": "P", "P/2": "P",
	"O": "O", "O/2": "O",
	"CO": "CO", "CO/2": "CO",
	"HL": "HL", "HL/2": "HL",
	"KL": "KL", "NB": "NB", "TS": "TS", "DS": "DS", "HS": "HS", "MC": "MC",
}

DEFAULT_FILE = os.path.join(os.path.dirname(__file__), "AL_data.xlsx")


# ---------------------------------------------------------------------------
# Đọc & phân loại
# ---------------------------------------------------------------------------


def _read(path: str) -> list:
	import openpyxl

	wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
	# Không neo theo tên sheet: mỗi lần HR lưu lại file, Excel đổi tên sheet
	# ("Sheet1" -> "Sheet2" ...). File chỉ có một sheet dữ liệu duy nhất.
	ws = wb[wb.sheetnames[0]]
	return [r for r in ws.iter_rows(min_row=3, values_only=True) if r[COL_EMP]]


# openpyxl trả về `datetime` cho ô có định dạng ngày. Nhưng khi HR mở/lưu lại file,
# hai cột ngày mất định dạng và về **serial number của Excel** (số ngày kể từ 30/12/1899)
# — toàn bộ 10.614 dòng của bản 27/08 đều là `int`. Không xử lý thì `getdate(46240)`
# trả None và `_split_if_daycount_mismatch()` nổ `NoneType - NoneType`.
_EXCEL_EPOCH = _date_cls(1899, 12, 30)


def _date(v):
	if v is None or v == "":
		return None
	if hasattr(v, "date"):
		return v.date()
	if isinstance(v, (int, float)):
		return _EXCEL_EPOCH + timedelta(days=int(v))
	return getdate(v)


def _code(row) -> str:
	return str(row[COL_CODE]).strip() if row[COL_CODE] is not None else ""


def _leave_type_by_abbr() -> dict:
	return {
		r.custom_abbreviation: r.name
		for r in frappe.get_all("Leave Type", fields=["name", "custom_abbreviation"])
		if r.custom_abbreviation
	}


def _holidays() -> set:
	"""Ngày nghỉ của mọi Holiday List đang gán (cấp Company, 0 nhân viên tự đặt riêng)."""
	lists = [
		a.holiday_list
		for a in frappe.get_all("Holiday List Assignment", fields=["holiday_list"])
	]
	return {
		h.holiday_date
		for h in frappe.get_all(
			"Holiday", filters={"parent": ["in", lists or [""]]}, fields=["holiday_date"]
		)
	}


# ---------------------------------------------------------------------------
# Gộp dòng → đơn nghỉ  (Cách C, PLAN_IMPORT_AL_2026.md mục 5)
# ---------------------------------------------------------------------------


def build_applications(rows: list, holidays: set) -> tuple:
	"""Trả về `(danh sách đơn, danh sách dòng bị loại)`.

	Mỗi đơn: `{employee, code, from_date, to_date, half_day, source_rows, stt, notes}`.
	"""
	leave_rows, skipped = [], []
	for r in rows:
		c = _code(r)
		if c in CODE_TO_ABBR:
			leave_rows.append(r)
		else:
			skipped.append((r, "đi trễ/về sớm hoặc mã bất thường: %r" % c))

	groups = collections.defaultdict(list)
	for r in leave_rows:
		groups[(r[COL_STT], r[COL_EMP], _code(r))].append(r)

	apps = []
	for (stt, emp, code), grp in groups.items():
		grp.sort(key=lambda r: _date(r[COL_DATE]))
		notes = [str(r[COL_NOTE]).strip() for r in grp if r[COL_NOTE]]

		if code.endswith("/2"):
			# Nửa ngày: HRMS chỉ cho MỘT half_day_date mỗi đơn ⇒ mỗi dòng một đơn
			for r in grp:
				d = _date(r[COL_DATE])
				apps.append(_app(emp, code, d, d, True, 1, stt, notes))
			continue

		# Trọn ngày: cắt thành đoạn liên tục, được nhảy qua ngày trong Holiday List
		run = [grp[0]]
		for prev, cur in zip(grp, grp[1:]):
			p, c_ = _date(prev[COL_DATE]), _date(cur[COL_DATE])
			gap = [p + timedelta(days=i) for i in range(1, (c_ - p).days)]
			if all(g in holidays for g in gap):
				run.append(cur)
			else:
				apps += _split_if_daycount_mismatch(emp, code, run, holidays, stt, notes)
				run = [cur]
		apps += _split_if_daycount_mismatch(emp, code, run, holidays, stt, notes)

	# Đơn nằm TRỌN trong ngày nghỉ (CN/lễ) của Leave Type có `include_holiday = 0`:
	# HRMS tính `total_leave_days = 0` ở `validate_balance_leaves()` rồi throw
	# "The day(s) on which you are applying for leave are holidays". Đây là dòng HR
	# ghi nhầm vào ngày vốn không đi làm, không phải lỗi hệ thống ⇒ loại ở đây để
	# không lẫn vào danh sách LỖI thật.
	#
	# ⚠ Chỉ loại khi TOÀN BỘ khoảng ngày là ngày nghỉ. Đoạn nhiều ngày chỉ VẮT QUA
	# Chủ Nhật thì giữ nguyên — HRMS trừ ngày đó khỏi `total_leave_days` và không throw.
	kept = []
	for a in apps:
		if _falls_entirely_on_holidays(a, holidays):
			skipped.append((None, "rơi trọn vào ngày nghỉ (CN/lễ): %s %s %s→%s"
			                % (a["employee"], a["code"], a["from_date"], a["to_date"])))
		else:
			kept.append(a)

	return kept, skipped


def _falls_entirely_on_holidays(a, holidays: set) -> bool:
	if frappe.db.get_value("Leave Type", _LT_CACHE[a["code"]], "include_holiday"):
		return False
	span = (a["to_date"] - a["from_date"]).days + 1
	return all((a["from_date"] + timedelta(days=i)) in holidays for i in range(span))


def _app(emp, code, f, t, half, n, stt, notes):
	return {
		"employee": emp, "code": code, "from_date": f, "to_date": t,
		"half_day": half, "source_rows": n, "stt": stt, "notes": notes,
	}


def _split_if_daycount_mismatch(emp, code, run, holidays, stt, notes):
	"""Guard mục 5.2 — chỉ gộp khi HRMS đếm ra ĐÚNG số dòng nguồn.

	`TS` có `include_holiday = 1` (HRMS đếm cả Chủ Nhật) và `DS` thì HR ghi nhiều dòng hơn số
	ngày HRMS đếm. Gộp trong hai ca đó làm sai sổ phép ⇒ tách về 1 đơn / 1 dòng.
	"""
	f, t = _date(run[0][COL_DATE]), _date(run[-1][COL_DATE])
	span = (t - f).days + 1
	include_holiday = frappe.db.get_value("Leave Type", _LT_CACHE[code], "include_holiday")
	hrms_days = span if include_holiday else span - sum(
		1 for i in range(span) if (f + timedelta(days=i)) in holidays
	)
	if abs(hrms_days - len(run)) < 0.001:
		return [_app(emp, code, f, t, False, len(run), stt, notes)]
	return [
		_app(emp, code, _date(r[COL_DATE]), _date(r[COL_DATE]), False, 1, stt, notes)
		for r in run
	]


_LT_CACHE = {}


# ---------------------------------------------------------------------------
# Chạy
# ---------------------------------------------------------------------------


def run(dry_run: int = 1, path: str | None = None, batch_size: int = 200, limit: int = 0):
	"""Import thành Leave Application **draft**. `dry_run=1` chỉ báo cáo, không ghi gì."""
	dry_run, limit = int(dry_run), int(limit)
	path = path or DEFAULT_FILE
	by_abbr = _leave_type_by_abbr()
	global _LT_CACHE
	_LT_CACHE = {c: by_abbr[a] for c, a in CODE_TO_ABBR.items() if a in by_abbr}

	missing = sorted({a for a in CODE_TO_ABBR.values() if a not in by_abbr})
	if missing:
		frappe.throw(f"Thiếu Leave Type cho mã: {missing}")

	rows = _read(path)
	holidays = _holidays()
	apps, skipped = build_applications(rows, holidays)
	apps.sort(key=lambda a: (a["employee"], a["from_date"]))
	if limit:
		apps = apps[:limit]

	print(f"Dòng đọc được          : {len(rows)}")
	print(f"Dòng/đơn bị loại        : {len(skipped)}")
	_reasons = collections.Counter(
		"rơi trọn vào ngày nghỉ (CN/lễ)" if r.startswith("rơi trọn") else "đi trễ/về sớm hoặc mã bất thường"
		for _, r in skipped
	)
	for _r, _n in _reasons.most_common():
		print(f"      {_n:>4} × {_r}")
	print(f"→ Leave Application     : {len(apps)}  "
	      f"({sum(1 for a in apps if a['from_date'] != a['to_date'])} phiếu nhiều ngày, "
	      f"{sum(1 for a in apps if a['half_day'])} nửa ngày)")
	if dry_run:
		print("\n[DRY RUN] không ghi gì. Bỏ `dry_run` để tạo draft.")
		_report_by_code(apps)
		return {"applications": len(apps), "skipped": len(skipped)}

	created = existing = 0
	failures = []
	frappe.flags.mute_messages = True
	for i, a in enumerate(apps, 1):
		# Savepoint cho TỪNG đơn. `frappe.db.rollback()` trần sẽ cuốn theo mọi bản ghi tốt
		# chưa commit trong lô — một dòng lỗi ở giữa lô 200 làm mất 199 dòng đã insert.
		frappe.db.savepoint("import_leave_row")
		try:
			if _already_there(a):
				existing += 1
				continue
			_insert_draft(a)
			created += 1
		except Exception as e:
			failures.append((a["employee"], a["code"], str(a["from_date"]),
			                 str(a["to_date"]), str(e).split("\n")[0][:120]))
			frappe.db.rollback(save_point="import_leave_row")
		if i % batch_size == 0:
			frappe.db.commit()
			print(f"  ... {i}/{len(apps)}  tạo {created}, bỏ qua {existing}, lỗi {len(failures)}")
	frappe.db.commit()
	frappe.flags.mute_messages = False

	print(f"\nTẠO MỚI (draft): {created} | đã có sẵn: {existing} | LỖI: {len(failures)}")
	_report_failures(failures)
	return {"created": created, "existing": existing, "failed": len(failures),
	        "failures": failures}


def _already_there(a) -> bool:
	return bool(frappe.db.exists("Leave Application", {
		"employee": a["employee"],
		"leave_type": _LT_CACHE[a["code"]],
		"from_date": a["from_date"],
		"to_date": a["to_date"],
		"docstatus": ["<", 2],
	}))


def _insert_draft(a):
	doc = frappe.new_doc("Leave Application")
	doc.employee = a["employee"]
	doc.leave_type = _LT_CACHE[a["code"]]
	doc.from_date = a["from_date"]
	doc.to_date = a["to_date"]
	if a["half_day"]:
		doc.half_day = 1
		doc.half_day_date = a["from_date"]
	# "Open" = bản nháp chưa duyệt; cũng để tránh PWA Notification rác (xem docstring module)
	doc.status = "Open"
	desc = [f"[Import AL_data] STT {a['stt']} · mã {a['code']}"]
	if a["notes"]:
		desc.append(" · ".join(dict.fromkeys(a["notes"])))
	doc.description = "\n".join(desc)
	doc.flags.ignore_permissions = True
	doc.insert()


def _report_by_code(apps):
	c = collections.Counter(a["code"] for a in apps)
	print("\nSố phiếu theo mã:")
	for k, v in sorted(c.items(), key=lambda x: -x[1]):
		print(f"   {k:>6} : {v}")


def _report_failures(failures):
	if not failures:
		return
	groups = collections.Counter(f[4][:70] for f in failures)
	print("\nLỗi theo nguyên nhân:")
	for msg, n in groups.most_common():
		print(f"   {n:>4} × {msg}")
	print("\n10 dòng đầu:")
	for f in failures[:10]:
		print(f"   {f[0]} {f[1]:>5} {f[2]}→{f[3]}: {f[4]}")


def submit_imported(batch_size: int = 200, limit: int = 0, to_date: str | None = None):
	"""Duyệt + submit các draft đã import. **Chỉ chạy khi HR đã rà soát xong.**

	🔴 Submit sẽ sinh Attendance và ghi Leave Ledger Entry — không còn quay lại được bằng
	rollback. Sau bước này phải chạy `bulk_update_attendance_optimized` cho toàn khoảng ngày
	để engine và luồng Leave Application đồng thuận.
	"""
	filters = {"docstatus": 0, "description": ["like", "[Import AL_data]%"]}
	if to_date:
		# Lọc theo `to_date` chứ không phải `from_date`: đơn vắt qua mốc sẽ sinh Attendance
		# cho cả những ngày SAU mốc, không nằm trong phạm vi được duyệt.
		filters["to_date"] = ["<=", to_date]
	names = [
		d.name for d in frappe.get_all(
			"Leave Application", filters=filters,
			fields=["name"], order_by="employee, from_date",
			limit_page_length=int(limit) or 0,
		)
	]
	print(f"Draft cần submit{f' (to_date <= {to_date})' if to_date else ''}: {len(names)}")
	done, failures = 0, []
	frappe.flags.mute_messages = True
	for i, name in enumerate(names, 1):
		frappe.db.savepoint("submit_leave_row")
		try:
			doc = frappe.get_doc("Leave Application", name)
			# status đã là "Approved" từ trước; chỉ gán khi cần để `has_value_changed()`
			# không kích hoạt notify_approval_status() -> đẻ PWA Notification rác.
			if doc.status != "Approved":
				doc.status = "Approved"
			doc.flags.ignore_permissions = True
			doc.submit()
			done += 1
		except Exception as e:
			failures.append((name, str(e).split("\n")[0][:120]))
			frappe.db.rollback(save_point="submit_leave_row")
		if i % batch_size == 0:
			frappe.db.commit()
			print(f"  ... {i}/{len(names)} submit {done}, lỗi {len(failures)}")
	frappe.db.commit()
	frappe.flags.mute_messages = False
	print(f"\nSUBMIT: {done} | LỖI: {len(failures)}")
	for f in failures[:10]:
		print("   ", f)
	return {"submitted": done, "failed": len(failures)}


def verify(path: str | None = None):
	"""Đối chiếu số ngày nghỉ ERP (draft) với file HR, theo từng mã. Ngưỡng 0,01 ngày."""
	path = path or DEFAULT_FILE
	rows = _read(path)
	hr = collections.Counter()
	for r in rows:
		c = _code(r)
		if c in CODE_TO_ABBR:
			hr[CODE_TO_ABBR[c]] += 0.5 if c.endswith("/2") else 1

	abbr = {r.name: r.custom_abbreviation for r in frappe.get_all(
		"Leave Type", fields=["name", "custom_abbreviation"])}
	erp = collections.Counter()
	imported = {"description": ("like", "[Import AL_data]%")}
	for la in frappe.get_all("Leave Application", filters=imported,
	                         fields=["leave_type", "total_leave_days"]):
		erp[abbr.get(la.leave_type)] += la.total_leave_days

	print("Leave Application từ import:", frappe.db.count("Leave Application", imported))
	for st in (0, 1, 2):
		print(f"   docstatus={st}: {frappe.db.count('Leave Application', {**imported, 'docstatus': st})}")

	print(f"\n{'mã':6}{'HR (ngày)':>12}{'ERP':>12}{'lệch':>9}")
	diffs = []
	for k in sorted(set(hr) | set(erp)):
		h, e = hr[k], erp[k]
		flag = ""
		if abs(h - e) > 0.01:
			diffs.append((k, h, e))
			flag = "  ← LỆCH"
		print(f"{k:6}{h:>12.1f}{e:>12.1f}{e - h:>9.1f}{flag}")
	print(f"{'TỔNG':6}{sum(hr.values()):>12.1f}{sum(erp.values()):>12.1f}"
	      f"{sum(erp.values()) - sum(hr.values()):>9.1f}")
	return diffs
