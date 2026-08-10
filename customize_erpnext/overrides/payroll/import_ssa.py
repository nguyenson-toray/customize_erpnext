# Copyright (c) 2026, IT Team - TIQN and contributors
# For license information, please see license.txt

"""Import lương HĐLĐ vào `Salary Structure Assignment` — nhận cả `.xlsx` lẫn `.csv`.

Vì sao phải viết parser riêng thay vì dùng Data Import của Frappe: file HR gửi là
**báo cáo đã định dạng** (tiêu đề gộp 2 dòng, chia mục "I. Management" / "II. Production",
có dòng Total, có dòng ẩn), không phải template import phẳng.

**Nên làm hai bước:** `to_csv()` trích file HR ra CSV phẳng (giữ lại trong repo để đối chiếu
và hồi quy về sau), rồi `run()` import từ CSV đó. CSV bền hơn — không phụ thuộc vị trí cột,
mở bằng Excel vẫn đúng dấu tiếng Việt (có BOM UTF-8), và đọc được bằng mắt khi cần truy vết.

    bench --site erp.tiqn.local execute \\
        customize_erpnext.overrides.payroll.import_ssa.run --kwargs "{'path': '...xlsx'}"

Mặc định **dry-run** — chỉ in ra xem sẽ tạo gì. Truyền `commit=True` mới ghi thật.

⚠ Chỉ lấy nhóm cột **"Salary in Labour Contract"** (cột 14–24) là giá trị ổn định theo
hợp đồng. Nhóm "Actual Salary in the Month" (cột 25+) là số thực trả của riêng kỳ đó,
thuộc về Salary Slip chứ không phải SSA.
"""

import frappe
from frappe import _
from frappe.utils import add_days, cint, flt, getdate

from customize_erpnext.overrides.salary_structure_assignment.salary_structure_assignment import (
	SI_BASE_FIELDS,
)

# Vị trí cột trong file Excel của HR. Đổi layout thì sửa ở đây.
COLUMNS = {
	"employee": 3,
	"employee_name": 4,
	"contract_style": 5,
	"hr_si_base": 12,  # "Salary for SI & HI" — mức HR đăng ký với cơ quan BHXH
	"base": 14,
	"custom_technical_allowance": 15,
	"custom_position_allowance": 16,
	"custom_accommodation_allowance": 17,
	"custom_commuting_allowance": 18,
	"custom_phone_allowance": 19,
	"custom_pccc_allowance": 20,
	"custom_atvs_allowance": 21,
	"custom_supporting_stages_allowance": 22,
	"custom_attendance_incentive": 23,
	"custom_responsibility_incentive": 24,
}
AMOUNT_FIELDS = [k for k in COLUMNS if k.startswith("custom_") or k == "base"]

HEADER_ROWS = 9  # dữ liệu bắt đầu từ dòng 10


def run(path: str, salary_structure: str = "TIQN - Standard (ALL components)",
        from_date: str | None = None, commit: bool = False) -> None:
	rows = read_rows(path)
	print(f"Đọc được {len(rows)} dòng nhân viên từ {path}\n")

	plan, skipped = [], []
	for row in rows:
		issue = _check(row, salary_structure, from_date)
		(skipped if issue else plan).append((row, issue))

	_report(plan, skipped, from_date)

	if not commit:
		print("\n>>> DRY-RUN — chưa ghi gì. Truyền commit=True để tạo thật.")
		return

	created = 0
	for row, _issue in plan:
		_create_ssa(row, salary_structure, from_date)
		created += 1
	frappe.db.commit()
	print(f"\n>>> ĐÃ TẠO {created} Salary Structure Assignment.")


def to_csv(xlsx_path: str, csv_path: str) -> int:
	"""Trích file Excel của HR ra CSV phẳng. Trả về số dòng đã ghi.

	Ghi bằng `utf-8-sig` (BOM) để Excel mở không vỡ dấu tiếng Việt — quy ước dự án.
	"""
	import csv

	rows = read_rows(xlsx_path)
	fields = list(COLUMNS)
	with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
		writer = csv.DictWriter(f, fieldnames=fields)
		writer.writeheader()
		for row in rows:
			writer.writerow({k: (int(flt(row.get(k))) if k in AMOUNT_FIELDS or k == "hr_si_base"
			                     else (row.get(k) or "")) for k in fields})
	print(f"Đã ghi {len(rows)} dòng vào {csv_path}")
	return len(rows)


def read_rows(path: str) -> list:
	"""Đọc file, bỏ dòng tiêu đề / dòng mục / dòng Total.

	Nhận diện dòng dữ liệu bằng mã nhân viên bắt đầu bằng `TIQN` — đơn giản và bền
	hơn là đếm dòng, vì file có dòng ẩn và số dòng mỗi mục thay đổi theo kỳ.
	"""
	if path.lower().endswith(".csv"):
		return _read_csv(path)

	import openpyxl

	ws = openpyxl.load_workbook(path, data_only=True).worksheets[0]
	rows = []
	for r in range(HEADER_ROWS + 1, ws.max_row + 1):
		employee = ws.cell(r, COLUMNS["employee"]).value
		if not (isinstance(employee, str) and employee.strip().upper().startswith("TIQN")):
			continue
		row = frappe._dict(excel_row=r, employee=employee.strip())
		for field, col in COLUMNS.items():
			if field == "employee":
				continue
			value = ws.cell(r, col).value
			row[field] = flt(value) if field in AMOUNT_FIELDS or field == "hr_si_base" else value
		rows.append(row)
	return rows


def _read_csv(path: str) -> list:
	"""CSV phẳng do `to_csv()` sinh ra — tiêu đề chính là tên field, không phụ thuộc vị trí."""
	import csv

	rows = []
	with open(path, encoding="utf-8-sig", newline="") as f:
		for i, raw in enumerate(csv.DictReader(f), start=2):
			employee = (raw.get("employee") or "").strip()
			if not employee.upper().startswith("TIQN"):
				continue
			row = frappe._dict(excel_row=i, employee=employee)
			for field in COLUMNS:
				if field == "employee":
					continue
				value = raw.get(field)
				row[field] = flt(value) if field in AMOUNT_FIELDS or field == "hr_si_base" else value
			rows.append(row)
	return rows


def _check(row, salary_structure: str, from_date: str | None) -> str | None:
	"""Trả về lý do bỏ qua, hoặc None nếu tạo được."""
	if not frappe.db.exists("Employee", row.employee):
		return _("không có trong hệ thống")
	if not flt(row.base):
		return _("lương HĐLĐ = 0")

	effective = from_date or frappe.db.get_value("Employee", row.employee, "date_of_joining")
	if not effective:
		return _("chưa có ngày vào làm")

	if frappe.db.exists("Salary Structure Assignment", {
		"employee": row.employee, "from_date": getdate(effective), "docstatus": 1,
	}):
		return _("đã có SSA cùng ngày hiệu lực")
	return None


def _create_ssa(row, salary_structure: str, from_date: str | None):
	employee = frappe.get_cached_doc("Employee", row.employee)
	doc = frappe.new_doc("Salary Structure Assignment")
	doc.employee = row.employee
	doc.salary_structure = salary_structure
	doc.from_date = getdate(from_date or employee.date_of_joining)
	doc.company = employee.company
	doc.currency = frappe.get_cached_value("Company", employee.company, "default_currency") or "VND"
	for field in AMOUNT_FIELDS:
		doc.set(field, flt(row.get(field)))

	# `custom_si_base` tự tính từ 8 khoản. Chỉ bật ghi đè khi mức HR đăng ký với cơ quan
	# BHXH lệch khỏi tổng — giữ đúng số HR đang đóng thay vì ép theo công thức.
	if flt(row.hr_si_base) and flt(row.hr_si_base) != _computed_si_base(row):
		doc.custom_si_base_override = 1
		doc.custom_si_base = flt(row.hr_si_base)

	doc.insert(ignore_permissions=True)
	doc.submit()
	return doc


def _computed_si_base(row) -> float:
	"""Căn cứ đóng BH theo công thức — dùng chung `SI_BASE_FIELDS` với controller SSA
	để hai nơi không trôi khỏi nhau."""
	return sum(flt(row.get(f)) for f in SI_BASE_FIELDS)


def _report(plan, skipped, from_date):
	if plan:
		print(f"SẼ TẠO {len(plan)} bản ghi:")
		print(f"  {'Mã NV':11} {'Họ tên':24} {'Hiệu lực':11} {'Lương HĐLĐ':>13} {'Căn cứ BH':>13}  Ghi chú")
		for row, _i in plan:
			eff = from_date or frappe.db.get_value("Employee", row.employee, "date_of_joining")
			computed = _computed_si_base(row)
			note = "" if flt(row.hr_si_base) == computed else f"⚠ HR khai {flt(row.hr_si_base):,.0f}"
			print(f"  {row.employee:11} {str(row.employee_name)[:22]:24} {str(eff):11} "
			      f"{flt(row.base):>13,.0f} {computed:>13,.0f}  {note}")
	if skipped:
		print(f"\nBỎ QUA {len(skipped)} dòng:")
		for row, issue in skipped:
			print(f"  {row.employee:11} {str(row.employee_name)[:24]:26} — {issue}")


# ---------------------------------------------------------------------------
# Đồng bộ loại hợp đồng
# ---------------------------------------------------------------------------

# Cột "Contract style" của HR -> Employment Type trong ERP.
# Chỉ ánh xạ những giá trị CHẮC CHẮN. Giá trị khác thì bỏ qua và báo ra, không đoán —
# `employment_type` quyết định có trừ BHXH/đoàn phí hay không, đoán sai là sai lương.
CONTRACT_STYLE_MAP = {
	"probation": "30 Days Probationary Contract",
}


def sync_employment_type(path: str, commit: bool = False) -> None:
	"""Điền `Employee.employment_type` từ cột `contract_style` của file lương.

	Vì sao cần: toàn bộ quy tắc thử việc trong payroll (không trừ BHXH/BHYT/BHTN, không
	trừ đoàn phí, hoàn 21,5%, thuế suất cố định 10%) đọc `Employee.employment_type`.
	Field này do module Labor Contract đồng bộ, mà module đó đang tắt trong `hooks.py`,
	nên nhân viên mới và nhân viên đã nghỉ đều đang để trống.

	Mặc định dry-run.
	"""
	updates, skipped = [], []
	for row in read_rows(path):
		style = (row.get("contract_style") or "").strip()
		target = CONTRACT_STYLE_MAP.get(style.lower())
		if not target:
			skipped.append((row, style or "(trống)"))
			continue
		current = frappe.db.get_value("Employee", row.employee, "employment_type")
		if current == target:
			continue
		updates.append((row, current, target))

	if updates:
		print(f"SẼ SỬA {len(updates)} nhân viên:")
		for row, current, target in updates:
			print(f"  {row.employee:11} {str(row.employee_name)[:22]:24} "
			      f"{current or '(trống)':38} -> {target}")
	if skipped:
		print(f"\nBỎ QUA {len(skipped)} — không có ánh xạ chắc chắn cho 'Contract style':")
		for row, style in skipped:
			print(f"  {row.employee:11} {str(row.employee_name)[:22]:24} contract_style = {style}")

	if not commit:
		print("\n>>> DRY-RUN — chưa ghi gì. Truyền commit=True để sửa thật.")
		return

	for row, _current, target in updates:
		frappe.db.set_value("Employee", row.employee, "employment_type", target,
		                    update_modified=False)
		frappe.clear_document_cache("Employee", row.employee)
	frappe.db.commit()
	print(f"\n>>> ĐÃ SỬA {len(updates)} nhân viên.")


# ---------------------------------------------------------------------------
# Suy luận loại hợp đồng từ ngày vào làm
# ---------------------------------------------------------------------------

# Chuỗi hợp đồng cố định của TIQN. Độ dài từng chặng lấy từ `Employment Type.custom_period`
# (nguồn duy nhất — module Labor Contract cũng đọc chỗ đó), không hardcode ở đây.
CONTRACT_CHAIN = [
	"30 Days Probationary Contract",
	"1 Year Employment Contract",
	"3 Year Employment Contract",
	"Indefinite-term Employment Contract",
]


def infer_employment_type(employee: str, as_on=None) -> tuple:
	"""Suy ra loại hợp đồng đang hiệu lực tại `as_on`, đi theo chuỗi từ ngày vào làm.

	Trả về `(employment_type, chain)` — `chain` là các chặng đã tính, để in ra cho người
	dùng tự kiểm chứng thay vì phải tin một con số.

	⚠ Đây là **suy luận, không phải hồ sơ**. Giả định: chuỗi hợp đồng chạy liên tục không
	gián đoạn kể từ ngày vào làm. Nếu NLĐ từng nghỉ rồi quay lại, hoặc HR ký lệch mốc,
	kết quả sẽ sai. Chỉ dùng khi **không có** bản ghi `Labor Contract` để tra.

	Thời gian thử việc lấy từ `Employee.custom_probation_days`; trống thì dùng 30 ngày —
	948/1010 nhân viên có khai đều là 30.
	"""
	emp = frappe.db.get_value(
		"Employee", employee,
		["date_of_joining", "relieving_date", "custom_probation_days"], as_dict=True,
	)
	if not emp or not emp.date_of_joining:
		return None, []

	as_on = getdate(as_on or emp.relieving_date or frappe.utils.nowdate())
	probation = "60 Days Probationary Contract" if cint(emp.custom_probation_days) == 60 else CONTRACT_CHAIN[0]
	stages = [probation] + CONTRACT_CHAIN[1:]

	start = getdate(emp.date_of_joining)
	chain = []
	for stage in stages:
		period = cint(frappe.db.get_value("Employment Type", stage, "custom_period"))
		end = add_days(start, period - 1) if period else None
		chain.append((stage, start, end))
		if end is None or as_on <= end:
			return stage, chain
		start = add_days(end, 1)
	return chain[-1][0], chain


def sync_inferred_employment_type(employees: list | str, as_on=None, commit: bool = False) -> None:
	"""Điền `employment_type` bằng suy luận cho các nhân viên chưa có.

	Chỉ động vào bản ghi đang **trống** — không ghi đè giá trị đã có, vì giá trị đã có
	có thể do module Labor Contract đồng bộ từ hợp đồng thật.
	"""
	if isinstance(employees, str):
		employees = [e.strip() for e in employees.split(",") if e.strip()]

	updates = []
	for employee in employees:
		current = frappe.db.get_value("Employee", employee, "employment_type")
		if current:
			print(f"  {employee}: đã có '{current}' — bỏ qua")
			continue
		inferred, chain = infer_employment_type(employee, as_on)
		if not inferred:
			print(f"  {employee}: thiếu ngày vào làm — bỏ qua")
			continue

		emp = frappe.db.get_value("Employee", employee,
		                          ["employee_name", "date_of_joining", "relieving_date"], as_dict=True)
		ref = getdate(as_on or emp.relieving_date or frappe.utils.nowdate())
		print(f"\n  {employee} — {emp.employee_name}")
		print(f"    vào làm {emp.date_of_joining} | mốc xét {ref}"
		      f"{' (ngày nghỉ việc)' if emp.relieving_date and not as_on else ''}")
		for stage, start, end in chain:
			mark = " <-- đang hiệu lực" if stage == inferred else ""
			print(f"      {str(start):10} → {str(end) if end else 'không thời hạn':14}  {stage}{mark}")
		updates.append((employee, inferred))

	if not commit:
		print("\n>>> DRY-RUN — chưa ghi gì. Truyền commit=True để gán thật.")
		return

	for employee, inferred in updates:
		frappe.db.set_value("Employee", employee, "employment_type", inferred, update_modified=False)
		frappe.clear_document_cache("Employee", employee)
	frappe.db.commit()
	print(f"\n>>> ĐÃ GÁN {len(updates)} nhân viên.")
