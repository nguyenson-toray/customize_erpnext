"""Phase B — Action APIs & Phase C — Dashboard APIs for Uniform Control."""
import frappe
from frappe import _
from frappe.utils import flt, cint, today, add_days

from customize_erpnext.uniform_control.utils import (
    get_eligible_employees_for_allocation,
    get_uniform_warehouse,
    get_item_available_qty,
)


def _check_permission(doctype="Uniform Allocation", ptype="read"):
    if not frappe.has_permission(doctype, ptype):
        frappe.throw(_("Not permitted"), frappe.PermissionError)


# ─────────────────────────────── Phase B ───────────────────────────────────

@frappe.whitelist()
def get_eligible_employees(
    allocation_type, uniform_type=None, department=None,
    joining_date_from=None, joining_date_to=None,
    group=None, gender=None, due_date_from=None, due_date_to=None,
    overdue_only=0,
):
    """
    Return employees eligible for the given allocation_type.
    Includes suggested item_code, qty, available_qty per row.
    """
    _check_permission()
    return get_eligible_employees_for_allocation(
        allocation_type, uniform_type, department,
        joining_date_from, joining_date_to, group, gender,
        due_date_from, due_date_to, overdue_only,
    )


ALLOWED_HEADER_FIELDS = {
    "naming_series", "posting_date", "allocation_type", "company", "set_warehouse",
    "uniform_type_filter", "department_filter", "group_filter", "gender_filter",
    "joining_date_from", "joining_date_to",
}
ALLOWED_ITEM_FIELDS = {
    "employee", "item_code", "qty", "issue_reason", "shoe_rack_location", "remark",
}


@frappe.whitelist(methods=["POST"])
def create_allocation(header, items):
    """
    Create a Draft Uniform Allocation from payload (only whitelisted fields).
    header: dict with posting_date, allocation_type, company, set_warehouse, etc.
    items: list of dicts (employee, item_code, qty, issue_reason, ...)
    """
    _check_permission(ptype="create")

    if isinstance(header, str):
        header = frappe.parse_json(header)
    if isinstance(items, str):
        items = frappe.parse_json(items)

    alloc = frappe.new_doc("Uniform Allocation")
    alloc.update({k: v for k, v in (header or {}).items() if k in ALLOWED_HEADER_FIELDS})
    for item in items or []:
        alloc.append("items", {k: v for k, v in item.items() if k in ALLOWED_ITEM_FIELDS})

    alloc.insert(ignore_permissions=True)
    return alloc.name


@frappe.whitelist(methods=["POST"])
def submit_allocation(name):
    """Submit Allocation → creates Material Issue + updates profiles."""
    _check_permission(ptype="submit")

    alloc = frappe.get_doc("Uniform Allocation", name)
    if alloc.docstatus != 0:
        frappe.throw(_("Allocation {0} is not in Draft state.").format(name))
    alloc.submit()
    return {"name": alloc.name, "stock_entry": alloc.stock_entry}


@frappe.whitelist(methods=["POST"])
def receive_stock(items):
    """
    Simplified stock receipt for HR.
    items: list of {item_code, qty, valuation_rate} dicts.
    Creates a Material Receipt Stock Entry into the Uniform warehouse.
    """
    _check_permission(doctype="Stock Entry", ptype="create")

    if isinstance(items, str):
        items = frappe.parse_json(items)

    warehouse = get_uniform_warehouse()
    if not warehouse:
        frappe.throw(_("Uniform Warehouse is not configured in Uniform Setting."))

    # Derive company from the warehouse — user default may belong to another company
    company = (
        frappe.db.get_value("Warehouse", warehouse, "company")
        or frappe.defaults.get_user_default("Company")
        or frappe.db.get_single_value("Global Defaults", "default_company")
    )

    se = frappe.new_doc("Stock Entry")
    se.stock_entry_type = "Material Receipt"
    se.posting_date = today()
    se.company = company
    se.to_warehouse = warehouse

    for item in items:
        se.append("items", {
            "item_code": item.get("item_code"),
            "qty": cint(item.get("qty")),
            "t_warehouse": warehouse,
            "basic_rate": flt(item.get("valuation_rate", 0)),
        })

    se.insert(ignore_permissions=True)
    se.submit()
    return {"name": se.name, "status": "Submitted"}


# ─────────────────────────────── Phase C ───────────────────────────────────

@frappe.whitelist()
def get_dashboard_summary():
    """
    Return summary data for the Uniform Control Dashboard:
    - total variants in stock
    - low stock count
    - employees due soon / overdue
    - recent allocations count
    """
    _check_permission()

    warehouse = get_uniform_warehouse()

    # Stock summary — without a configured warehouse there is nothing to count
    bins = []
    reorder_levels = {}
    if warehouse:
        bins = frappe.get_all(
            "Bin",
            filters={"warehouse": warehouse},
            fields=["item_code", "actual_qty"],
            order_by="item_code asc",
        )
        reorder_levels = {
            r.parent: flt(r.warehouse_reorder_level)
            for r in frappe.get_all(
                "Item Reorder",
                filters={"warehouse": warehouse},
                fields=["parent", "warehouse_reorder_level"],
            )
        }
    low_stock_count = sum(
        1 for b in bins
        if reorder_levels.get(b.item_code, 0) > 0
        and flt(b.actual_qty) <= reorder_levels[b.item_code]
    )

    # Due / overdue — computed from next_due_date (the stored `status` field is
    # only refreshed on profile save, so it goes stale between saves), and
    # scoped to Active managed employees like the due list itself.
    from customize_erpnext.uniform_control.utils import get_employee_id_prefix

    setting_reminder = cint(
        frappe.db.get_single_value("Uniform Setting", "reminder_days_before")
    ) or 30
    prefix = get_employee_id_prefix()
    prefix_cond = "AND p.employee LIKE %(prefix)s" if prefix else ""
    due_soon, overdue = frappe.db.sql(
        f"""
        SELECT
            SUM(eui.next_due_date >= %(today)s AND eui.next_due_date <= %(cutoff)s),
            SUM(eui.next_due_date < %(today)s)
        FROM `tabEmployee Uniform Item` eui
        INNER JOIN `tabEmployee Uniform Profile` p ON p.name = eui.parent
        INNER JOIN `tabEmployee` e ON e.name = p.employee
        WHERE eui.next_due_date IS NOT NULL
          AND e.status = 'Active' {prefix_cond}
        """,
        {"today": today(), "cutoff": add_days(today(), setting_reminder),
         "prefix": f"{prefix}%"},
    )[0]

    # Recent allocations (last 30 days)
    recent_allocs = frappe.db.count(
        "Uniform Allocation",
        {"docstatus": 1, "posting_date": [">=", frappe.utils.add_days(today(), -30)]},
    )

    return {
        "total_variants_in_stock": len(bins),
        "low_stock_variants": low_stock_count,
        "employees_due_soon": cint(due_soon),
        "employees_overdue": cint(overdue),
        "allocations_last_30_days": cint(recent_allocs),
        "warehouse": warehouse,
    }


@frappe.whitelist()
def get_due_items(limit=100, due_before=None):
    """Employees with uniform items due for reissue up to `due_before` (Active
    managed employees), counting EVERY reissue cycle in the horizon.

    Default `due_before` = today + 1 year. Each row carries qty_per_cycle,
    cycles and total_qty; `needed` aggregates total_qty (multi-cycle) per variant.
    Returns {"rows": [...], "needed": {variant: total_qty}}."""
    _check_permission(doctype="Employee Uniform Profile")

    from math import ceil
    from frappe.utils import getdate, date_diff
    from customize_erpnext.uniform_control.utils import (
        reissue_demand, leaver_missed_demand,
    )

    setting = frappe.get_single("Uniform Setting")
    to_date = due_before or add_days(today(), 365)

    res = reissue_demand(to_date, setting)

    # "Est. for Leavers" — MEASURED model: missed re-issues of actual leavers
    # in the Attrition Window → monthly average × months of this horizon,
    # rounded UP, clamped to the variant's gross demand.
    enabled = bool(cint(setting.get("consider_attrition")))
    est, attrition = {}, {"enabled": enabled, "months": 0, "persons": 0,
                          "monthly_total": 0, "period_months": 0}
    if enabled:
        miss = leaver_missed_demand(setting.get("attrition_months"))
        # day-based months so a 1-month horizon inside one calendar month ≠ 0
        period = max(0.0, date_diff(getdate(to_date), getdate(today())) / 30.44)
        for v, monthly in miss["monthly"].items():
            e = min(cint(res["needed"].get(v, 0)), ceil(monthly * period))
            if e > 0:
                est[v] = e
        attrition.update({
            "months": miss["months"], "persons": miss["persons"],
            "monthly_total": round(miss["total_qty"] / miss["months"], 1),
            "period_months": round(period, 1), "window": miss["window"],
        })

    rows = res["rows"]
    # `qty` kept as the per-cycle qty for backward-compatible callers
    for r in rows:
        r["qty"] = r["qty_per_cycle"]
    if cint(limit):
        rows = rows[: cint(limit)]
    return {
        "rows": rows,
        "needed": res["needed"],        # full (gross) demand, from tracking
        "est_for_leavers": est,         # measured slice per variant (positive)
        "attrition": attrition,
    }


@frappe.whitelist(methods=["POST"])
def rebuild_tracking(employee):
    """Re-sync an employee's Issuance Tracking from SUBMITTED Uniform Allocations
    (the source of truth). Templates with allocations are recomputed; legacy rows
    with no allocation are left untouched. Then next_due/status are refreshed."""
    _check_permission(doctype="Employee Uniform Profile", ptype="write")
    from customize_erpnext.uniform_control.doctype.employee_uniform_profile.employee_uniform_profile import _template_of

    name = frappe.db.get_value("Employee Uniform Profile", {"employee": employee})
    if not name:
        return {"ok": False}

    rows = frappe.db.sql(
        """
        SELECT uai.item_code, uai.qty, ua.posting_date,
               COALESCE(NULLIF(i.variant_of, ''), i.name) AS template
        FROM `tabUniform Allocation Item` uai
        JOIN `tabUniform Allocation` ua ON ua.name = uai.parent
        JOIN `tabItem` i ON i.name = uai.item_code
        WHERE uai.docstatus = 1 AND uai.employee = %s
        ORDER BY ua.posting_date ASC
        """,
        employee, as_dict=True,
    )
    agg = {}
    for r in rows:
        a = agg.setdefault(r.template, {"total": 0, "last": None, "item": None, "last_qty": 0})
        a["total"] += cint(r.qty)
        d = frappe.utils.getdate(r.posting_date)
        if a["last"] is None or d > a["last"]:
            a["last"], a["item"], a["last_qty"] = d, r.item_code, cint(r.qty)
        elif d == a["last"]:
            a["last_qty"] += cint(r.qty)
            a["item"] = r.item_code

    profile = frappe.get_doc("Employee Uniform Profile", name)
    by_tmpl = {_template_of(row.item_template): row for row in profile.all_tracking_rows()}
    for tmpl, a in agg.items():
        row = by_tmpl.get(tmpl) or profile.append(profile.tracking_field_for(tmpl), {})
        row.item_template = a["item"]
        row.last_issue_date = a["last"]
        row.last_issue_qty = a["last_qty"]
        row.total_issued_qty = a["total"]
    profile.save(ignore_permissions=True)  # validate refreshes next_due/status
    return {"ok": True, "rebuilt": len(agg)}


@frappe.whitelist(methods=["POST"])
def recompute_all_tracking():
    """Recompute next_due_date + status on every managed Uniform Profile.
    Run this after changing a rule's Reissue Cycle (Months): the stored
    next_due_date is only refreshed when a profile is saved, so existing
    tracking keeps the OLD cycle until this bulk recompute runs.
    Profile.validate() does the recomputation on save."""
    frappe.only_for(("System Manager", "Uniform Manager"))

    from customize_erpnext.uniform_control.utils import get_employee_id_prefix
    prefix = get_employee_id_prefix()
    filters = {}
    if prefix:
        filters["employee"] = ["like", f"{prefix}%"]

    names = frappe.get_all("Employee Uniform Profile", filters=filters, pluck="name")
    done = 0
    for i, name in enumerate(names, start=1):
        try:
            frappe.get_doc("Employee Uniform Profile", name).save(ignore_permissions=True)
            done += 1
        except Exception:
            frappe.log_error(frappe.get_traceback(), f"Recompute tracking failed: {name}")
        if i % 200 == 0:
            frappe.db.commit()
    frappe.db.commit()
    return {"ok": True, "profiles": len(names), "recomputed": done}


@frappe.whitelist(methods=["POST"])
def apply_default_rules(employee=None):
    """Apply Default Rules to one profile (employee given) or ALL managed
    profiles without Manual Override. Returns count of profiles changed."""
    _check_permission(doctype="Employee Uniform Profile", ptype="write")

    from customize_erpnext.uniform_control import utils

    setting = frappe.get_single("Uniform Setting")
    filters = {}
    if employee:
        filters["employee"] = employee
    else:
        filters["manual_override"] = 0

    # Single employee = explicit intent → overwrite. Bulk = fill empty only,
    # so real per-person assignments (e.g. imported shirts) are preserved.
    force = bool(employee)
    names = frappe.get_all("Employee Uniform Profile", filters=filters, pluck="name")
    changed = 0
    for name in names:
        profile = frappe.get_doc("Employee Uniform Profile", name)
        if not employee and profile.manual_override:
            continue
        fields = utils.apply_default_rules(profile, setting=setting, force=force)
        if fields:
            profile.save(ignore_permissions=True)
            changed += 1

    return {"profiles": len(names), "changed": changed}


@frappe.whitelist()
def export_dashboard_excel(due_before=None):
    """Download dashboard data as .xlsx — matches the on-screen view:
      sheet 1 "Stock Plan": Stock | Reissue Need (gross, from tracking) |
        Est. for Leavers (negative, informational) | Total | Shortfall;
      sheet 2 "Employees Due for Issue": reissue rows with cycles (multi-cycle).
    `due_before` = reissue horizon (default today)."""
    _check_permission()

    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Font

    from customize_erpnext.uniform_control.api.uniform_excel_api import get_uniform_stock_excel

    # Default horizon = today (same as the dashboard's default filter)
    horizon = due_before or today()
    due = get_due_items(limit=100000, due_before=horizon)
    due_rows, gross = due["rows"], due["needed"]
    est_map = due.get("est_for_leavers") or {}
    attr = due.get("attrition") or {}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)

    # Sheet 1 — purchasing plan. Fixed columns, same as on screen:
    # Reissue Need = gross tracking demand; Est. for Leavers = NEGATIVE
    # informational slice; Total = the two combined; Shortfall = Total − Stock.
    ws_plan = wb.create_sheet("Stock Plan")
    ws_plan.append([
        "Uniform Type", "Item", "Stock", "Reissue Need",
        "Est. for Leavers", "Total", "Shortfall", "Warehouse",
    ])
    for r in get_uniform_stock_excel():
        code = r.get("item_code")
        stock = cint(r.get("actual_qty"))
        g = cint(gross.get(code, 0))
        leavers = -cint(est_map.get(code, 0))
        total = g + leavers
        ws_plan.append([
            r.get("template"), r.get("item_name") or code, stock, g, leavers,
            total, max(0, total - stock), r.get("warehouse"),
        ])

    # Sheet 2 — employees due, with reissue cycles in the horizon
    ws_due = wb.create_sheet("Employees Due for Issue")
    ws_due.append([
        "Employee", "Employee Name", "Department", "Section", "Group", "Uniform Type",
        "Size", "Qty/Cycle", "Cycles", "Total Qty", "Last Issue Date", "Next Due Date", "Status",
    ])
    for r in due_rows:
        ws_due.append([
            r.get("employee"), r.get("employee_name"), r.get("department"),
            r.get("custom_section"), r.get("custom_group"), r.get("item_template"),
            r.get("size"), r.get("qty_per_cycle"), r.get("cycles"), r.get("total_qty"),
            str(r.get("last_issue_date") or ""), str(r.get("next_due_date") or ""), r.get("status"),
        ])

    ws_plan.append([])
    if attr.get("enabled"):
        ws_plan.append([
            f"Note: 'Est. for Leavers' (negative) is MEASURED: in the last {attr['months']} "
            f"full months, {attr['persons']} leavers missed re-issues (avg "
            f"{attr['monthly_total']}/month) → × {attr['period_months']} month(s) of this "
            f"horizon, rounded up. Total = Reissue Need + Est.; Shortfall = Total − Stock."
        ])
    else:
        ws_plan.append([
            "Note: Deduct Attrition is OFF in Uniform Setting — Est. for Leavers is 0 "
            "and Total equals Reissue Need."
        ])

    for ws in (ws_plan, ws_due):
        for cell in ws[1]:
            cell.font = bold

    buf = BytesIO()
    wb.save(buf)

    frappe.response["filename"] = f"uniform-dashboard-{today()}.xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"] = "binary"


@frappe.whitelist()
def get_employee_uniform_profile(employee):
    """Return Employee Uniform Profile + issuance history."""
    _check_permission(doctype="Employee Uniform Profile")

    profile_name = frappe.db.get_value(
        "Employee Uniform Profile", {"employee": employee}
    )
    if not profile_name:
        return None

    profile = frappe.get_doc("Employee Uniform Profile", profile_name)

    # Allocation history
    history = frappe.get_all(
        "Uniform Allocation Item",
        filters={"employee": employee, "docstatus": 1},
        fields=["parent", "item_code", "item_name", "qty", "issue_reason"],
        order_by="creation desc",
        limit=50,
    )

    return {
        "profile": profile.as_dict(),
        "history": history,
    }
