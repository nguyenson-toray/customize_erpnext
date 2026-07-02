"""Uniform demand forecast for planned new hires, driven by an HRMS Staffing Plan.

Headcount per designation comes from the Staffing Plan (vacancies / positions).
Staffing Plan has no gender/size, while Uniform Rules + variants do — so gender
split and size mix are inferred from CURRENT employees in the same designation
(fallback: by gender only, then company-wide). Output = demand per item variant
vs current stock. HR can then edit forecast_qty manually."""
import re

import frappe
from frappe import _
from frappe.utils import cint, getdate, today, add_days

from customize_erpnext.uniform_control.utils import (
    get_rules_by_category,
    build_variant_cache,
    get_variant_for_profile,
    get_item_available_qty,
    get_employee_id_prefix,
)

# Uniform Rule category -> (variant_source for get_variant_for_profile, profile size field)
CATEGORY_SIZE = {"Shirt": ("Shirt Size", "shirt_size"), "Shoe": ("Shoe Size", "shoe_size")}


@frappe.whitelist()
def compute(forecast):
    """Fill the forecast's items. Mode decides the source:
      - New Hires: the recruitment plan lines (designation + headcount);
      - Re-issue: upcoming reissue demand up to To Date (multi-cycle);
      - Both: the sum.
    Returns a summary (+ unmapped designations for New Hires)."""
    if not frappe.has_permission("Uniform Demand Forecast", "write"):
        frappe.throw(_("Not permitted"), frappe.PermissionError)

    doc = frappe.get_doc("Uniform Demand Forecast", forecast)
    setting = frappe.get_single("Uniform Setting")
    warehouse = doc.warehouse or setting.uniform_warehouse
    prefix = get_employee_id_prefix()
    mode = doc.mode or "New Hires"

    hire = {}      # variant -> {"people":float,"per":int,"template","size","category"}
    unmapped = []  # (designation, headcount) with no segment data to infer from
    total_hc = 0.0  # headcount spread across segments (New Hires)
    covered = {}    # category -> headcount for which a rule matched
    missing = {}    # (template, size) -> qty with no matching item variant
    seg_records = []  # per segment: which categories it matched (for gap detail)

    # ── New-hire demand from the recruitment plan lines ──
    if mode in ("New Hires", "Both"):
        if not doc.lines:
            frappe.throw(_("Add at least one designation in the Recruitment Plan."))
        seg_cache, size_cache = {}, {}
        for d in doc.lines:
            if not d.designation:
                continue
            headcount = cint(d.headcount)
            if headcount <= 0:
                continue
            # Spread headcount across the (grade, gender, group, section) segments
            # of current employees of this designation so grade-based shirt rules
            # and group/section caps resolve correctly.
            segs = _segments(d.designation, prefix, seg_cache)
            if not segs:
                unmapped.append((d.designation, headcount))
                continue
            for seg, frac in segs:
                count = headcount * frac
                if count <= 0:
                    continue
                total_hc += count
                rules = get_rules_by_category(seg, setting)
                for cat in rules:
                    covered[cat] = covered.get(cat, 0.0) + count
                seg_records.append({"designation": d.designation, "grade": seg.get("grade"),
                                    "count": count, "cats": set(rules)})
                cache = build_variant_cache([r.item for r in rules.values() if r.item])
                for cat, rule in rules.items():
                    # Split PEOPLE by size (each person keeps one size), qty ×
                    # first_qty applied after rounding people → no half-person sizes.
                    _spread(hire, cat, rule, count, d.designation, seg.get("gender"),
                            prefix, cache, size_cache, missing)

    # Round people per template (largest-remainder), then pieces = people × first_qty
    demand = _round_hire_by_template(hire)  # variant -> {"qty":int,"template","size","category"}

    # ── Re-issue demand up to To Date (multi-cycle) — piece-based, already whole ──
    if mode in ("Re-issue", "Both"):
        from customize_erpnext.uniform_control.utils import reissue_demand
        to_date = doc.to_date or add_days(today(), 365)
        rd = reissue_demand(to_date, setting, prefix)
        for variant, qty in rd["needed"].items():
            if qty <= 0:
                continue
            m = rd["meta"].get(variant, {})
            if variant in demand:
                demand[variant]["qty"] += cint(qty)
            else:
                demand[variant] = {"qty": cint(qty), "template": m.get("template") or variant,
                                   "size": m.get("size") or "", "category": m.get("category") or ""}

    doc.set("items", [])
    for variant in sorted(demand):
        info = demand[variant]
        if info["qty"] <= 0:
            continue
        stock = int(get_item_available_qty(variant, warehouse)) if warehouse else 0
        doc.append("items", {
            "item_code": variant,
            "template": info["template"],
            "size": info["size"],
            "category": info["category"],
            "forecast_qty": info["qty"],
            "current_stock": stock,
        })
    doc.save(ignore_permissions=True)  # validate() recomputes totals

    # Partial coverage: a category whose rule matched SOME but not all headcount
    # (e.g. shirts need a Grade, but some staff have none) leaves the rest with
    # nothing. Flag it WITH the who/why breakdown (designation + missing grade).
    from collections import defaultdict
    coverage_gaps = []
    for cat, cov in covered.items():
        miss = total_hc - cov
        if not (cov > 0 and miss >= 0.5):
            continue
        by = defaultdict(float)  # (designation, grade) -> headcount without this cat
        for rec in seg_records:
            if cat not in rec["cats"]:
                by[(rec["designation"], rec["grade"] or "")] += rec["count"]
        detail = [
            {"designation": d, "grade": g, "count": int(round(c))}
            for (d, g), c in sorted(by.items(), key=lambda x: -x[1]) if round(c) >= 1
        ]
        coverage_gaps.append({"category": cat, "uncovered": int(round(miss)), "detail": detail})

    return {
        "items": len(doc.items),
        "total_forecast": doc.total_forecast_qty,
        "total_shortfall": doc.total_shortfall,
        # designations with no current staff to infer from → HR adds manually
        "unmapped": [{"designation": d, "headcount": h} for d, h in unmapped],
        # categories not fully covered by any rule (partial-match headcount)
        "coverage_gaps": coverage_gaps,
        # sizes with no matching item variant (create the variant or add manually)
        "missing_variants": [
            {"template": t, "size": s, "qty": int(round(q))}
            for (t, s), q in sorted(missing.items()) if round(q) >= 1
        ],
    }


@frappe.whitelist()
def current_shirt_ratio(forecast, basis="Company"):
    """Current employees' shirt distribution — the reference ratio for the split.
    Grouped by shirt template (type + gender) and size, as of the forecast's
    creation date. Scope depends on `basis`:
      - "Company": all managed employees (always covers both shirt types);
      - "Recruited": only the designations in the plan's Recruitment lines
        (the true basis of THIS forecast). Falls back to Company if no lines.
    Returns {"as_of": date, "basis": ..., "rows": [{template, size, qty}]}."""
    doc = frappe.get_doc("Uniform Demand Forecast", forecast)
    setting = frappe.get_single("Uniform Setting")
    prefix = get_employee_id_prefix()
    as_of = getdate(doc.creation) if doc.creation else getdate(today())

    designations = [l.designation for l in (doc.lines or []) if l.designation]
    use_recruited = basis == "Recruited" and designations
    if basis == "Recruited" and not designations:
        basis = "Company"  # nothing to scope to → fall back

    params = {"as_of": as_of, "prefix": f"{prefix}%"}
    conds = ["p.shirt_size IS NOT NULL AND p.shirt_size != ''",
             "e.date_of_joining <= %(as_of)s",
             "(e.relieving_date IS NULL OR e.relieving_date > %(as_of)s)"]
    if prefix:
        conds.append("e.name LIKE %(prefix)s")
    if use_recruited:
        conds.append("e.designation IN %(designations)s")
        params["designations"] = designations
    where = " AND ".join(conds)

    rows = frappe.db.sql(
        f"""
        SELECT e.grade AS grade, p.uniform_gender AS gender,
               e.custom_group AS custom_group, e.custom_section AS custom_section,
               p.shirt_size AS size, COUNT(*) c
        FROM `tabEmployee` e
        JOIN `tabEmployee Uniform Profile` p ON p.employee = e.name
        WHERE {where}
        GROUP BY e.grade, p.uniform_gender, e.custom_group, e.custom_section, p.shirt_size
        """,
        params, as_dict=True,
    )

    out, rule_cache = {}, {}
    for r in rows:
        key = (r.grade, r.gender, r.custom_group, r.custom_section)
        if key not in rule_cache:
            seg = {"designation": None, "grade": r.grade, "gender": r.gender,
                   "custom_group": r.custom_group, "custom_section": r.custom_section}
            shirt = get_rules_by_category(seg, setting).get("Shirt")
            rule_cache[key] = shirt.item if shirt else None
        template = rule_cache[key]
        if not template:
            continue
        out[(template, r.size)] = out.get((template, r.size), 0) + r.c

    return {
        "as_of": str(as_of),
        "basis": basis,
        "rows": [{"template": t, "size": s, "qty": q} for (t, s), q in out.items()],
    }


@frappe.whitelist()
def export_forecast_excel(forecast, basis="Company"):
    """Download the whole forecast as .xlsx:
      - Info: header fields (mode, dates, warehouse, totals, notes);
      - Recruitment Plan: the Designations to Recruit lines;
      - Forecast: the computed items;
      - Current Ratio (Recruited) and Current Ratio (Company): both scopes."""
    if not frappe.has_permission("Uniform Demand Forecast", "read"):
        frappe.throw(_("Not permitted"), frappe.PermissionError)

    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Font

    doc = frappe.get_doc("Uniform Demand Forecast", forecast)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    header_rows = []  # (worksheet, row_index_of_header) to bold afterwards

    # 1) Info — the record header
    ws_info = wb.create_sheet("Info")
    for k, v in [
        ("Forecast", doc.name),
        ("Mode", doc.mode),
        ("Company", doc.company),
        ("Warehouse", doc.warehouse),
        ("From Date", str(doc.from_date or "")),
        ("To Date", str(doc.to_date or "")),
        ("Total Forecast Qty", doc.total_forecast_qty),
        ("Total Shortfall vs Stock", doc.total_shortfall),
        ("Notes", doc.notes or ""),
    ]:
        ws_info.append([k, v])
    for row in ws_info.iter_rows(min_col=1, max_col=1):
        row[0].font = bold

    # 2) Recruitment Plan — the Designations to Recruit lines
    ws_plan = wb.create_sheet("Recruitment Plan")
    ws_plan.append(["Designation", "Headcount to Recruit"])
    header_rows.append((ws_plan, 1))
    for l in doc.lines:
        ws_plan.append([l.designation, l.headcount])

    # 3) Forecast — computed items
    ws = wb.create_sheet("Forecast")
    ws.append(["Item", "Uniform Type", "Size", "Category", "Forecast Qty", "Current Stock"])
    header_rows.append((ws, 1))
    for r in doc.items:
        ws.append([r.item_code, r.template, r.size, r.category, r.forecast_qty, r.current_stock])

    # 4) Current Ratio — BOTH scopes
    for scope, title in [("Recruited", "Current Ratio (Recruited)"),
                         ("Company", "Current Ratio (Company)")]:
        ws_r = wb.create_sheet(title)
        ws_r.append(["Type", "Gender", "Size", "Employees"])
        header_rows.append((ws_r, 1))
        for row in current_shirt_ratio(forecast, scope).get("rows", []):
            t = (row["template"] or "").lower()
            typ = "Áo sơ mi" if "sơ mi" in t else ("Áo thun" if "thun" in t else row["template"])
            gender = "Nữ" if "nữ" in t else ("Nam" if "nam" in t else "")
            ws_r.append([typ, gender, row["size"], row["qty"]])

    for ws_x, idx in header_rows:
        for cell in ws_x[idx]:
            cell.font = bold

    buf = BytesIO()
    wb.save(buf)
    frappe.response["filename"] = f"{forecast}-{today()}.xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"] = "binary"


# ─────────────────────────────── helpers ───────────────────────────────────

def _add_hire(hire, variant, template, size, category, people, per):
    e = hire.setdefault(
        variant, {"people": 0.0, "per": per, "template": template, "size": size, "category": category}
    )
    e["people"] += people


def _round_hire_by_template(hire):
    """Round PEOPLE per template with the largest-remainder method (each person
    keeps a single size), then pieces = rounded_people × first_qty. Ensures size
    quantities are whole-person multiples of first_qty (no half-person sizes).
    Returns {variant: {"qty":int,"template","size","category"}}."""
    from collections import defaultdict

    groups = defaultdict(list)  # template -> [(variant, e)]
    for variant, e in hire.items():
        groups[e["template"]].append((variant, e))

    out = {}
    for _tmpl, entries in groups.items():
        target = int(round(sum(e["people"] for _v, e in entries)))
        rows = [[v, e, int(e["people"]), e["people"] - int(e["people"])] for v, e in entries]
        remainder = target - sum(r[2] for r in rows)
        rows.sort(key=lambda r: r[3], reverse=True)
        for i in range(max(0, remainder)):
            rows[i % len(rows)][2] += 1
        for v, e, ppl, _frac in rows:
            out[v] = {"qty": ppl * e["per"], "template": e["template"],
                      "size": e["size"], "category": e["category"]}
    return out


def _spread(hire, cat, rule, count, designation, gender, prefix, cache, size_cache, missing):
    """Spread `count` PEOPLE across item variants by the size mix (each person
    keeps one size). `per` = first_qty is stored for later × after rounding.
    People for a size with no matching variant are recorded in `missing`
    (as pieces) instead of being silently dropped."""
    per = cint(rule.first_qty) or 1
    template = rule.item
    entry = cache.get(template) or {}
    sized = CATEGORY_SIZE.get(cat)

    # Non-sized item (a Cap variant / Bottle) or a template without variants → as-is
    if not (sized and entry.get("has_variants")):
        base = frappe.db.get_value("Item", template, "variant_of") or template
        _add_hire(hire, template, base, _suffix(template, base), cat, count, per)
        return

    source, field = sized
    mix = _size_mix(designation, gender, field, prefix, size_cache)
    if not mix:
        # No size data → single default variant (e.g. shoes 'Free Size')
        default = "Free Size" if field == "shoe_size" else None
        variant = None
        if default:
            variant, _err = get_variant_for_profile(
                template, frappe._dict({field: default}), cache, source
            )
        if variant:
            _add_hire(hire, variant, template, default, cat, count, per)
        else:
            _add_hire(hire, template, template, "", cat, count, per)  # template-level fallback
        return

    for sizeval, frac in mix.items():
        variant, _err = get_variant_for_profile(
            template, frappe._dict({field: sizeval}), cache, source
        )
        if variant:
            _add_hire(hire, variant, template, sizeval, cat, count * frac, per)
        else:
            # No variant for this size → don't lose it silently (record as pieces)
            missing[(template, sizeval)] = missing.get((template, sizeval), 0.0) + count * frac * per


def _suffix(item_code, base):
    """Variant label = item name with the template name stripped off."""
    name = frappe.db.get_value("Item", item_code, "item_name") or item_code
    bname = frappe.db.get_value("Item", base, "item_name") or base
    return name[len(bname):].strip() if name.startswith(bname) else ""


def _ratio_sql(field, where, params):
    rows = frappe.db.sql(
        f"""
        SELECT p.{field} AS k, COUNT(*) c
        FROM `tabEmployee Uniform Profile` p
        JOIN `tabEmployee` e ON e.name = p.employee
        WHERE p.{field} IS NOT NULL AND p.{field} != '' {where}
        GROUP BY p.{field}
        """,
        params, as_dict=True,
    )
    total = sum(r.c for r in rows)
    return {r.k: r.c / total for r in rows} if total else {}


def _segments(designation, prefix, cache):
    """Distribution of (grade, gender, group, section) among current employees of
    this designation → list of (emp_data, fraction). Drives rule matching so a
    planned hire's likely grade/group/section/gender mix is reflected.

    A designation with no current staff (e.g. a brand-new "...-Trainee") falls
    back to the base designation before the first dash ("Sewing Worker-Trainee"
    → "Sewing Worker"). Still empty → caller reports it as unmapped."""
    if designation in cache:
        return cache[designation]
    segs = _segments_query(designation, prefix)
    if not segs:
        base = re.split(r"[-–—]", designation, maxsplit=1)[0].strip()
        if base and base != designation:
            segs = _segments_query(base, prefix)
    cache[designation] = segs
    return segs


def _segments_query(designation, prefix):
    pcond = " AND e.name LIKE %(prefix)s" if prefix else ""
    rows = frappe.db.sql(
        f"""
        SELECT e.grade AS grade, p.uniform_gender AS gender,
               e.custom_group AS custom_group, e.custom_section AS custom_section,
               COUNT(*) c
        FROM `tabEmployee` e
        JOIN `tabEmployee Uniform Profile` p ON p.employee = e.name
        WHERE e.status = 'Active' AND e.designation = %(designation)s {pcond}
        GROUP BY e.grade, p.uniform_gender, e.custom_group, e.custom_section
        """,
        {"designation": designation, "prefix": f"{prefix}%"}, as_dict=True,
    )
    total = sum(r.c for r in rows)
    return [
        ({"designation": designation, "grade": r.grade, "gender": r.gender,
          "custom_group": r.custom_group, "custom_section": r.custom_section}, r.c / total)
        for r in rows
    ] if total else []


def _size_mix(designation, gender, field, prefix, cache):
    key = (designation, gender, field)
    if key in cache:
        return cache[key]
    pcond = " AND e.name LIKE %(prefix)s" if prefix else ""
    base = {"prefix": f"{prefix}%"}
    r = _ratio_sql(
        field,
        "AND e.designation = %(designation)s AND p.uniform_gender = %(gender)s" + pcond,
        {**base, "designation": designation, "gender": gender},
    )
    if not r:  # by gender only
        r = _ratio_sql(field, "AND p.uniform_gender = %(gender)s" + pcond, {**base, "gender": gender})
    if not r:  # company-wide
        r = _ratio_sql(field, pcond, base)
    cache[key] = r
    return r
