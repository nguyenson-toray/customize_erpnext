import frappe
from frappe import _

# Python code (get_role_profile)
@frappe.whitelist()
def get_role_profile(email=None):
    """
    Lấy role_profile_name và roles của user
    Args:
        email: Email của user. Nếu không có sẽ lấy user hiện tại
        
    Returns:
        dict: Thông tin roles của user
    """
    try:
        if not email:
            email = frappe.session.user
             
        # Lấy role_profile và roles
        user = frappe.get_doc('User', email)
        role_profile = user.role_profile_name
        roles = [role.role for role in user.roles]
        # return role_profile
        return {
            'role_profile': role_profile,
            'roles': roles,
            'has_item_manager_role': 'Item Manager' in roles
        }
        
    except Exception as e:
        frappe.log_error(f"Error in get_role_profile: {str(e)}")
        return None
    
# Get color options for the selected item template 

@frappe.whitelist()
def get_colors_for_template(item_template):
    """Get all colors for an item template"""
    colors = frappe.db.sql("""
        SELECT DISTINCT attribute_value 
        FROM `tabItem Variant Attribute` 
        WHERE attribute = 'Color' 
        AND parent IN (
            SELECT name 
            FROM `tabItem` 
            WHERE variant_of = %s
        )
    """, item_template, as_list=1)
    
    return [color[0] for color in colors] if colors else []

def get_item_default_warehouse(item_code):
    """Get default warehouse for an item"""
    try:
        # Get the first default warehouse from item_defaults
        default_warehouse = frappe.db.sql("""
            SELECT default_warehouse
            FROM `tabItem Default`
            WHERE parent = %s
            AND default_warehouse IS NOT NULL
            ORDER BY idx
            LIMIT 1
        """, item_code, as_dict=True)
        
        return default_warehouse[0].default_warehouse if default_warehouse else None
    except Exception:
        return None

@frappe.whitelist()
def export_master_data_item_attribute():
    """Export master data item attributes to Excel file"""
    import io
    import base64
    from datetime import datetime
    
    try:
        # Import openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError:
            frappe.throw(_("openpyxl library is required for Excel export"))
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Item sheet
        item_sheet = wb.create_sheet("Item")
        
        # Item sheet headers
        item_headers = [
            'Item Code', 'Item Name', 'Item Name Detail', 'Item Group', 
            'Default Unit of Measure', 'Color', 'Size', 'Brand', 
            'Season', 'Info', 'Default Warehouse', 'Bin Qty', 'Creation'
        ]
        
        # Add headers to Item sheet
        for col, header in enumerate(item_headers, 1):
            cell = item_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get all items with their attributes
        items_data = frappe.db.sql("""
            SELECT 
                i.item_code,
                i.item_name,
                i.custom_item_name_detail,
                i.item_group,
                i.stock_uom,
                i.creation
            FROM `tabItem` i
            WHERE i.disabled = 0
            ORDER BY i.item_code ASC
        """, as_dict=True)
        
        # Define attribute order
        attribute_order = ['Color', 'Size', 'Brand', 'Season', 'Info']
        
        # Process each item
        for row_num, item in enumerate(items_data, 2):
            # Get item attributes
            item_attributes = frappe.db.sql("""
                SELECT attribute, attribute_value
                FROM `tabItem Variant Attribute`
                WHERE parent = %s
                ORDER BY idx
            """, item.item_code, as_dict=True)
            
            # Create attribute dictionary
            attr_dict = {}
            for attr in item_attributes:
                attr_dict[attr.attribute] = attr.attribute_value
            
            # Get bin quantity
            bin_qty = frappe.db.sql("""
                SELECT SUM(actual_qty) as total_qty
                FROM `tabBin`
                WHERE item_code = %s
            """, item.item_code, as_dict=True)
            
            total_qty = bin_qty[0].total_qty if bin_qty and bin_qty[0].total_qty else 0
            
            # Get default warehouse for this item
            default_warehouse = get_item_default_warehouse(item.item_code)
            
            # Fill item data
            item_sheet.cell(row=row_num, column=1).value = item.item_code
            item_sheet.cell(row=row_num, column=2).value = item.item_name
            item_sheet.cell(row=row_num, column=3).value = item.custom_item_name_detail
            item_sheet.cell(row=row_num, column=4).value = item.item_group
            item_sheet.cell(row=row_num, column=5).value = item.stock_uom
            
            # Fill attributes in order
            for idx, attr_name in enumerate(attribute_order):
                col = 6 + idx  # Start from column 6 (Color)
                item_sheet.cell(row=row_num, column=col).value = attr_dict.get(attr_name, '')
            
            # Fill warehouse and bin qty
            item_sheet.cell(row=row_num, column=11).value = default_warehouse
            item_sheet.cell(row=row_num, column=12).value = total_qty
            
            # Fill creation time
            item_sheet.cell(row=row_num, column=13).value = item.creation
        
        # Format Item sheet as table
        if len(items_data) > 0:
            item_table = Table(
                displayName="ItemTable",
                ref=f"A1:M{len(items_data) + 1}"
            )
            # Add style
            item_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            item_sheet.add_table(item_table)
        
        # Create Attribute sheet
        attr_sheet = wb.create_sheet("Attribute")
        
        # Get all attribute values for each attribute type
        col = 1
        for attr_name in attribute_order:
            # Add attribute name header for abbreviation column
            cell = attr_sheet.cell(row=1, column=col)
            cell.value = f"{attr_name} Abbr"
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
            # Add attribute name header for value column
            cell = attr_sheet.cell(row=1, column=col + 1)
            cell.value = f"{attr_name} Value"
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
            # Get all values for this attribute with abbreviation from Item Attribute Value
            attr_values = frappe.db.sql("""
                SELECT DISTINCT iav.abbr, iav.attribute_value
                FROM `tabItem Attribute Value` iav
                WHERE iav.parent = %s
                ORDER BY iav.abbr ASC
            """, attr_name, as_dict=True)
            
            # Add abbreviations and values in separate columns
            for row_num, value in enumerate(attr_values, 2):
                if value:
                    # Abbreviation column
                    attr_sheet.cell(row=row_num, column=col).value = value.abbr or ''
                    # Value column
                    attr_sheet.cell(row=row_num, column=col + 1).value = value.attribute_value or ''
            
            # Format each attribute pair as table
            if len(attr_values) > 0:
                # Get column letters
                start_col = chr(64 + col)  # Convert to letter (A, B, C, etc.)
                end_col = chr(64 + col + 1)
                
                attr_table = Table(
                    displayName=f"{attr_name}Table",
                    ref=f"{start_col}1:{end_col}{len(attr_values) + 1}"
                )
                # Add style
                attr_table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                attr_sheet.add_table(attr_table)
            
            # Move to next pair of columns (skip one column for spacing)
            col += 3
        
        # Auto-adjust column widths
        for sheet in [item_sheet, attr_sheet]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Convert to base64
        xlsx_data = output.getvalue()
        output.close()
        
        # Create filename with current date
        filename = f"Master_Data_Item_Attribute_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Return file data
        return {
            'file_data': base64.b64encode(xlsx_data).decode(),
            'filename': filename,
            'items_count': len(items_data)
        }
        
    except Exception as e:
        frappe.log_error(f"Error in export_master_data_item_attribute: {str(e)}")
        frappe.throw(_("Error generating Excel file: {0}").format(str(e)))


# ===========================
# FINGERPRINT SYNC FUNCTIONS
# ===========================

@frappe.whitelist()
def get_employee_fingerprint_data(employee_id):
    """Get employee fingerprint data for display"""
    emp = frappe.get_doc("Employee", employee_id)
    return {
        "employee": emp.employee,
        "employee_name": emp.employee_name,
        "attendance_device_id": emp.attendance_device_id or "Not Set"
    }

@frappe.whitelist()
def save_fingerprint_data(employee_id, finger_index, template_data, quality_score=0):
    """Saves fingerprint template data to the Employee document."""
    try:
        employee = frappe.get_doc("Employee", employee_id)
        
        # Assuming a child table with fieldname 'custom_fingerprints' in Employee Doctype
        # And the child Doctype has fields 'finger_index' and 'template_data'
        
        existing_fingerprint = None
        for fp in employee.get("custom_fingerprints"):
            if str(fp.finger_index) == str(finger_index):
                existing_fingerprint = fp
                break
        
        # Get finger name from index
        finger_name = get_finger_name(int(finger_index))
        
        fingerprint_id = None
        if existing_fingerprint:
            existing_fingerprint.template_data = template_data
            existing_fingerprint.finger_name = finger_name
            existing_fingerprint.quality_score = int(quality_score) if quality_score else 0
            fingerprint_id = existing_fingerprint.name
        else:
            new_fingerprint = employee.append("custom_fingerprints", {
                "finger_index": finger_index,
                "finger_name": finger_name,
                "template_data": template_data,
                "quality_score": int(quality_score) if quality_score else 0
            })
            fingerprint_id = new_fingerprint.name

        employee.save(ignore_permissions=True)
        frappe.db.commit()
        
        return {"success": True, "fingerprint_id": fingerprint_id}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Failed to save fingerprint data")
        return {"success": False, "message": str(e)}

@frappe.whitelist()
def update_fingerprint_names(employee_id):
    """Update finger names for existing fingerprints that don't have names set"""
    try:
        employee = frappe.get_doc("Employee", employee_id)
        updated_count = 0
        
        if employee.get("custom_fingerprints"):
            for fp in employee.get("custom_fingerprints"):
                if not fp.finger_name or fp.finger_name.strip() == "":
                    fp.finger_name = get_finger_name(fp.finger_index)
                    updated_count += 1
                    
            if updated_count > 0:
                employee.save(ignore_permissions=True)
                frappe.db.commit()
                
        return {
            "success": True, 
            "updated_count": updated_count,
            "message": f"Updated {updated_count} fingerprint names"
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Failed to update fingerprint names")
        return {"success": False, "message": str(e)}

@frappe.whitelist()
def get_employee_fingerprints_status(employee_id):
    """Returns a list of finger_index for which fingerprint data exists for the given employee."""
    try:
        employee = frappe.get_doc("Employee", employee_id)
        existing_fingers = []
        if employee.get("custom_fingerprints"):
            for fp in employee.get("custom_fingerprints"):
                existing_fingers.append(fp.finger_index)
        return {"success": True, "existing_fingers": existing_fingers}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Failed to get employee fingerprints status")
        return {"success": False, "message": str(e)}

def get_finger_name(finger_index):
    """Get standardized finger name from index"""
    finger_names = {
        0: "Left Little",1: "Left Ring",  2: "Left Middle", 3: "Left Index", 5: "Left Thumb", 
        5: "Right Thumb", 6: "Right Index", 7: "Right Middle", 8: "Right Ring", 9: "Right Little"
    }
    return finger_names.get(finger_index, f"Finger {finger_index}")

def get_finger_index(finger_name):
    """Get finger index from standardized name"""
    finger_map = {
        'Left Little': 0, 'Left Ring': 1,  'Left Middle': 2, 'Left Index': 3, 'Left Thumb': 4,  
        'Right Thumb': 5, 'Right Index': 6, 'Right Middle': 7, 'Right Ring': 8, 'Right Little': 9
    }
    return finger_map.get(finger_name, -1)

@frappe.whitelist()
def get_employees_for_fingerprint():
    """Get employees sorted by employee code (descending) for fingerprint dialog"""
    try:
        employees = frappe.get_list("Employee", 
            fields=["name", "employee", "employee_name", "attendance_device_id"],
            filters={"status": "Active"},
            order_by="employee desc"
        )
        return {"success": True, "employees": employees}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Failed to get employees for fingerprint")
        return {"success": False, "message": str(e)}

def _prepare_employee_sync_data(employee_id):
    """Helper: Prepare employee data for fingerprint sync (DRY principle)

    Now supports syncing employees without fingerprint data - will sync user info only
    """
    # Get employee data
    employee = frappe.get_doc("Employee", employee_id)

    if not employee.attendance_device_id:
        return None, {
            "success": False,
            "message": f"Employee {employee.employee} does not have attendance_device_id"
        }

    # Get employee fingerprints (optional now)
    fingerprints_data = frappe.db.sql("""
        SELECT finger_index, template_data
        FROM `tabFingerprint Data`
        WHERE parent = %s AND template_data IS NOT NULL AND template_data != ''
        ORDER BY finger_index
    """, employee_id, as_dict=True)

    # Prepare fingerprints list (empty if no data)
    fingerprints = [{"finger_index": fp.finger_index, "template_data": fp.template_data} for fp in fingerprints_data] if fingerprints_data else []

    # Prepare privilege
    try:
        from zk import const
        privilege_str = getattr(employee, 'custom_privilege', 'USER_DEFAULT')
        privilege = const.USER_ADMIN if privilege_str == "USER_ADMIN" else const.USER_DEFAULT
    except ImportError:
        privilege = 14 if getattr(employee, 'custom_privilege', 'USER_DEFAULT') == "USER_ADMIN" else 0

    # Prepare password
    password_int = getattr(employee, 'custom_password', None)
    password = str(password_int) if password_int and password_int != 0 else ""

    employee_data = {
        "employee": employee.employee,
        "employee_name": employee.employee_name,
        "attendance_device_id": employee.attendance_device_id,
        "password": password,
        "privilege": privilege,
        "fingerprints": fingerprints
    }

    return employee_data, None


@frappe.whitelist()
def sync_employee_fingerprint_to_machines(employee_id):
    """Sync employee fingerprint data to all enabled attendance machines

    NOTE: This function is kept for backward compatibility.
    New code should use sync_employee_to_single_machine() for better parallelization.
    """
    try:
        # Use helper to prepare employee data (DRY)
        employee_data, error = _prepare_employee_sync_data(employee_id)
        if error:
            return error

        # Get enabled attendance machines
        machines = frappe.get_list("Attendance Machine",
            filters={"enable": 1},
            fields=["name", "device_name", "ip_address", "port", "timeout", "force_udp", "ommit_ping"]
        )

        if not machines:
            return {
                "success": False,
                "message": "No enabled attendance machines found"
            }

        sync_results = []
        success_count = 0

        # Sync to each machine
        for machine in machines:
            try:
                result = sync_to_single_machine(machine, employee_data)
                sync_results.append({
                    "machine": machine.device_name,
                    "ip": machine.ip_address,
                    "success": result["success"],
                    "message": result["message"]
                })

                if result["success"]:
                    success_count += 1

            except Exception as e:
                sync_results.append({
                    "machine": machine.device_name,
                    "ip": machine.ip_address,
                    "success": False,
                    "message": str(e)
                })

        # Create summary message
        total_machines = len(machines)
        if success_count == total_machines:
            status = "success"
            message = f"Successfully synced to all {total_machines} machines"
        elif success_count > 0:
            status = "partial"
            message = f"Synced to {success_count}/{total_machines} machines"
        else:
            status = "failed"
            message = f"Failed to sync to any machines"

        # Get privilege string for summary
        try:
            from zk import const
            privilege_str = "USER_ADMIN" if employee_data["privilege"] == const.USER_ADMIN else "USER_DEFAULT"
        except:
            privilege_str = "USER_ADMIN" if employee_data["privilege"] == 14 else "USER_DEFAULT"

        return {
            "success": success_count > 0,
            "status": status,
            "message": message,
            "sync_results": sync_results,
            "summary": {
                "employee": employee_data["employee"],
                "employee_name": employee_data["employee_name"],
                "attendance_device_id": employee_data["attendance_device_id"],
                "privilege": f"{privilege_str} ({employee_data['privilege']})",
                "password": employee_data["password"] if employee_data["password"] else "No password set",
                "fingerprints_count": len(employee_data["fingerprints"]),
                "machines_total": total_machines,
                "machines_success": success_count,
                "machines_failed": total_machines - success_count
            }
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Failed to sync employee fingerprint to machines")
        return {
            "success": False,
            "message": f"Error: {str(e)}"
        }

@frappe.whitelist()
def sync_employee_to_single_machine(employee_id, machine_name):
    """Sync one employee to one specific machine (for parallel processing)"""
    try:
        # Use helper to prepare employee data (DRY)
        employee_data, error = _prepare_employee_sync_data(employee_id)
        if error:
            return error

        # Get specific machine
        machine = frappe.get_doc("Attendance Machine", machine_name)

        if not machine.enable:
            return {
                "success": False,
                "message": f"Machine {machine.device_name} is disabled"
            }

        # Sync to single machine
        result = sync_to_single_machine(machine, employee_data)

        if result["success"]:
            return {
                "success": True,
                "message": f"Synced {employee_data['employee_name']} to {machine.device_name}"
            }
        else:
            return {
                "success": False,
                "message": result["message"]
            }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), f"sync_employee_to_single_machine: {employee_id} -> {machine_name}")
        return {
            "success": False,
            "message": f"Error: {str(e)}"
        }

def sync_to_single_machine(machine_config, employee_data):
    """Sync employee data to a single attendance machine

    IMPROVEMENTS:
    - Only sends fingers with actual data (bandwidth optimized)
    - Detailed logging for troubleshooting
    - Faster network connectivity check (3s timeout)
    - Proper error handling and messages
    """
    try:
        import socket
        from zk import ZK
        from zk.base import Finger
        import base64
        import time

        # Prepare device config
        device_config = {
            "device_name": machine_config.device_name,
            "ip_address": machine_config.ip_address,
            "port": machine_config.port or 4370,
            "timeout": machine_config.timeout or 10,
            "force_udp": machine_config.force_udp or True,
            "ommit_ping": machine_config.ommit_ping or True
        }

        # Log sync start
        frappe.logger().info(f"🔄 Starting sync: {employee_data['employee']} -> {device_config['device_name']} ({device_config['ip_address']})")

        # Check network connectivity
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(3)
            result = sock.connect_ex((device_config["ip_address"], device_config["port"]))
            sock.close()

            if result != 0:
                error_msg = f"Cannot connect to {device_config['ip_address']}:{device_config['port']}"
                frappe.logger().error(f"❌ Network check failed: {error_msg}")
                return {
                    "success": False,
                    "message": error_msg
                }
        except Exception as e:
            error_msg = f"Network error: {str(e)}"
            frappe.logger().error(f"❌ Network exception: {error_msg}")
            return {
                "success": False,
                "message": error_msg
            }
        
        # Connect to device
        frappe.logger().info(f"🔌 Connecting to device {device_config['device_name']}...")
        zk = ZK(
            device_config["ip_address"],
            port=device_config["port"],
            timeout=device_config["timeout"],
            force_udp=device_config["force_udp"],
            ommit_ping=device_config["ommit_ping"]
        )

        conn = zk.connect()
        if not conn:
            frappe.logger().error(f"❌ Failed to connect to device {device_config['device_name']}")
            return {
                "success": False,
                "message": "Failed to connect to device"
            }

        frappe.logger().info(f"✅ Connected to {device_config['device_name']}")

        try:
            # Disable device during sync
            conn.disable_device()

            attendance_device_id = employee_data["attendance_device_id"]

            # Check if user exists and delete if found
            existing_users = conn.get_users()
            user_exists = any(u.user_id == attendance_device_id for u in existing_users)

            if user_exists:
                frappe.logger().info(f"🗑️  User {attendance_device_id} exists, deleting old data...")
                conn.delete_user(user_id=attendance_device_id)
                time.sleep(0.1)  # Give device time to process

            # Create new user
            full_name = employee_data["employee_name"]
            shortened_name = shorten_name(full_name, 24)
            frappe.logger().info(f"➕ Creating user: {shortened_name} (ID: {attendance_device_id})")
            
            # Set user with correct parameter order and group_id
            password = employee_data.get("password", "")
            privilege = employee_data.get("privilege", 0)  # Default to USER_DEFAULT
            
            if password:
                # Create user with password
                conn.set_user(
                    name=shortened_name,
                    privilege=privilege,
                    password=password,
                    group_id='',
                    user_id=attendance_device_id
                )
            else:
                # Create user without password
                conn.set_user(
                    name=shortened_name,
                    privilege=privilege,
                    group_id='',
                    user_id=attendance_device_id
                )
            
            # Get user info after creation
            users = conn.get_users()
            user = next((u for u in users if u.user_id == attendance_device_id), None)
            
            if not user:
                return {
                    "success": False,
                    "message": f"Could not create or find user {attendance_device_id}"
                }
            
            # ============================================================
            # OPTIMIZED: Only send fingers with actual data
            # This reduces bandwidth and increases sync speed significantly
            # ============================================================

            # Prepare fingerprint templates - only for fingers with data
            templates_to_send = []
            fingerprint_count = 0

            # Create fingerprint data lookup for fast access
            fingerprint_lookup = {fp.get("finger_index"): fp for fp in employee_data["fingerprints"] if fp.get("template_data")}

            # Pre-decode all template data to avoid repeated base64 operations
            decoded_templates = {}
            for finger_index, fp in fingerprint_lookup.items():
                try:
                    decoded_templates[finger_index] = base64.b64decode(fp["template_data"])
                except Exception:
                    pass  # Skip invalid templates

            # IMPROVEMENT: Only create Finger objects for fingers with actual data
            # This is more efficient than sending 10 fingers with empty templates
            for finger_index, template_data in decoded_templates.items():
                finger_obj = Finger(uid=user.uid, fid=finger_index, valid=True, template=template_data)
                templates_to_send.append(finger_obj)
                fingerprint_count += 1

            # Send fingerprint templates if available, otherwise just sync user info
            if templates_to_send:
                # Send only valid templates to device (bandwidth optimized)
                frappe.logger().info(f"📤 Sending {fingerprint_count} fingerprint templates to device...")
                conn.save_user_template(user, templates_to_send)
                frappe.logger().info(f"✅ Successfully synced {fingerprint_count} fingerprints for {employee_data['employee']}")

                return {
                    "success": True,
                    "message": f"Successfully synced {fingerprint_count} fingerprints for user {attendance_device_id}"
                }
            else:
                # No fingerprints, but user info was synced successfully
                frappe.logger().info(f"✅ Successfully synced user info (no fingerprints) for {employee_data['employee']}")

                return {
                    "success": True,
                    "message": f"Successfully synced user info for {attendance_device_id} (no fingerprint data)"
                }

        finally:
            # Re-enable device and disconnect
            try:
                conn.enable_device()
                conn.disconnect()
                frappe.logger().info(f"🔌 Disconnected from {device_config['device_name']}")
            except Exception as e:
                frappe.logger().error(f"⚠️  Error during disconnect: {str(e)}")

    except Exception as e:
        frappe.logger().error(f"❌ Sync error for {employee_data.get('employee', 'Unknown')}: {str(e)}")
        frappe.log_error(frappe.get_traceback(), f"Sync error: {employee_data.get('employee', 'Unknown')}")
        return {
            "success": False,
            "message": f"Sync error: {str(e)}"
        }

@frappe.whitelist()
def get_enabled_attendance_machines():
    """Get list of enabled attendance machines with parallel connection checks"""
    try:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        # Get enabled machines
        # filter : only get machines that are enabled =1
        machines = frappe.get_list("Attendance Machine",
            filters={"enable": 1},
            fields=["name", "device_name", "ip_address", "port", "timeout", "force_udp", "ommit_ping", "location"]
        )

        if not machines:
            return {
                "success": True,
                "machines": [],
                "message": "No enabled attendance machines found"
            }

        machines_with_status = []

        # Parallel connection checking with ThreadPoolExecutor
        # Max workers: 15 concurrent threads (or number of machines if less)
        with ThreadPoolExecutor(max_workers=min(len(machines), 15)) as executor:
            # Submit all machine connection checks simultaneously
            future_to_machine = {
                executor.submit(check_machine_connection_fast, machine): machine
                for machine in machines
            }

            # Collect results as they complete (with 10s total timeout)
            for future in as_completed(future_to_machine, timeout=10):
                machine = future_to_machine[future]
                try:
                    connection_status = future.result()
                    machine_info = {
                        "name": machine.get("name"),
                        "device_name": machine.get("device_name"),
                        "ip_address": machine.get("ip_address"),
                        "port": machine.get("port") or 4370,
                        "location": machine.get("location") or "",
                        "connection_status": connection_status["status"],
                        "connection_message": connection_status["message"],
                        "response_time": connection_status.get("response_time", 0)
                    }
                    machines_with_status.append(machine_info)
                except Exception as e:
                    # Fallback for failed connection checks
                    machines_with_status.append({
                        "name": machine.get("name"),
                        "device_name": machine.get("device_name"),
                        "ip_address": machine.get("ip_address"),
                        "port": machine.get("port") or 4370,
                        "location": machine.get("location") or "",
                        "connection_status": "error",
                        "connection_message": f"Check timeout: {str(e)}",
                        "response_time": 0
                    })

        return {
            "success": True,
            "machines": machines_with_status,
            "total_machines": len(machines_with_status),
            "online_machines": len([m for m in machines_with_status if m["connection_status"] == "online"]),
            "offline_machines": len([m for m in machines_with_status if m["connection_status"] == "offline"])
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Failed to get enabled attendance machines")
        return {
            "success": False,
            "message": f"Error: {str(e)}"
        }

def check_machine_connection_fast(machine_config, timeout=2):
    """Fast connection check with 2s timeout and caching"""
    import socket
    import time

    try:
        # Handle both dict and object access
        if isinstance(machine_config, dict):
            ip_address = machine_config.get('ip_address')
            port = machine_config.get('port') or 4370
        else:
            ip_address = machine_config.ip_address
            port = machine_config.port or 4370

        # Cache key for this machine
        cache_key = f"machine_conn_{ip_address}_{port}"

        # Check cache first (30 second cache)
        try:
            cached_status = frappe.cache().get_value(cache_key)
            if cached_status:
                return cached_status
        except:
            pass  # Cache might fail in thread context

        start_time = time.time()

        # Test TCP connection with 2s timeout
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        result = sock.connect_ex((ip_address, port))
        sock.close()

        response_time = round((time.time() - start_time) * 1000, 2)  # ms

        if result == 0:
            status = {
                "status": "online",
                "message": f"Connected ({response_time}ms)",
                "response_time": response_time
            }
        else:
            status = {
                "status": "offline",
                "message": f"Connection failed (Error: {result})",
                "response_time": 0
            }

        # Cache the result for 30 seconds
        try:
            frappe.cache().set_value(cache_key, status, expires_in_sec=30)
        except:
            pass  # Cache might fail in thread context

        return status

    except Exception as e:
        import traceback
        return {
            "status": "error",
            "message": f"Connection error: {str(e)}",
            "response_time": 0,
            "traceback": traceback.format_exc()
        }

def check_machine_connection(machine_config):
    """Check if attendance machine is reachable (legacy method - calls fast version)"""
    return check_machine_connection_fast(machine_config, timeout=2)

def shorten_name(full_name, max_length=24):
    """Shorten employee name if it exceeds max length and convert Vietnamese to non-accented"""
    from unidecode import unidecode
    if not full_name:
        return full_name 
    text_processed = unidecode(full_name)  # Convert to non-accented text
    # Remove extra spaces
    text_processed = ' '.join(text_processed.split()).strip()
    
    if len(text_processed) > max_length:
        parts = text_processed.split()
        if len(parts) > 1:
            # Take first letter of all parts except the last one
            initials = "".join(part[0].upper() for part in parts[:-1])
            last_part = parts[-1]
            return f"{initials} {last_part}"
        else:
            # If only one word and too long, truncate it
            return text_processed[:max_length]
    else:
        return text_processed


@frappe.whitelist()
def set_default_warehouse_by_brand(template_item, brand_warehouse_map):
    """
    Set default warehouse for item variants based on their Brand attribute
    Only applies if is_customer_provided_item = 1
    If is_customer_provided_item = 0, uses template's default warehouse

    Args:
        template_item: The template item name
        brand_warehouse_map: Dictionary mapping Brand values to warehouse names
    """
    import json

    # Parse the brand_warehouse_map if it's a string
    if isinstance(brand_warehouse_map, str):
        brand_warehouse_map = json.loads(brand_warehouse_map)

    # Get template item's default warehouse
    template_doc = frappe.get_doc("Item", template_item)
    template_warehouse = None
    default_company = frappe.defaults.get_user_default("company")

    for item_default in template_doc.item_defaults:
        if item_default.company == default_company:
            template_warehouse = item_default.default_warehouse
            break

    # Get all variants of the template item
    variants = frappe.get_all(
        "Item",
        filters={"variant_of": template_item},
        fields=["name"]
    )

    if not variants:
        return

    updated_count = 0

    for variant in variants:
        # Get the item document
        item_doc = frappe.get_doc("Item", variant.name)

        # Determine which warehouse to use based on is_customer_provided_item
        if item_doc.is_customer_provided_item == 1:
            # Get the Brand attribute value for this variant
            brand_attr = frappe.db.get_value(
                "Item Variant Attribute",
                {
                    "parent": variant.name,
                    "attribute": "Brand"
                },
                "attribute_value"
            )

            if brand_attr and brand_attr in brand_warehouse_map:
                warehouse = brand_warehouse_map[brand_attr]
            else:
                # If brand not found in map, skip
                continue
        else:
            # Use template's default warehouse
            if template_warehouse:
                warehouse = template_warehouse
            else:
                # If template has no default warehouse, skip
                continue

        # Check if item_defaults already exists
        existing_default = None
        for item_default in item_doc.item_defaults:
            if item_default.company == default_company:
                existing_default = item_default
                break

        if existing_default:
            # Update existing item_default
            existing_default.default_warehouse = warehouse
        else:
            # Add new item_default
            item_doc.append("item_defaults", {
                "company": default_company,
                "default_warehouse": warehouse
            })

        # Save the document
        item_doc.save(ignore_permissions=True)
        updated_count += 1

    if updated_count > 0:
        frappe.db.commit()
        frappe.msgprint(f"Updated default warehouse for {updated_count} variants")

    return updated_count

@frappe.whitelist()
def delete_fingerprint_data(employee_id, finger_index):
    """Delete fingerprint template data for a specific finger"""
    try:
        employee = frappe.get_doc("Employee", employee_id)

        # Find and remove the fingerprint with matching finger_index
        fingerprint_to_remove = None
        for fp in employee.get("custom_fingerprints"):
            if str(fp.finger_index) == str(finger_index):
                fingerprint_to_remove = fp
                break

        if fingerprint_to_remove:
            employee.remove(fingerprint_to_remove)
            employee.save(ignore_permissions=True)
            frappe.db.commit()

            return {
                "success": True,
                "message": f"Deleted fingerprint for finger index {finger_index}"
            }
        else:
            return {
                "success": False,
                "message": f"No fingerprint found for finger index {finger_index}"
            }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Failed to delete fingerprint data")
        return {"success": False, "message": str(e)}
