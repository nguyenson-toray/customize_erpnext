import os
import re
import json
import base64
from io import BytesIO

import frappe
import qrcode
from PIL import Image

LOGO_PATH = os.path.join(
	frappe.get_app_path("customize_erpnext"), "public", "images", "logo_500.jpg"
)


# ── Internal helpers ──────────────────────────────────────────────────────────

def _make_qr_image(content, size, use_logo):
	ec = qrcode.constants.ERROR_CORRECT_H if use_logo else qrcode.constants.ERROR_CORRECT_L
	qr = qrcode.QRCode(version=None, error_correction=ec, box_size=10, border=2)
	qr.add_data(content)
	qr.make(fit=True)
	img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")
	img = img.resize((size, size), Image.LANCZOS)

	if use_logo and os.path.exists(LOGO_PATH):
		logo = Image.open(LOGO_PATH).convert("RGBA")
		logo_max = size // 4
		logo.thumbnail((logo_max, logo_max), Image.LANCZOS)
		lw, lh = logo.size
		pad = 6
		bg = Image.new("RGBA", (lw + pad * 2, lh + pad * 2), "white")
		bg.paste(logo, (pad, pad), logo)
		pos = ((size - bg.width) // 2, (size - bg.height) // 2)
		img.paste(bg, pos, bg)

	buf = BytesIO()
	img.convert("RGB").save(buf, format="PNG")
	return buf.getvalue()


def _make_linear_barcode(barcode_type, content, width, height=None, write_text=True):
	import barcode as bc_lib
	from barcode.writer import ImageWriter

	type_map = {
		"code128": "code128",
		"code39": "code39",
		"ean13": "ean13",
		"ean8": "ean8",
		"upca": "upca",
		"itf": "itf",
		"itf14": "itf",  # python-barcode ITF handles 14 digits
	}
	# Normalize ITF-14 to full 14 digits
	if barcode_type == "itf14":
		content = _itf14_full(content)
	writer = ImageWriter()
	BarcodeClass = bc_lib.get_barcode_class(type_map[barcode_type])
	# Code 39: disable mod-43 checksum to match JsBarcode (manual mode) behaviour
	kwargs = {"add_checksum": False} if barcode_type == "code39" else {}
	bc_obj = BarcodeClass(content, writer=writer, **kwargs)
	buf = BytesIO()
	# Pass original `content` as the `text` arg so python-barcode displays it as-is
	# instead of calling get_fullcode() which appends the internal check character.
	display_text = content
	bc_obj.write(buf, options={
		"module_height": 15.0,
		"text_distance": 4.0,
		"font_size": 10,
		"quiet_zone": 2.5,
		"write_text": write_text,
		"dpi": 150,
	}, text=display_text if write_text else None)
	buf.seek(0)
	src = Image.open(buf)
	w, h = src.size
	new_h = height if height else max(1, int(h * width / w))
	img = src.resize((width, new_h), Image.LANCZOS)
	out = BytesIO()
	img.convert("RGB").save(out, format="PNG")
	return out.getvalue()


def _validate(barcode_type, content):
	"""Return error string or None if valid."""
	if not content:
		return "Content is empty"

	if barcode_type == "qr":
		if len(content.encode("utf-8")) > 2953:
			return "Content too long for QR code (max ~2953 bytes / ~4296 alphanumeric chars)"
		return None

	if barcode_type == "code128":
		if any(ord(c) > 127 for c in content):
			return "Code 128 only supports ASCII characters (0–127)"
		return None

	if barcode_type == "code39":
		allowed = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-. $/+%")
		invalid = set(content.upper()) - allowed
		if invalid:
			return f"Code 39: invalid chars {''.join(sorted(invalid))}. Allowed: A-Z, 0-9, - . $ / + % space"
		return None

	digit_types = {"ean13", "ean8", "upca", "itf"}
	if barcode_type in digit_types:
		if not re.match(r"^\d+$", content):
			return f"{barcode_type.upper()} requires digits only (0–9)"
		required_len = {"ean13": 12, "ean8": 7, "upca": 11}
		if barcode_type in required_len:
			n = required_len[barcode_type]
			if len(content) != n:
				return f"{barcode_type.upper()} requires exactly {n} digits (checksum auto-added), got {len(content)}"
		if barcode_type == "itf" and len(content) % 2 != 0:
			return f"ITF requires an even number of digits, got {len(content)} (pad with a leading 0)"
		return None

	if barcode_type == "itf14":
		if not re.match(r"^\d+$", content):
			return "ITF-14 requires digits only (0–9)"
		if len(content) not in (13, 14):
			return f"ITF-14 requires exactly 13 digits (14th is checksum), got {len(content)}"
		return None

	return None


def _itf14_full(digits):
	"""Return 14-digit ITF-14 string, calculating check digit if only 13 given."""
	if len(digits) == 14:
		return digits
	weights = [3, 1] * 7  # alternating 3,1 for positions 1-13
	total = sum(int(d) * w for d, w in zip(digits, weights))
	check = (10 - (total % 10)) % 10
	return digits + str(check)


# ── Public API ────────────────────────────────────────────────────────────────

@frappe.whitelist()
def generate_qr(content, size=500, use_logo=1):
	"""Legacy single-QR endpoint — kept for backward compatibility."""
	size = max(100, min(2000, int(size or 500)))
	use_logo = int(use_logo)
	return base64.b64encode(_make_qr_image(content, size, use_logo)).decode()


@frappe.whitelist()
def generate_barcode_excel(barcode_type, file_b64, has_header=1, img_width=200, img_height=None, use_logo=1, show_text=1):
	"""
	Read column A from an uploaded Excel file, generate a barcode for each row,
	and embed the barcode image into column B.  Returns a base64-encoded .xlsx.

	Args:
		barcode_type : same codes as generate_barcode
		file_b64     : base64-encoded .xlsx/.xls file content
		has_header   : 1 → first row is a header (skip for generation, add "Barcode" label)
		img_width    : barcode image display width (px) inside the Excel cell
		use_logo     : embed logo in QR codes (1/0)
	Returns:
		{ file_b64: <base64 xlsx>, total: N, errors: M }
	"""
	import openpyxl
	from openpyxl.drawing.image import Image as XLImage
	from openpyxl.styles import Font, Alignment, PatternFill

	has_header = int(has_header)
	img_width  = max(60, min(600, int(img_width or 200)))
	img_height = max(20, min(1000, int(img_height))) if img_height else None
	use_logo   = int(use_logo)
	show_text  = int(show_text)

	# Decode uploaded file
	excel_bytes = base64.b64decode(file_b64)
	wb = openpyxl.load_workbook(BytesIO(excel_bytes))
	ws = wb.active

	start_row = 2 if has_header else 1

	# Add header for column B
	if has_header:
		cell = ws.cell(row=1, column=2, value="Barcode")
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center")
		# Match style of column A header (copy fill if any)
		a1_fill = ws.cell(row=1, column=1).fill
		if a1_fill and a1_fill.fill_type not in (None, "none"):
			cell.fill = a1_fill

	# Determine Excel display dimensions per barcode type
	is_qr = barcode_type == "qr"
	if is_qr:
		xl_w = img_width
		xl_h = img_width          # QR is square
	else:
		xl_w = img_width
		# height will be calculated proportionally after generating first image

	# Column B width  (openpyxl unit ≈ ~7 px per char)
	ws.column_dimensions["B"].width = round(xl_w / 7) + 2

	total  = 0
	errors = 0

	for row_idx in range(start_row, ws.max_row + 1):
		raw = ws.cell(row=row_idx, column=1).value
		if raw is None:
			continue
		# openpyxl returns numeric cells as float (e.g. 590123412345.0)
		# Convert to int first to avoid "590123412345.0" being encoded into barcode
		if isinstance(raw, float) and raw == int(raw):
			item = str(int(raw))
		else:
			item = str(raw).strip()
		if not item:
			continue

		# Normalize
		if barcode_type == "code39":
			item = item.upper()

		err = _validate(barcode_type, item)
		if err:
			ws.cell(row=row_idx, column=2).value = f"[Error] {err}"
			ws.cell(row=row_idx, column=2).font = Font(color="FF0000", italic=True)
			errors += 1
			continue

		try:
			# Generate at 2× for quality, then display at img_width
			gen_size = max(400, img_width * 2)
			if is_qr:
				img_bytes = _make_qr_image(item, gen_size, use_logo)
				xl_h_row  = img_height if img_height else img_width
			else:
				gen_height = max(40, img_height * 2) if img_height else None
				img_bytes = _make_linear_barcode(barcode_type, item, gen_size, height=gen_height, write_text=bool(show_text))
				src       = Image.open(BytesIO(img_bytes))
				xl_h_row  = img_height if img_height else max(30, round(src.height * img_width / src.width))

			xl_img        = XLImage(BytesIO(img_bytes))
			xl_img.width  = img_width
			xl_img.height = xl_h_row
			xl_img.anchor = f"B{row_idx}"
			ws.add_image(xl_img)

			# Row height: convert px → points (1 pt ≈ 1.333 px), add 6 px padding
			ws.row_dimensions[row_idx].height = round((xl_h_row + 6) / 1.333)
			total += 1

		except Exception as e:
			ws.cell(row=row_idx, column=2).value = f"[Error] {e}"
			ws.cell(row=row_idx, column=2).font = Font(color="FF0000", italic=True)
			errors += 1

	out = BytesIO()
	wb.save(out)
	return {
		"file_b64": base64.b64encode(out.getvalue()).decode(),
		"total":  total,
		"errors": errors,
	}


@frappe.whitelist()
def generate_barcode(barcode_type, contents, size=500, height=None, use_logo=1):
	"""
	Generate one or many barcodes in a single call.

	Args:
		barcode_type: 'qr' | 'code128' | 'code39' | 'ean13' | 'ean8' | 'upca' | 'itf'
		contents    : JSON-encoded list of strings (one per barcode)
		size        : output width in px
		height      : output height in px (optional; proportional if omitted; ignored for QR)
		use_logo    : 1/0 — embed logo in QR codes only

	Returns:
		list of { content, b64, error } dicts
	"""
	size     = max(100, min(2000, int(size or 500)))
	height   = max(20, min(1000, int(height))) if height else None
	use_logo = int(use_logo)

	try:
		items = json.loads(contents) if isinstance(contents, str) else list(contents)
	except Exception:
		items = [str(contents)]

	# Code39 is uppercase-only — normalize silently
	if barcode_type == "code39":
		items = [str(i).upper() for i in items]

	results = []
	for item in items:
		item = str(item).strip()
		err = _validate(barcode_type, item)
		if err:
			results.append({"content": item, "b64": None, "error": err})
			continue
		try:
			if barcode_type == "qr":
				data = _make_qr_image(item, size, use_logo)
			else:
				data = _make_linear_barcode(barcode_type, item, size, height=height)
			results.append({"content": item, "b64": base64.b64encode(data).decode(), "error": None})
		except Exception as e:
			results.append({"content": item, "b64": None, "error": str(e)})

	return results
