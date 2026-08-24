"""
Employee Self Update Info API

A dynamic, field-picker driven self-service page. HR chooses any Employee field
(including custom fields) in `Employee Self Update Info Setting`; employees review
the current value of each field and update what changed. Submissions are stored as
JSON on `Employee Self Update Info` and exported to Excel — they are NOT synced back
to the Employee record.

Public (allow_guest=True):
    get_field_config        -> field list built from settings + Employee meta
    get_eligible_employees  -> employees configured in the setting
    get_form_data           -> current Employee values overlaid with any draft
    save_form_data          -> store submitted values as JSON

HR (login required):
    download_excel          -> xlsx with old (Employee) vs new (submitted) columns
"""

import json

import frappe
from frappe import _
from frappe.utils import now_datetime

SETTING_DT = "Employee Self Update Info Setting"
INFO_DT = "Employee Self Update Info"
# Reserved key inside data_json for the employee's free-text remarks.
REMARKS_KEY = "__remarks"


def _submit_device_info():
	"""Basic device/browser info of the submitting request: IP + User-Agent.

	Read-only audit only; no permission prompt, not synced to Employee.
	Returns '' when there is no HTTP request context (e.g. bench execute).
	"""
	try:
		ip = frappe.local.request_ip or ""
	except Exception:
		ip = ""
	def _h(name):
		try:
			return (frappe.get_request_header(name) or "").strip().strip('"')
		except Exception:
			return ""

	# User-Agent Client Hints (Chromium/Android; empty on Safari/iOS). Requires
	# the Accept-CH opt-in sent by the web page (see www/.../index.py).
	model = _h("Sec-CH-UA-Model")
	platform = " ".join(x for x in (_h("Sec-CH-UA-Platform"), _h("Sec-CH-UA-Platform-Version")) if x)

	parts = []
	if ip:
		parts.append(f"IP: {ip}")
	if model:
		parts.append(f"Model: {model}")
	if platform:
		parts.append(f"Platform: {platform}")
	return "\n".join(parts)

# Fieldtypes the dynamic renderer knows how to handle (v1: flat + Link + Text).
_ALLOWED_FIELDTYPES = {
	"Data", "Date", "Datetime", "Time", "Int", "Float", "Currency",
	"Select", "Check", "Small Text", "Text", "Long Text", "Link", "Phone",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _require_hr():
	if frappe.session.user == "Guest":
		frappe.throw(_("Authentication required"), frappe.AuthenticationError)
	roles = frappe.get_roles()
	if not any(r in roles for r in ("HR Manager", "HR User", "System Manager")):
		frappe.throw(_("Not permitted"), frappe.PermissionError)


def _get_setting():
	return frappe.get_cached_doc(SETTING_DT)


def _selected_rows(setting):
	# Only rows explicitly enabled (enable=1) are exposed on the public page.
	# This single chokepoint governs rendering, save validation and export.
	return [r for r in (setting.selected_fields or []) if r.employee_fieldname and r.enable]


def _build_config():
	"""Resolve the configured fields against the live Employee meta."""
	setting = _get_setting()
	rows = _selected_rows(setting)
	meta = frappe.get_meta("Employee")
	df_by_name = {df.fieldname: df for df in meta.fields}

	sections = {}
	order = []
	for row in rows:
		# Skip fields explicitly disabled via the "Enable" toggle.
		if not row.enable:
			continue
		if row.is_custom:
			# Free-form field that does NOT exist on Employee — defined entirely
			# by the row. Stored in the submission only.
			fieldtype = row.custom_fieldtype or "Data"
			if fieldtype not in _ALLOWED_FIELDTYPES:
				continue
			options = None
			if fieldtype == "Select" and row.custom_options:
				options = [o.strip() for o in row.custom_options.split("\n")]
			field = {
				"fieldname": row.employee_fieldname,
				"label": row.label_vi or row.employee_fieldname,
				"employee_label": row.label_vi or row.employee_fieldname,
				"detail": row.detail or "",
				"placeholder": row.placeholder or "",
				"fieldtype": fieldtype,
				"options": options,
				"required": bool(row.required),
				"read_only": bool(row.read_only),
				"widget": "Auto",
				"custom": True,
				"auto_fill": bool(row.auto_fill_data),
				**_validation_meta(row),
			}
		else:
			df = df_by_name.get(row.employee_fieldname)
			if not df or df.fieldtype not in _ALLOWED_FIELDTYPES:
				# Field removed/renamed on Employee, or type unsupported — skip.
				continue

			options = None
			if df.fieldtype == "Select" and df.options:
				options = [o for o in df.options.split("\n")]
			elif df.fieldtype == "Link":
				options = df.options  # the linked doctype name

			field = {
				"fieldname": df.fieldname,
				# UI label: prefer the Vietnamese label, fall back to the Employee
				# field's default label.
				"label": row.label_vi or df.label or df.fieldname,
				# Excel/import label: always the Employee field's own label so the
				# exported file can be re-imported into Employee (Data Import).
				"employee_label": df.label or df.fieldname,
				# Optional guidance shown under the label on the page.
				"detail": row.detail or "",
				# Optional hint text inside the empty input.
				"placeholder": row.placeholder or "",
				"fieldtype": df.fieldtype,
				"options": options,
				"required": bool(row.required),
				"read_only": bool(row.read_only),
				"widget": row.widget or "Auto",
				"custom": False,
				"auto_fill": bool(row.auto_fill_data),
				**_validation_meta(row),
			}

		sec = row.section_label or "General"
		if sec not in sections:
			sections[sec] = {"label": sec, "fields": []}
			order.append(sec)
		sections[sec]["fields"].append(field)

	return {"sections": [sections[s] for s in order]}


def _validation_meta(row):
	"""Validation attributes copied from a config row into the field dict."""
	return {
		"validation": row.validation or "",
		"min_length": int(row.min_length or 0),
		"max_length": int(row.max_length or 0),
		"regex": row.regex or "",
	}


# Built-in patterns for the preset validation types.
_VALIDATION_PATTERNS = {
	"Digits": (r"^\d+$", "chỉ gồm chữ số"),
	"Phone": (r"^0\d{9}$", "số điện thoại VN 10 số, bắt đầu bằng 0"),
	"Email": (r"^[^@\s]+@[^@\s]+\.[^@\s]+$", "email hợp lệ"),
	"CCCD": (r"^\d{12}$", "đúng 12 chữ số"),
	"CMND": (r"^\d{9}$", "đúng 9 chữ số"),
}


def _validate_value(field, value):
	"""Return an error message (str) if `value` fails the field's validation,
	else None. Empty values pass here (handled by the required check)."""
	import re

	val = "" if value is None else str(value).strip()
	if val == "":
		return None

	label = field.get("label") or field.get("fieldname")
	min_len = field.get("min_length") or 0
	max_len = field.get("max_length") or 0
	if min_len and len(val) < min_len:
		return _("{0}: tối thiểu {1} ký tự").format(label, min_len)
	if max_len and len(val) > max_len:
		return _("{0}: tối đa {1} ký tự").format(label, max_len)

	vtype = field.get("validation") or ""
	if not vtype:
		return None

	if vtype in ("Past", "Future"):
		try:
			dv = frappe.utils.getdate(val)
		except Exception:
			return None  # not a parseable date → don't block
		tv = frappe.utils.getdate(frappe.utils.today())
		if vtype == "Past" and dv > tv:
			return _("{0}: không được ở tương lai").format(label)
		if vtype == "Future" and dv < tv:
			return _("{0}: không được ở quá khứ").format(label)
		return None

	if vtype == "Regex":
		pattern = field.get("regex") or ""
		if not pattern:
			return None
		try:
			ok = re.match(pattern, val) is not None
		except re.error:
			return None  # invalid config regex → don't block the employee
		return None if ok else _("{0}: không đúng định dạng").format(label)

	spec = _VALIDATION_PATTERNS.get(vtype)
	if not spec:
		return None
	pattern, desc = spec
	if re.match(pattern, val) is None:
		return _("{0}: phải là {1}").format(label, desc)
	return None


def _config_fieldnames(config):
	return [f["fieldname"] for sec in config["sections"] for f in sec["fields"]]


def _eligible_ids(setting):
	ids = [r.employee for r in (setting.employees or []) if r.employee]
	return ids


def _dob_day(employee_id):
	"""Day-of-month of the employee's date of birth, zero-padded, e.g.
	1984-09-01 -> '01'."""
	dob = frappe.db.get_value("Employee", employee_id, "date_of_birth")
	if not dob:
		return None
	return str(dob)[-2:]  # 'YYYY-MM-DD' -> last 2 chars = 'DD'


def _num_eq(a, b):
	"""Compare two short numeric codes, tolerant of leading zeros."""
	try:
		return int(a) == int(b)
	except (TypeError, ValueError):
		return str(a).strip() == str(b).strip()


def _code_ok(setting, employee_id, code):
	"""True if the supplied code matches the DOB day or the bypass code."""
	code = (str(code or "")).strip()
	if not code:
		return False
	bypass = setting.bypass_code
	if bypass and _num_eq(code, bypass):
		return True
	day = _dob_day(employee_id)
	return bool(day) and _num_eq(code, day)


def _is_hr():
	"""True if the current session is a logged-in HR / System Manager."""
	if frappe.session.user in ("Guest", None):
		return False
	return any(r in frappe.get_roles() for r in ("HR Manager", "HR User", "System Manager"))


def _ensure_eligible(setting, employee_id):
	"""HR may edit ANY employee; employees themselves only those in the setting."""
	if _is_hr():
		if not frappe.db.exists("Employee", employee_id):
			frappe.throw(_("Employee not found"))
		return
	if employee_id not in _eligible_ids(setting):
		frappe.throw(_("Employee not eligible for self update"), frappe.PermissionError)


def _gate(setting, employee_id, code):
	"""Throw unless verification passes. When validate_by_dob is on, EVERYONE
	(including HR) must verify — HR uses the admin bypass_code to edit any
	employee. No-op only when validate_by_dob is off."""
	if not setting.validate_by_dob:
		return
	if not _code_ok(setting, employee_id, code):
		frappe.throw(_("Verification failed. Please check the code."), frappe.ValidationError)


def _already_submitted(employee_id):
	"""True once the employee has submitted the form at least once."""
	return bool(frappe.db.get_value(INFO_DT, employee_id, "submitted_on"))


def _unlock_ok(setting, unlock_code):
	"""True if the supplied unlock code matches bypass_code_for_unlock."""
	unlock_code = (str(unlock_code or "")).strip()
	if not unlock_code:
		return False
	return _num_eq(unlock_code, setting.bypass_code_for_unlock)


def _is_locked(setting, employee_id):
	"""True if the form is locked for this employee (submitted once + lock on)."""
	return bool(setting.get("lock_after_submit")) and _already_submitted(employee_id)


def _gate_edit(setting, employee_id, code, unlock_code=None):
	"""Gate for loading/saving the form. DOB check (as _gate) plus, when the
	form is locked after a first submission, BOTH the DOB digits AND the unlock
	code are required (employee must contact HR for the unlock code)."""
	_gate(setting, employee_id, code)
	if _is_locked(setting, employee_id):
		if not (_code_ok(setting, employee_id, code) and _unlock_ok(setting, unlock_code)):
			frappe.throw(
				_("This form is locked. Enter your date-of-birth digits and the unlock code from HR."),
				frappe.PermissionError,
			)


# --- Short-lived download token: lets the just-submitted session download the
# receipt (PDF/PNG) without the unlock code, while a cold visit stays locked. ---

def _dl_secret():
	return (frappe.local.conf.get("encryption_key") or frappe.local.conf.get("secret") or "csi").encode()


def _dl_windows():
	"""Current and previous 15-minute windows (token validity ~15–30 min)."""
	w = int(now_datetime().timestamp() // 900)
	return (w, w - 1)


def _make_download_token(employee_id):
	import hashlib
	import hmac

	msg = f"{employee_id}:{_dl_windows()[0]}".encode()
	return hmac.new(_dl_secret(), msg, hashlib.sha256).hexdigest()[:32]


def _download_token_ok(employee_id, token):
	import hashlib
	import hmac

	token = (str(token or "")).strip()
	if not token:
		return False
	for w in _dl_windows():
		good = hmac.new(_dl_secret(), f"{employee_id}:{w}".encode(), hashlib.sha256).hexdigest()[:32]
		if hmac.compare_digest(good, token):
			return True
	return False


def _gate_receipt(setting, employee_id, code, unlock_code=None, token=None):
	"""Gate for the PDF/PNG receipt. A valid fresh-submit token allows the
	download (same session that just submitted); otherwise the full edit gate
	applies so a locked form does not leak personal data on a later visit."""
	if _download_token_ok(employee_id, token):
		return
	_gate_edit(setting, employee_id, code, unlock_code)


# ---------------------------------------------------------------------------
# Public APIs
# ---------------------------------------------------------------------------

@frappe.whitelist(allow_guest=True)
def get_field_config():
	"""Return the dynamic field configuration for the web form."""
	setting = _get_setting()
	config = _build_config()
	config["require_dob"] = bool(setting.validate_by_dob)
	# Whether the page may keep a local draft (localStorage).
	config["allow_local_draft"] = bool(setting.allow_browser_local_storage)
	# In Zalo's in-app browser, force the user to open in the device browser
	# (PDF download is blocked inside the Zalo webview).
	config["force_device_browser"] = bool(setting.get("force_open_devices_browser"))
	# Receipt file type offered on the success screen: "PDF" (default) or "PNG".
	# PNG renders inline so it can be long-pressed to save inside the Zalo webview.
	config["save_file_type"] = setting.get("save_file_type") or "PDF"
	# Lock the form (and receipts) after the first submission. The secret
	# unlock code (bypass_code_for_unlock) is NEVER sent to the client.
	config["lock_after_submit"] = bool(setting.get("lock_after_submit"))
	return config


@frappe.whitelist(allow_guest=True)
def verify_employee(employee_id, code):
	"""Verify the DOB digits / bypass code before showing the form.

	Returns {valid: bool}. Always valid when validate_by_dob is off.
	"""
	if not employee_id:
		frappe.throw(_("Missing employee"))
	setting = _get_setting()
	_ensure_eligible(setting, employee_id)
	if not setting.validate_by_dob:
		return {"valid": True}
	return {"valid": _code_ok(setting, employee_id, code)}


@frappe.whitelist(allow_guest=True)
def get_access_info(employee_id):
	"""Report whether this employee's form is locked after a first submission.

	Returns {submitted, submitted_on (display), locked}. The unlock code is
	never exposed. Low sensitivity: same as the `submitted` flag already
	returned by get_eligible_employees.
	"""
	if not employee_id:
		frappe.throw(_("Missing employee"))
	setting = _get_setting()
	_ensure_eligible(setting, employee_id)
	submitted_on = frappe.db.get_value(INFO_DT, employee_id, "submitted_on")
	return {
		"submitted": bool(submitted_on),
		"submitted_on": frappe.utils.format_datetime(submitted_on, "dd/MM/yyyy HH:mm") if submitted_on else "",
		"locked": bool(setting.get("lock_after_submit")) and bool(submitted_on),
	}


@frappe.whitelist(allow_guest=True)
def unlock_access(employee_id, code, unlock_code):
	"""Pre-check the unlock: needs BOTH the DOB digits and the unlock code.

	Returns {valid}. Server-side enforcement still happens in get_form_data /
	save_form_data, so bypassing this cannot unlock the form.
	"""
	if not employee_id:
		frappe.throw(_("Missing employee"))
	setting = _get_setting()
	_ensure_eligible(setting, employee_id)
	valid = _code_ok(setting, employee_id, code) and _unlock_ok(setting, unlock_code)
	return {"valid": valid}


@frappe.whitelist(allow_guest=True)
def get_eligible_employees():
	"""Return employees configured in the setting (optionally filtered by group)."""
	setting = _get_setting()
	ids = _eligible_ids(setting)
	if not ids:
		return []

	rows = frappe.get_all(
		"Employee",
		filters={"name": ["in", ids]},
		fields=["name as employee_id", "employee_name as display_name"],
		order_by="employee_name asc",
	)
	submitted = set(
		frappe.get_all(
			INFO_DT,
			filters={"employee": ["in", ids], "status": "Submitted"},
			pluck="employee",
		)
	)
	for r in rows:
		r["submitted"] = r["employee_id"] in submitted
	return rows


@frappe.whitelist(allow_guest=True)
def get_form_data(employee_id, code=None, unlock_code=None):
	"""Return current Employee values overlaid with any saved draft/submission.

	`code` = DOB digits / bypass code, required only when validate_by_dob is on.
	`unlock_code` = required together with `code` when the form is locked after
	a first submission (lock_after_submit).

	Returns:
	    {
	      original: {fieldname: value},   # live Employee values
	      values:   {fieldname: value},   # what to show (draft overrides original)
	      has_existing: bool,
	      status: "Draft"|"Submitted"|None,
	      employee_name: str,
	    }
	"""
	if not employee_id:
		frappe.throw(_("Missing employee"))

	setting = _get_setting()
	_ensure_eligible(setting, employee_id)
	_gate_edit(setting, employee_id, code, unlock_code)

	config = _build_config()
	fieldnames = _config_fieldnames(config)
	# Only real Employee fields can be read from the Employee record; custom
	# fields have no source value.
	real_fields = [
		f["fieldname"]
		for sec in config["sections"]
		for f in sec["fields"]
		if not f.get("custom")
	]

	emp = frappe.db.get_value(
		"Employee", employee_id, ["employee_name"] + real_fields, as_dict=True
	) or {}
	employee_name = emp.get("employee_name")
	original = {fn: (emp.get(fn) if fn in real_fields else None) for fn in fieldnames}

	values = dict(original)
	status = None
	has_existing = False
	remarks = ""
	if frappe.db.exists(INFO_DT, employee_id):
		doc = frappe.get_doc(INFO_DT, employee_id)
		status = doc.status
		has_existing = True
		saved = json.loads(doc.data_json or "{}")
		remarks = saved.get(REMARKS_KEY, "")
		for fn in fieldnames:
			if fn in saved:
				values[fn] = saved[fn]

	# Fields with auto_fill = 0 must be re-entered every visit: never prefill
	# them from Employee or from a previous submission.
	no_fill = {
		f["fieldname"]
		for sec in config["sections"]
		for f in sec["fields"]
		if not f.get("auto_fill")
	}
	for fn in no_fill:
		values[fn] = None

	return {
		"original": original,
		"values": values,
		"has_existing": has_existing,
		"status": status,
		"employee_name": employee_name,
		"remarks": remarks,
	}


@frappe.whitelist(allow_guest=True)
def save_form_data(employee_id, data, code=None, unlock_code=None):
	"""Store submitted values as JSON. Does NOT write back to Employee."""
	if not employee_id:
		frappe.throw(_("Missing employee"))

	setting = _get_setting()
	_ensure_eligible(setting, employee_id)
	_gate_edit(setting, employee_id, code, unlock_code)

	if isinstance(data, str):
		data = json.loads(data or "{}")

	config = _build_config()
	allowed = set(_config_fieldnames(config))

	# Keep only configured, editable fields.
	editable = {
		f["fieldname"]
		for sec in config["sections"]
		for f in sec["fields"]
		if not f["read_only"]
	}
	clean = {k: v for k, v in (data or {}).items() if k in allowed and k in editable}

	# Free-text remarks (always allowed, stored under a reserved key).
	remarks = (data or {}).get(REMARKS_KEY)
	if remarks not in (None, ""):
		clean[REMARKS_KEY] = str(remarks).strip()

	# Validate required fields.
	missing = []
	for sec in config["sections"]:
		for f in sec["fields"]:
			if f["required"] and not f["read_only"]:
				val = clean.get(f["fieldname"])
				if val is None or str(val).strip() == "":
					missing.append(f["label"])
	if missing:
		frappe.throw(_("Please fill required fields: {0}").format(", ".join(missing)))

	# Format validation (mirrors the client; server is the source of truth).
	errors = []
	for sec in config["sections"]:
		for f in sec["fields"]:
			if f["read_only"]:
				continue
			err = _validate_value(f, clean.get(f["fieldname"]))
			if err:
				errors.append(err)
	if errors:
		frappe.throw("<br>".join(errors))

	employee_name = frappe.db.get_value("Employee", employee_id, "employee_name")

	if frappe.db.exists(INFO_DT, employee_id):
		doc = frappe.get_doc(INFO_DT, employee_id)
	else:
		doc = frappe.new_doc(INFO_DT)
		doc.employee = employee_id

	doc.employee_name = employee_name
	doc.data_json = json.dumps(clean, ensure_ascii=False)
	doc.status = "Submitted"
	doc.submitted_on = now_datetime()
	# Append device info each submit (keep history) instead of replacing.
	info = _submit_device_info()
	if info:
		stamp = frappe.utils.format_datetime(now_datetime(), "yyyy-MM-dd HH:mm")
		entry = f"[{stamp}]\n{info}"
		prev = (doc.device_info or "").strip()
		doc.device_info = (prev + "\n\n" + entry) if prev else entry
	doc.save(ignore_permissions=True)
	frappe.db.commit()

	# Short-lived token so THIS session can download the receipt right away,
	# even though the form is now locked for later (cold) visits.
	return {
		"status": "success",
		"message": _("Your information has been submitted."),
		"download_token": _make_download_token(employee_id),
	}


@frappe.whitelist(allow_guest=True)
def download_submission_pdf(employee_id, code=None, unlock_code=None, token=None):
	"""Return a PDF receipt of the employee's submitted information."""
	html, base = _load_submission_receipt(employee_id, code, unlock_code, token)
	from frappe.utils.pdf import get_pdf

	frappe.response["filename"] = f"{base}.pdf"
	frappe.response["filecontent"] = get_pdf(html)
	frappe.response["type"] = "pdf"


@frappe.whitelist(allow_guest=True)
def download_submission_image(employee_id, code=None, unlock_code=None, token=None):
	"""Return a PNG image of the employee's submitted information as a file download."""
	html, base = _load_submission_receipt(employee_id, code, unlock_code, token)
	frappe.response["filename"] = f"{base}.png"
	frappe.response["filecontent"] = _html_to_png(html)
	frappe.response["type"] = "download"
	frappe.response["content_type"] = "image/png"


def _load_submission_receipt(employee_id, code, unlock_code=None, token=None):
	"""Shared gate + HTML build for the PDF/PNG receipt. Returns (html, base_filename)."""
	if not employee_id:
		frappe.throw(_("Missing employee"))
	setting = _get_setting()
	_ensure_eligible(setting, employee_id)
	_gate_receipt(setting, employee_id, code, unlock_code, token)

	if not frappe.db.exists(INFO_DT, employee_id):
		frappe.throw(_("No submission found for this employee."))

	doc = frappe.get_doc(INFO_DT, employee_id)
	saved = json.loads(doc.data_json or "{}")
	config = _build_config()
	html = _build_submission_html(doc, saved, config)

	stamp = frappe.utils.format_datetime(now_datetime(), "yyyyMMdd_HHmm")
	name_part = (doc.employee_name or "").strip()
	base = f"{employee_id} {name_part} {stamp}".strip()
	return html, base


def _html_to_png(html):
	"""Render HTML to a PNG (bytes) via wkhtmltoimage (stdin → stdout)."""
	import subprocess

	cmd = [
		"wkhtmltoimage",
		"--format", "png",
		"--encoding", "utf-8",
		"--quality", "94",
		"--width", "820",
		"--enable-local-file-access",
		"--quiet",
		"-", "-",
	]
	try:
		proc = subprocess.run(
			cmd, input=html.encode("utf-8"), stdout=subprocess.PIPE, stderr=subprocess.PIPE
		)
	except FileNotFoundError:
		frappe.throw(_("wkhtmltoimage is not installed on the server."))
	if not proc.stdout:
		frappe.log_error(proc.stderr.decode("utf-8", "ignore"), "wkhtmltoimage failed")
		frappe.throw(_("Could not render the image."))

	# wkhtmltoimage emits a large RGBA PNG. Flatten onto white + optimize to
	# keep the file small enough to save comfortably on a phone.
	try:
		import io

		from PIL import Image

		img = Image.open(io.BytesIO(proc.stdout))
		if img.mode in ("RGBA", "LA", "P"):
			bg = Image.new("RGB", img.size, "#FFFFFF")
			img = img.convert("RGBA")
			bg.paste(img, mask=img.split()[-1])
			img = bg
		else:
			img = img.convert("RGB")
		out = io.BytesIO()
		img.save(out, format="PNG", optimize=True)
		return out.getvalue()
	except Exception:
		return proc.stdout


def _logo_data_uri():
	"""Return the company logo as a base64 data URI (for the PDF), or ''."""
	import base64
	import os

	path = frappe.get_app_path("customize_erpnext", "public", "images", "logo_500.jpg")
	if not os.path.exists(path):
		return ""
	with open(path, "rb") as fh:
		b64 = base64.b64encode(fh.read()).decode()
	return f"data:image/jpeg;base64,{b64}"


def _strip_section_no(label):
	"""Drop a leading ordinal prefix from a section name for the receipt.

	e.g. "1. Thông tin chung" / "2) Trình độ" / "3 - CCCD" -> "Thông tin chung"…
	so the PDF/PNG shows the plain Section name without the numbering HR typed.
	"""
	import re

	cleaned = re.sub(r"^\s*\d+\s*[.)\-–—]\s*", "", label or "").strip()
	return cleaned or (label or "")


def _build_submission_html(doc, saved, config):
	company = (
		frappe.defaults.get_global_default("company")
		or frappe.db.get_single_value("Global Defaults", "default_company")
		or ""
	)
	logo = _logo_data_uri()
	submitted = frappe.utils.format_datetime(doc.submitted_on, "dd/MM/yyyy HH:mm") if doc.submitted_on else ""

	# The employee code is already shown in the header — skip any field that just
	# repeats it (e.g. a field mapped to `employee` / `name`).
	skip_fields = {"employee", "name"}
	rows = []
	for sec in config["sections"]:
		sec_rows = []
		for f in sec["fields"]:
			if f["fieldname"] in skip_fields:
				continue
			val = saved.get(f["fieldname"])
			val = "" if val is None else str(val)
			sec_rows.append(
				"<tr><td class='lbl'>{0}</td><td class='val'>{1}</td></tr>".format(
					frappe.utils.escape_html(f["label"]),
					frappe.utils.escape_html(val) or "&mdash;",
				)
			)
		if not sec_rows:
			continue  # drop a section that became empty
		rows.append(
			f'<tr><td colspan="2" class="sec">{frappe.utils.escape_html(_strip_section_no(sec["label"]))}</td></tr>'
		)
		rows.extend(sec_rows)

	remarks = saved.get(REMARKS_KEY) or ""
	remarks_block = ""
	if remarks:
		remarks_block = (
			"<div class='remarks'><div class='rlabel'>Ghi chú thêm</div>"
			f"<div class='rtext'>{frappe.utils.escape_html(remarks)}</div></div>"
		)

	logo_img = f"<img src='{logo}' class='logo'/>" if logo else ""

	return f"""
<!DOCTYPE html><html><head><meta charset="utf-8"><style>
  body{{font-family:'Helvetica Neue',Arial,sans-serif;color:#1f2733;font-size:12px;margin:0}}
  .head{{text-align:left;border-bottom:2px solid #1e3a8a;padding-bottom:10px;margin-bottom:14px;overflow:hidden}}
  .logo{{height:36px;float:left;margin:2px 12px 0 0}}
  .company{{font-size:12px;font-weight:bold;color:#1e3a8a;text-transform:uppercase}}
  .title{{font-size:21px;font-weight:bold;color:#1e3a8a;margin-top:2px}}
  .meta{{margin:10px 0;font-size:12px}}
  .meta b{{display:inline-block;min-width:130px}}
  table{{width:100%;border-collapse:collapse;margin-top:6px}}
  td{{border:1px solid #d5dbe3;padding:6px 8px;vertical-align:top}}
  td.sec{{background:#eef2ff;font-weight:bold;color:#1e3a8a}}
  td.lbl{{width:40%;background:#f7f9fc;color:#475569}}
  td.val{{width:60%}}
  .remarks{{margin-top:14px;border:1px solid #d5dbe3;border-radius:6px;padding:10px}}
  .rlabel{{font-weight:bold;color:#1e3a8a;margin-bottom:4px}}
</style></head><body>
  <div class="head">
    {logo_img}
    <div class="company">{frappe.utils.escape_html(company)}</div>
    <div class="title">PHIẾU CẬP NHẬT THÔNG TIN NHÂN VIÊN</div>
  </div>
  <div class="meta">
    <div><b>Mã nhân viên:</b> {frappe.utils.escape_html(doc.employee)}</div>
    <div><b>Họ và tên:</b> {frappe.utils.escape_html(doc.employee_name or "")}</div>
    <div><b>Thời điểm gửi:</b> {submitted}</div>
  </div>
  <table>{''.join(rows)}</table>
  {remarks_block}
</body></html>
"""


# ---------------------------------------------------------------------------
# HR APIs
# ---------------------------------------------------------------------------

@frappe.whitelist()
def download_excel(names=None):
	"""Export submissions to xlsx with two sheets — "New Data" (values submitted
	by employees) and "Old Data" (current Employee values).

	The file is designed to be re-imported into Employee via Data Import, so:
	  - the first column is "ID" (= Employee name → maps to the record on update),
	  - field column headers are the Employee field's own label (NOT label_vi),
	  - no Status / Submitted On columns (they would collide with Employee fields).

	"New Data" is the first sheet (the importable one); changed cells are
	highlighted. `names` = JSON list of Employee Self Update Info names, or None
	for all.
	"""
	_require_hr()
	import io

	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill

	if isinstance(names, str):
		names = json.loads(names or "null")

	filters = {"name": ["in", names]} if names else {}
	records = frappe.get_all(
		INFO_DT,
		filters=filters,
		fields=["name", "employee", "employee_name", "data_json"],
		order_by="employee asc",
	)

	config = _build_config()
	fields = [f for sec in config["sections"] for f in sec["fields"]]
	# Headers use the Employee field label so the file can be imported back.
	# A trailing "Ghi chú" column holds the employee's free-text remarks.
	header = ["ID", "Employee Name"] + [f["employee_label"] for f in fields] + ["Ghi chú (nhân viên)"]
	real_fieldnames = [f["fieldname"] for f in fields if not f.get("custom")]

	bold = Font(bold=True)
	changed_fill = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")

	# Pre-compute old (Employee) and new (submitted) value maps per record.
	rows_data = []
	for rec in records:
		saved = json.loads(rec.get("data_json") or "{}")
		old_vals = frappe.db.get_value(
			"Employee", rec.employee, real_fieldnames, as_dict=True
		) or {} if real_fieldnames else {}
		rows_data.append({"rec": rec, "saved": saved, "old": old_vals})

	def _write_header(ws):
		ws.append(header)
		for cell in ws[1]:
			cell.font = bold

	wb = Workbook()

	# Sheet 1 — New Data (submitted values; this is the importable sheet).
	ws_new = wb.active
	ws_new.title = "New Data"
	_write_header(ws_new)
	for rd in rows_data:
		rec, saved, old_vals = rd["rec"], rd["saved"], rd["old"]
		row = [rec.employee, rec.employee_name]
		row += [_fmt(saved.get(f["fieldname"])) for f in fields]
		row.append(_fmt(saved.get(REMARKS_KEY)))
		ws_new.append(row)
		excel_row = ws_new.max_row
		for idx, f in enumerate(fields):
			fn = f["fieldname"]
			if fn in saved and _fmt(saved.get(fn)) != _fmt(old_vals.get(fn)):
				ws_new.cell(row=excel_row, column=3 + idx).fill = changed_fill

	# Sheet 2 — Old Data (current values on the Employee record, for reference).
	ws_old = wb.create_sheet("Old Data")
	_write_header(ws_old)
	for rd in rows_data:
		rec = rd["rec"]
		row = [rec.employee, rec.employee_name]
		row += [_fmt(rd["old"].get(f["fieldname"])) for f in fields]
		row.append("")  # no "old" remarks
		ws_old.append(row)

	buf = io.BytesIO()
	wb.save(buf)
	frappe.response["filename"] = "employee_self_update_info.xlsx"
	frappe.response["filecontent"] = buf.getvalue()
	frappe.response["type"] = "binary"


def _fmt(value):
	if value is None:
		return ""
	return str(value)


# Field types HR may edit inline on the Desk form (plain text-like inputs).
# Excludes Select/Check/Date/Datetime/Time/Link and the address widgets.
_TEXT_EDITABLE_TYPES = {"Data", "Small Text", "Text", "Long Text", "Int", "Float", "Currency", "Phone"}
_MULTILINE_TYPES = {"Small Text", "Text", "Long Text"}


def _is_text_editable(f):
	"""True if HR can edit this field's value directly on the Desk form."""
	return (
		f.get("fieldtype") in _TEXT_EDITABLE_TYPES
		and f.get("widget") in (None, "", "Auto")
		and not f.get("read_only")
	)


@frappe.whitelist()
def update_submission_values(name, values):
	"""HR edits text-type submitted values inline on the Desk form.

	Only text-like fields (see _is_text_editable) are written back into
	`data_json`; Select/Date/Link/address widgets are ignored. Not allowed once
	the record is Synced. `values` = JSON {fieldname: value}.
	"""
	_require_hr()
	if isinstance(values, str):
		values = json.loads(values or "{}")
	doc = frappe.get_doc(INFO_DT, name)
	if doc.status == "Synced":
		frappe.throw(_("Đã đồng bộ — không sửa được. Dùng Edit in Portal nếu cần."))

	config = _build_config()
	editable = {
		f["fieldname"]
		for sec in config["sections"]
		for f in sec["fields"]
		if _is_text_editable(f)
	}
	saved = json.loads(doc.data_json or "{}")
	changed = 0
	for k, v in (values or {}).items():
		if k not in editable:
			continue
		nv = "" if v is None else str(v).strip()
		if str(saved.get(k, "")) != nv:
			saved[k] = nv
			changed += 1
	if changed:
		doc.data_json = json.dumps(saved, ensure_ascii=False)
		doc.save(ignore_permissions=True)
		frappe.db.commit()
	return {"ok": True, "changed": changed}


@frappe.whitelist()
def get_submission_view(name):
	"""Human-readable view of a submission for HR (instead of raw JSON).

	Returns {sections:[{label, rows:[{label, value, old, changed, custom}]}],
	remarks}. `changed` compares the submitted value with the current Employee
	value (only for real Employee fields).
	"""
	_require_hr()
	doc = frappe.get_doc(INFO_DT, name)
	saved = json.loads(doc.data_json or "{}")
	config = _build_config()

	real = [f["fieldname"] for sec in config["sections"] for f in sec["fields"] if not f.get("custom")]
	emp_vals = frappe.db.get_value("Employee", doc.employee, real, as_dict=True) or {} if real else {}

	sections = []
	for sec in config["sections"]:
		rows = []
		for f in sec["fields"]:
			fn = f["fieldname"]
			if fn in ("employee", "name"):
				continue  # identity — already shown on the record header
			new_v = saved.get(fn)
			new_s = "" if new_v is None else str(new_v)
			old_s = "" if f.get("custom") else _fmt(emp_vals.get(fn))
			rows.append({
				"fieldname": fn,
				"label": f["label"],
				"value": new_s,
				"old": old_s,
				"changed": (not f.get("custom")) and fn in saved and new_s != old_s,
				"custom": bool(f.get("custom")),
				"editable": _is_text_editable(f),
				"multiline": f.get("fieldtype") in _MULTILINE_TYPES,
			})
		if rows:
			sections.append({"label": sec["label"], "rows": rows})

	return {"sections": sections, "remarks": saved.get(REMARKS_KEY) or ""}


# ---------------------------------------------------------------------------
# Review + Sync to Employee
# ---------------------------------------------------------------------------

# Address groups on Employee whose "_full" is rebuilt = village, commune, province.
_ADDRESS_GROUPS = [
	{
		"province": "custom_current_address_province",
		"commune": "custom_current_address_commune",
		"village": "custom_current_address_village",
		"full": "custom_current_address_full",
	},
	{
		"province": "custom_permanent_address_province",
		"commune": "custom_permanent_address_commune",
		"village": "custom_permanent_address_village",
		"full": "custom_permanent_address_full",
	},
	# 🚧 TẠM TẮT 21/08/2026 — field custom_place_of_origin_address_* đã bị gỡ khỏi Employee.
	#    Giữ nguyên để khai lại sau; bỏ comment cả khối là chạy như cũ.
	# {
	# "province": "custom_place_of_origin_address_province",
	# "commune": "custom_place_of_origin_address_commune",
	# "village": "custom_place_of_origin_address_village",
	# "full": "custom_place_of_origin_address_full",
	# },
]


def _coerce_for_employee(value, fieldtype):
	"""Convert a submitted (string) value to something safe for emp.set()."""
	if value is None:
		return None
	sval = str(value).strip()
	if sval == "":
		# Empty → clear numeric/date/link fields; keep "" for text.
		if fieldtype in ("Int", "Float", "Currency", "Date", "Datetime", "Time", "Link"):
			return None
		return ""
	if fieldtype == "Check":
		return 1 if sval in ("1", "true", "True", "on", "yes", "Có") else 0
	if fieldtype == "Int":
		return int(float(sval))
	if fieldtype in ("Float", "Currency"):
		return float(sval)
	return sval


def _rebuild_address_full(emp, changed_fieldnames):
	"""Recompute custom_*_address_full = village + commune + province (like the
	Employee form) for any address group whose parts were touched."""
	for g in _ADDRESS_GROUPS:
		if not any(g[k] in changed_fieldnames for k in ("province", "commune", "village")):
			continue
		parts = [emp.get(g["village"]), emp.get(g["commune"]), emp.get(g["province"])]
		emp.set(g["full"], ", ".join([p for p in parts if p]))


@frappe.whitelist()
def review_forms(names):
	"""Mark Submitted forms as Reviewed (required before syncing).

	`names` = JSON list of Employee Self Update Info names.
	Returns {reviewed, skipped, results:[{employee, ok, message}]}.
	"""
	_require_hr()
	if isinstance(names, str):
		names = json.loads(names or "[]")
	if not names:
		frappe.throw(_("No records selected."))

	reviewed, skipped, results = 0, 0, []
	for name in names:
		doc = frappe.get_doc(INFO_DT, name)
		if doc.status != "Submitted":
			skipped += 1
			results.append({"employee": doc.employee, "ok": False,
				"message": _("Bỏ qua — trạng thái {0} (chỉ review được bản Submitted)").format(doc.status)})
			continue
		doc.status = "Reviewed"
		doc.reviewed_on = now_datetime()
		doc.reviewed_by = frappe.session.user
		doc.save(ignore_permissions=True)
		reviewed += 1
		results.append({"employee": doc.employee, "ok": True, "message": _("Đã review")})
	frappe.db.commit()
	return {"reviewed": reviewed, "skipped": skipped, "results": results}


@frappe.whitelist()
def get_syncable_fields():
	"""Return the real Employee fields (non-custom) that Sync can write, for the
	HR field-picker dialog. `[{fieldname, label}]`."""
	_require_hr()
	config = _build_config()
	return [
		{"fieldname": f["fieldname"], "label": f["label"]}
		for sec in config["sections"]
		for f in sec["fields"]
		if not f.get("custom") and f["fieldname"] not in ("employee", "name")
	]


@frappe.whitelist()
def sync_to_employee(names, fields=None):
	"""Write submissions into the Employee record.

	Only fields that exist on Employee (incl. custom_ fields) are written; custom
	free-form fields and remarks are ignored. `fields` (optional JSON list of
	fieldnames) limits Sync to the HR-selected fields; omitted → all configured
	fields. When Disable Review is on, records are synced straight from Submitted;
	otherwise only Reviewed records are synced. Each record is saved independently
	so one failure does not block the rest. Returns {synced, failed, skipped, results}.
	"""
	_require_hr()
	if isinstance(names, str):
		names = json.loads(names or "[]")
	if not names:
		frappe.throw(_("No records selected."))
	if isinstance(fields, str):
		fields = json.loads(fields or "null")
	if fields is not None and not fields:
		frappe.throw(_("Select at least one field to sync."))

	setting = _get_setting()
	# Disable Review → sync from Submitted; else require Reviewed. Reviewed is
	# always accepted so legacy already-reviewed records still sync.
	allowed = {"Submitted", "Reviewed"} if setting.get("disable_review") else {"Reviewed"}

	config = _build_config()
	# Real Employee fields only (non-custom, excluding identity fields).
	emp_fields = {
		f["fieldname"]: f
		for sec in config["sections"]
		for f in sec["fields"]
		if not f.get("custom") and f["fieldname"] not in ("employee", "name")
	}
	# Limit to the HR-selected fields when provided.
	if fields is not None:
		sel = set(fields)
		emp_fields = {fn: f for fn, f in emp_fields.items() if fn in sel}

	synced, failed, skipped, results = 0, 0, 0, []
	for name in names:
		doc = frappe.get_doc(INFO_DT, name)
		if doc.status not in allowed:
			skipped += 1
			need = _("Submitted") if setting.get("disable_review") else _("Reviewed")
			results.append({"employee": doc.employee, "employee_name": doc.employee_name, "ok": False,
				"message": _("Bỏ qua — cần trạng thái {0} để Sync (hiện tại: {1})").format(need, doc.status)})
			continue

		saved = json.loads(doc.data_json or "{}")
		try:
			emp = frappe.get_doc("Employee", doc.employee)
			meta = emp.meta
			changed = []
			for fn, f in emp_fields.items():
				if fn not in saved:
					continue
				if not meta.has_field(fn):
					continue  # field no longer on Employee
				emp.set(fn, _coerce_for_employee(saved.get(fn), f["fieldtype"]))
				changed.append(fn)
			_rebuild_address_full(emp, set(changed))
			emp.save(ignore_permissions=True)

			doc.status = "Synced"
			doc.synced_on = now_datetime()
			doc.synced_by = frappe.session.user
			doc.save(ignore_permissions=True)
			frappe.db.commit()
			synced += 1
			results.append({"employee": doc.employee, "employee_name": doc.employee_name, "ok": True,
				"message": _("Đã đồng bộ {0} trường").format(len(changed))})
		except Exception as e:
			frappe.db.rollback()
			failed += 1
			results.append({"employee": doc.employee, "employee_name": doc.employee_name, "ok": False,
				"message": str(e)})
	return {"synced": synced, "failed": failed, "skipped": skipped, "results": results}
