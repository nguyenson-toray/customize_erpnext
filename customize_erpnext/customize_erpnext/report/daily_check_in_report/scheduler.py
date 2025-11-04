import frappe
from frappe import _
from frappe.utils import today, formatdate, get_datetime, add_days
from datetime import datetime, time as time_obj, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import tempfile

@frappe.whitelist()
def send_daily_check_in_report():
    """
    Scheduled job to send Daily Check-in Report every day at 8:15 AM
    Sends email with Excel attachment
    Can also be called manually from Web Console for testing
    """
    try:
        # Get current date
        report_date = today()

        # Import the report execute function
        from customize_erpnext.customize_erpnext.report.daily_check_in_report.daily_check_in_report import get_data

        # Prepare filters
        filters = {
            "date": report_date,
            "status": "Absent",
            "show_maternity_leave": 1,
            "show_all_checkins": 0
        }

        # Get report data
        data = get_data(filters)

        # Get incomplete check-ins from day 26 of previous month to yesterday
        from frappe.utils import get_first_day, add_months
        current_month_first = get_first_day(report_date)
        prev_month_26 = add_days(add_months(current_month_first, -1), 25)  # Day 26 of previous month
        yesterday = add_days(report_date, -1)
        incomplete_checkins = get_incomplete_checkins(prev_month_26, yesterday)

        # Calculate statistics
        stats = calculate_statistics(report_date, data, incomplete_checkins)

        # Get last employee checkin time
        last_checkin_time = get_last_employee_checkin_time()

        # Generate email content
        email_subject = f"Báo cáo hiện diện / vắng ngày {formatdate(report_date, 'dd/MM/yyyy')}"
        email_content = generate_email_content(report_date, stats, data, last_checkin_time)

        # Generate Excel file
        excel_file_path, excel_file_name = generate_excel_report(report_date, stats)

        # Send email with Excel attachment
        recipients = ["it@tiqn.com.vn", "ni.nht@tiqn.com.vn", "hoanh.ltk@tiqn.com.vn","loan.ptk@tiqn.com.vn"]
        # recipients = ["son.nt@tiqn.com.vn"]
        frappe.sendmail(
            recipients=recipients,
            subject=email_subject,
            message=email_content,
            attachments=[{
                'fname': excel_file_name,
                'fcontent': open(excel_file_path, 'rb').read()
            }],
            delayed=False
        )

        # Clean up temporary file
        try:
            os.remove(excel_file_path)
        except Exception as cleanup_error:
            frappe.logger().warning(f"Failed to clean up temp file {excel_file_path}: {str(cleanup_error)}")

        frappe.logger().info(f"Daily Check-in Report with Excel attachment sent successfully for {report_date}")

    except Exception as e:
        frappe.logger().error(f"Error sending Daily Check-in Report: {str(e)}")
        frappe.log_error(
            title="Daily Check-in Report Scheduler Error",
            message=frappe.get_traceback()
        )

def get_incomplete_checkins(start_date, end_date):
    """
    Get list of employees who have incomplete check-ins from start_date to end_date
    This includes:
    - Employees with only 1 automatic check-in (device_id IS NOT NULL) (missing check-out)
    - Employees with NO check-in AFTER shift begin time (all check-ins are <= begin time)
    - Employees with NO check-out BEFORE shift end time (all check-outs are >= end time)

    Also retrieves manual check-ins (device_id IS NULL) for "Đã xử lý" column
    Returns list of employees with incomplete check-ins, sorted by date DESC, employee_code ASC
    """
    # Get all employees with their automatic check-ins (device_id IS NOT NULL) and shift information
    # For each employee-date combination
    employees_with_checkins = frappe.db.sql(f"""
        SELECT
            e.attendance_device_id,
            e.name AS employee_code,
            e.employee_name,
            e.department,
            e.custom_group,
            e.designation,
            DATE(ec.time) AS checkin_date,
            MIN(ec.time) AS first_check_in,
            MAX(ec.time) AS last_check_out,
            COUNT(ec.name) AS checkin_count,
            COALESCE(
                (SELECT srd.shift
                 FROM `tabShift Registration Detail` srd
                 JOIN `tabShift Registration` sr ON srd.parent = sr.name
                 WHERE srd.employee = e.name
                   AND srd.begin_date <= DATE(ec.time)
                   AND srd.end_date >= DATE(ec.time)
                   AND sr.docstatus = 1
                 ORDER BY sr.creation DESC
                 LIMIT 1),
                CASE
                    WHEN e.custom_group = 'Canteen' THEN 'Canteen'
                    ELSE 'Day'
                END
            ) AS shift,
            sn.begin_time,
            sn.end_time
        FROM
            `tabEmployee` e
        INNER JOIN
            `tabEmployee Checkin` ec
        ON
            e.name = ec.employee
            AND DATE(ec.time) >= '{start_date}'
            AND DATE(ec.time) <= '{end_date}'
            AND ec.device_id IS NOT NULL
        LEFT JOIN
            `tabShift Name` sn
        ON
            sn.shift_name = COALESCE(
                (SELECT srd.shift
                 FROM `tabShift Registration Detail` srd
                 JOIN `tabShift Registration` sr ON srd.parent = sr.name
                 WHERE srd.employee = e.name
                   AND srd.begin_date <= DATE(ec.time)
                   AND srd.end_date >= DATE(ec.time)
                   AND sr.docstatus = 1
                 ORDER BY sr.creation DESC
                 LIMIT 1),
                CASE
                    WHEN e.custom_group = 'Canteen' THEN 'Canteen'
                    ELSE 'Day'
                END
            )
        WHERE
            e.status = 'Active'
        GROUP BY
            e.name, DATE(ec.time)
        ORDER BY
            DATE(ec.time) DESC, e.name ASC
    """, as_dict=1)

    # Filter employees with incomplete check-ins
    incomplete_list = []
    for emp in employees_with_checkins:
        is_incomplete = False
        checkin_count = emp.get('checkin_count')
        checkin_date = emp.get('checkin_date')

        # Rule 1: Only 1 check-in
        if checkin_count == 1:
            is_incomplete = True
        # Rule 2 & 3: >= 2 check-ins but all before begin_time OR all after end_time
        elif checkin_count >= 2:
            # Convert shift times from timedelta to time objects FIRST
            begin_time = emp.get('begin_time')
            end_time = emp.get('end_time')

            # Convert begin_time to time object
            if begin_time:
                if isinstance(begin_time, timedelta):
                    total_seconds = int(begin_time.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    begin_time = time_obj(hours, minutes, seconds)
                elif not isinstance(begin_time, type(time_obj(0, 0, 0))):
                    begin_time = None  # Invalid type, skip

            # Convert end_time to time object
            if end_time:
                if isinstance(end_time, timedelta):
                    total_seconds = int(end_time.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    end_time = time_obj(hours, minutes, seconds)
                elif not isinstance(end_time, type(time_obj(0, 0, 0))):
                    end_time = None  # Invalid type, skip

            # Get all automatic check-in times for this employee on this date
            all_checkins = frappe.db.sql(f"""
                SELECT time
                FROM `tabEmployee Checkin`
                WHERE employee = '{emp.get('employee_code')}'
                  AND DATE(time) = '{checkin_date}'
                  AND device_id IS NOT NULL
                ORDER BY time
            """, as_dict=1)

            # Rule 2: All check-ins <= begin_time (missing proper check-in)
            # Only check if begin_time is a valid time object
            if begin_time and all_checkins and isinstance(begin_time, type(time_obj(0, 0, 0))):
                all_before_or_at_begin = True
                for checkin in all_checkins:
                    checkin_time = checkin.get('time').time()
                    if checkin_time > begin_time:
                        all_before_or_at_begin = False
                        break

                if all_before_or_at_begin:
                    is_incomplete = True

            # Rule 3: All check-ins >= end_time (missing proper check-out)
            # Only check if end_time is a valid time object
            if end_time and all_checkins and not is_incomplete and isinstance(end_time, type(time_obj(0, 0, 0))):
                all_after_or_at_end = True
                for checkin in all_checkins:
                    checkin_time = checkin.get('time').time()
                    if checkin_time < end_time:
                        all_after_or_at_end = False
                        break

                if all_after_or_at_end:
                    is_incomplete = True

        if is_incomplete:
            # Get manual check-ins (device_id IS NULL) for this employee on this date
            manual_checkins = frappe.db.sql(f"""
                SELECT
                    TIME(time) as checkin_time,
                    custom_reason_for_manual_check_in,
                    custom_other_reason_for_manual_check_in
                FROM `tabEmployee Checkin`
                WHERE employee = '{emp.get('employee_code')}'
                  AND DATE(time) = '{checkin_date}'
                  AND device_id IS NULL
                ORDER BY time
            """, as_dict=1)

            # Format manual check-ins for "Đã xử lý" column
            manual_checkin_times = []
            reasons = []
            other_reasons = []

            for mc in manual_checkins:
                if mc.get('checkin_time'):
                    manual_checkin_times.append(str(mc.get('checkin_time')))
                if mc.get('custom_reason_for_manual_check_in'):
                    reasons.append(mc.get('custom_reason_for_manual_check_in'))
                if mc.get('custom_other_reason_for_manual_check_in'):
                    other_reasons.append(mc.get('custom_other_reason_for_manual_check_in'))

            emp['manual_checkins'] = ', '.join(manual_checkin_times) if manual_checkin_times else ''
            emp['reason_for_manual'] = ', '.join(set(reasons)) if reasons else ''
            emp['other_reason_for_manual'] = ', '.join(set(other_reasons)) if other_reasons else ''

            incomplete_list.append(emp)

    return incomplete_list

def get_last_employee_checkin_time():
    """
    Get the last employee checkin time from the system
    Returns formatted time string or None if no checkins found
    """
    try:
        last_checkin = frappe.db.sql("""
            SELECT MAX(time) as last_time
            FROM `tabEmployee Checkin`
        """, as_dict=True)

        if last_checkin and last_checkin[0].get('last_time'):
            last_time = last_checkin[0].get('last_time')
            # Format as datetime string
            if isinstance(last_time, str):
                last_time = get_datetime(last_time)
            return last_time.strftime("%H:%M:%S %d/%m/%Y")
        return None
    except Exception as e:
        frappe.logger().error(f"Error getting last checkin time: {str(e)}")
        return None

def calculate_statistics(report_date, absent_data, incomplete_checkins):
    """
    Calculate statistics for the report:
    - Total active employees
    - Total present
    - Total absent (excluding maternity leave)
    - Total on maternity leave
    - Incomplete check-ins from previous day
    """
    # Count total active employees
    total_employees = frappe.db.count("Employee", filters={"status": "Active"})

    # Separate absent data into regular absent and maternity leave
    absent_regular = []
    absent_maternity = []

    for emp in absent_data:
        status_info = emp.get("status_info") or ""
        if status_info == "Maternity Leave":
            absent_maternity.append(emp)
        else:
            absent_regular.append(emp)

    # Sort both lists by custom_group (A-Z)
    absent_regular = sorted(absent_regular, key=lambda x: (x.get("custom_group") or "").lower())
    absent_maternity = sorted(absent_maternity, key=lambda x: (x.get("custom_group") or "").lower())

    # Sort incomplete check-ins by date DESC (newest first), then by employee_code ASC (A-Z)
    incomplete_sorted = sorted(
        incomplete_checkins,
        key=lambda x: (
            x.get("checkin_date") if x.get("checkin_date") else "",
            x.get("employee_code") or ""
        ),
        reverse=True  # This will sort date DESC, but we need to handle employee_code separately
    )
    # Better sorting: date DESC, employee_code ASC
    from datetime import date as date_type
    incomplete_sorted = sorted(
        incomplete_checkins,
        key=lambda x: (
            -(x.get("checkin_date").toordinal() if isinstance(x.get("checkin_date"), date_type) else 0),
            x.get("employee_code") or ""
        )
    )

    # Count absent employees (excluding maternity leave)
    total_absent = len(absent_regular)

    # Count employees on maternity leave
    maternity_count = len(absent_maternity)

    # Count incomplete check-ins
    incomplete_count = len(incomplete_checkins)

    # Count incomplete check-ins that have been processed (have manual check-ins)
    incomplete_processed = len([emp for emp in incomplete_sorted if emp.get('manual_checkins', '')])

    # Calculate present employees
    total_present = total_employees - total_absent - maternity_count

    return {
        "total_employees": total_employees,
        "total_present": total_present,
        "total_absent": total_absent,
        "maternity_count": maternity_count,
        "incomplete_count": incomplete_count,
        "incomplete_processed": incomplete_processed,
        "absent_regular": absent_regular,
        "absent_maternity": absent_maternity,
        "incomplete_checkins": incomplete_sorted
    }

def generate_email_content(report_date, stats, absent_data, last_checkin_time=None):
    """
    Generate HTML email content with statistics and two separate absent employee lists
    """
    formatted_date = formatdate(report_date, "dd/MM/yyyy")
    current_time = datetime.now().strftime("%H:%M:%S %d/%m/%Y")

    # Format last checkin time message
    last_data_time_msg = ""
    if last_checkin_time:
        last_data_time_msg = f"<br>Thời điểm chấm công sau cùng: {last_checkin_time}"

    # Build regular absent employee table (excluding maternity leave)
    absent_regular_rows = ""
    absent_regular_list = stats.get('absent_regular', [])
    formatted_date = formatdate(report_date, "dd/MM/yyyy")

    if absent_regular_list:
        for idx, emp in enumerate(absent_regular_list, 1):
            status_info = emp.get("status_info") or ""
            absent_regular_rows += f"""
            <tr>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{idx}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{formatted_date}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('attendance_device_id') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('employee_code') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('employee_name') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('department') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('custom_group') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('shift') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('designation') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{status_info}</td>
            </tr>
            """
    else:
        absent_regular_rows = """
        <tr>
            <td colspan="10" style="border: 1px solid #ddd; padding: 8px; text-align: center; color: #999;">
                Không có nhân viên vắng
            </td>
        </tr>
        """

    # Build maternity leave employee table
    maternity_leave_rows = ""
    absent_maternity_list = stats.get('absent_maternity', [])
    if absent_maternity_list:
        for idx, emp in enumerate(absent_maternity_list, 1):
            status_info = emp.get("status_info") or ""
            maternity_leave_rows += f"""
            <tr>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{idx}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{formatted_date}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('attendance_device_id') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('employee_code') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('employee_name') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('department') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('custom_group') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('shift') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('designation') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{status_info}</td>
            </tr>
            """
    else:
        maternity_leave_rows = """
        <tr>
            <td colspan="10" style="border: 1px solid #ddd; padding: 8px; text-align: center; color: #999;">
                Không có nhân viên nghỉ thai sản
            </td>
        </tr>
        """

    # Build incomplete check-ins table (from day 26 of previous month to yesterday)
    incomplete_checkin_rows = ""
    incomplete_list = stats.get('incomplete_checkins', [])
    from frappe.utils import get_first_day, add_months
    current_month_first = get_first_day(report_date)
    prev_month_26 = add_days(add_months(current_month_first, -1), 25)
    yesterday = add_days(report_date, -1)
    date_range_formatted = f"{formatdate(prev_month_26, 'dd/MM/yyyy')} - {formatdate(yesterday, 'dd/MM/yyyy')}"

    if incomplete_list:
        for idx, emp in enumerate(incomplete_list, 1):
            checkin_count = emp.get('checkin_count') or 0
            checkin_date = emp.get('checkin_date')
            checkin_date_formatted = formatdate(checkin_date, "dd/MM/yyyy") if checkin_date else ""

            first_checkin = ""
            last_checkout = ""

            if checkin_count == 1:
                # Rule 1: Single checkin - determine if it's check-in or check-out based on distance to shift times
                single_time = emp.get("first_check_in")
                if single_time:
                    if isinstance(single_time, str):
                        single_time = get_datetime(single_time)
                    formatted_time = single_time.strftime("%H:%M:%S %d/%m/%Y")
                    single_time_only = single_time.time()

                    # Get shift begin and end times
                    begin_time = emp.get('begin_time')
                    end_time = emp.get('end_time')

                    # Convert timedelta to time
                    if begin_time and isinstance(begin_time, timedelta):
                        total_seconds = int(begin_time.total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        begin_time = time_obj(hours, minutes, seconds)

                    if end_time and isinstance(end_time, timedelta):
                        total_seconds = int(end_time.total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        end_time = time_obj(hours, minutes, seconds)

                    # Calculate distance from begin and end time
                    if begin_time and end_time:
                        # Convert time to seconds for easier calculation
                        single_seconds = single_time_only.hour * 3600 + single_time_only.minute * 60 + single_time_only.second
                        begin_seconds = begin_time.hour * 3600 + begin_time.minute * 60 + begin_time.second
                        end_seconds = end_time.hour * 3600 + end_time.minute * 60 + end_time.second

                        # Calculate absolute distance
                        distance_to_begin = abs(single_seconds - begin_seconds)
                        distance_to_end = abs(single_seconds - end_seconds)

                        # If closer to begin time, it's a check-in; if closer to end time, it's a check-out
                        if distance_to_begin < distance_to_end:
                            first_checkin = formatted_time
                            last_checkout = ""
                        else:
                            first_checkin = ""
                            last_checkout = formatted_time
                    else:
                        # No shift info - use simple time-based logic (before 12:00 = check-in, after 12:00 = check-out)
                        noon_seconds = 12 * 3600  # 12:00:00
                        single_seconds = single_time_only.hour * 3600 + single_time_only.minute * 60 + single_time_only.second

                        if single_seconds < noon_seconds:
                            first_checkin = formatted_time
                            last_checkout = ""
                        else:
                            first_checkin = ""
                            last_checkout = formatted_time
            else:
                # Rule 2 & 3: Multiple checkins (>=2) - always show first as check-in and last as check-out
                first_checkin_time = emp.get("first_check_in")
                if first_checkin_time:
                    if isinstance(first_checkin_time, str):
                        first_checkin_time = get_datetime(first_checkin_time)
                    first_checkin = first_checkin_time.strftime("%H:%M:%S %d/%m/%Y")

                last_checkout_time = emp.get("last_check_out")
                if last_checkout_time:
                    if isinstance(last_checkout_time, str):
                        last_checkout_time = get_datetime(last_checkout_time)
                    last_checkout = last_checkout_time.strftime("%H:%M:%S %d/%m/%Y")

            # Get manual check-in info
            manual_checkins = emp.get('manual_checkins', '')
            reason_for_manual = emp.get('reason_for_manual', '')
            other_reason_for_manual = emp.get('other_reason_for_manual', '')

            incomplete_checkin_rows += f"""
            <tr>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{idx}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{checkin_date_formatted}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('attendance_device_id') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('employee_code') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('employee_name') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('department') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('custom_group') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('shift') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{emp.get('designation') or ''}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{first_checkin}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{last_checkout}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{checkin_count}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{manual_checkins}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{reason_for_manual}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{other_reason_for_manual}</td>
            </tr>
            """
    else:
        incomplete_checkin_rows = """
        <tr>
            <td colspan="15" style="border: 1px solid #ddd; padding: 8px; text-align: center; color: #999;">
                Không có nhân viên chấm công thiếu
            </td>
        </tr>
        """

    html_content = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
            .summary {{ margin: 20px 0; padding: 15px; background-color: #f5f5f5; border-radius: 5px; }}
            .summary-item {{ margin: 10px 0; font-size: 14px; }}
            .summary-item strong {{ color: #333; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background-color: #4CAF50; color: white; padding: 10px; text-align: left; border: 1px solid #ddd; }}
            .number {{ color: #2196F3; font-weight: bold; }}
            .present {{ color: #4CAF50; font-weight: bold; }}
            .absent {{ color: #f44336; font-weight: bold; }}
            .maternity {{ color: #FF9800; font-weight: bold; }}
            .incomplete {{ color: #9C27B0; font-weight: bold; }}
        </style>
    </head>
    <body>
        <h2 style="color: #333;">Báo cáo hiện diện / vắng ngày {formatted_date}</h2>

        <div class="summary">
        <h3 style="margin-top: 0; color: #555;">Email này được gửi tự động từ hệ thống ERPNext (Site: {get_current_frappe_site_name()}) vào lúc {current_time}{last_data_time_msg}</h3>
            <h3 style="margin-top: 0; color: #555;">Tổng quan:</h3>
            <div class="summary-item">
                <strong>Số lượng nhân viên (Active):</strong>
                <span class="number">{stats['total_employees']}</span> người
            </div>
            <div class="summary-item">
                <strong>Số lượng hiện diện:</strong>
                <span class="present">{stats['total_present']}</span> người
            </div>
            <div class="summary-item">
                <strong>Số lượng vắng (không bao gồm nghỉ thai sản):</strong>
                <span class="absent">{stats['total_absent']}</span> người
            </div>
            <div class="summary-item">
                <strong>Số lượng nghỉ thai sản:</strong>
                <span class="maternity">{stats['maternity_count']}</span> người
            </div>
            <div class="summary-item">
                <strong>Số lượng chấm công thiếu từ {date_range_formatted}:</strong>
                <span class="incomplete">{stats['incomplete_count']}</span> người
                <br>
                <strong>Đã xử lý:</strong>
                <span style="color: #4CAF50; font-weight: bold;">{stats['incomplete_processed']}</span> / <span class="incomplete">{stats['incomplete_count']}</span>
                <br>
                <small style="color: #666;">(Chỉ có 1 lần chấm công hoặc thiếu giờ chấm công vào/ra theo ca (Trên máy chấm công))</small>
            </div>
        </div>

        <h3 style="color: #555;">Danh sách nhân viên vắng (không bao gồm nghỉ thai sản):</h3>
        <table>
            <thead>
                <tr>
                    <th style="width: 4%; text-align: center;">STT</th>
                    <th style="width: 6%; text-align: center;">Ngày</th>
                    <th style="width: 7%;">Att ID</th>
                    <th style="width: 9%;">Employee</th>
                    <th style="width: 16%;">Employee Name</th>
                    <th style="width: 11%;">Department</th>
                    <th style="width: 11%;">Group</th>
                    <th style="width: 9%;">Shift</th>
                    <th style="width: 13%;">Designation</th>
                    <th style="width: 9%; text-align: center;">Status Info</th>
                </tr>
            </thead>
            <tbody>
                {absent_regular_rows}
            </tbody>
        </table>

        <h3 style="color: #555; margin-top: 30px;">Danh sách nhân viên nghỉ thai sản:</h3>
        <table>
            <thead>
                <tr>
                    <th style="width: 4%; text-align: center;">STT</th>
                    <th style="width: 6%; text-align: center;">Ngày</th>
                    <th style="width: 7%;">Att ID</th>
                    <th style="width: 9%;">Employee</th>
                    <th style="width: 16%;">Employee Name</th>
                    <th style="width: 11%;">Department</th>
                    <th style="width: 11%;">Group</th>
                    <th style="width: 9%;">Shift</th>
                    <th style="width: 13%;">Designation</th>
                    <th style="width: 9%; text-align: center;">Status Info</th>
                </tr>
            </thead>
            <tbody>
                {maternity_leave_rows}
            </tbody>
        </table>

        <h3 style="color: #555; margin-top: 30px;">Danh sách nhân viên chấm công thiếu từ {date_range_formatted}:</h3>
        <table>
            <thead>
                <tr>
                    <th style="width: 3%; text-align: center;">STT</th>
                    <th style="width: 6%; text-align: center;">Ngày</th>
                    <th style="width: 6%;">Att ID</th>
                    <th style="width: 8%;">Employee</th>
                    <th style="width: 12%;">Employee Name</th>
                    <th style="width: 10%;">Department</th>
                    <th style="width: 8%;">Group</th>
                    <th style="width: 7%;">Shift</th>
                    <th style="width: 10%;">Designation</th>
                    <th style="width: 8%; text-align: center;">Check-in</th>
                    <th style="width: 8%; text-align: center;">Check-out</th>
                    <th style="width: 4%; text-align: center;">Số lần chấm</th>
                    <th style="width: 8%; text-align: center;">Đã xử lý</th>
                    <th style="width: 6%;">Reason</th>
                    <th style="width: 6%;">Other Reason</th>
                </tr>
            </thead>
            <tbody>
                {incomplete_checkin_rows}
            </tbody>
        </table>

    </body>
    </html>
    """

    return html_content

def generate_excel_report(report_date, stats):
    """
    Generate Excel file with 3 sheets:
    1. Danh sách nhân viên vắng (không bao gồm nghỉ thai sản)
    2. Danh sách nhân viên nghỉ thai sản
    3. Danh sách nhân viên chấm công thiếu
    """
    # Create a new workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(name='Arial', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Headers definition
    headers = ['STT', 'Ngày', 'Att ID', 'Employee', 'Employee Name', 'Department', 'Group', 'Shift', 'Designation', 'Status Info']
    column_widths = [8, 12, 12, 15, 25, 20, 20, 15, 20, 20]
    formatted_date = formatdate(report_date, "dd/MM/yyyy")

    # Sheet 1: Danh sách nhân viên vắng (không bao gồm nghỉ thai sản)
    ws1 = wb.create_sheet("Absent")
    ws1.append(headers)

    # Apply header styles for sheet 1
    for col_num, (header, width) in enumerate(zip(headers, column_widths), 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Add data for sheet 1
    absent_regular = stats.get('absent_regular', [])
    for idx, emp in enumerate(absent_regular, 1):
        row_data = [
            idx,
            formatted_date,
            emp.get('attendance_device_id') or '',
            emp.get('employee_code') or '',
            emp.get('employee_name') or '',
            emp.get('department') or '',
            emp.get('custom_group') or '',
            emp.get('shift') or '',
            emp.get('designation') or '',
            emp.get('status_info') or ''
        ]
        ws1.append(row_data)

        # Apply styles to data cells
        for col_num in range(1, len(headers) + 1):
            cell = ws1.cell(row=idx + 1, column=col_num)
            cell.font = cell_font
            # Center align: STT (1), Ngày (2)
            cell.alignment = center_alignment if col_num in [1, 2] else cell_alignment
            cell.border = border

    # Add Table format for sheet 1
    if len(absent_regular) > 0:
        table1 = Table(displayName="absent", ref=f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(absent_regular) + 1}")
        style1 = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table1.tableStyleInfo = style1
        ws1.add_table(table1)

    # Sheet 2: Danh sách nhân viên nghỉ thai sản
    ws2 = wb.create_sheet("Maternity Leave")
    ws2.append(headers)

    # Apply header styles for sheet 2
    for col_num, (header, width) in enumerate(zip(headers, column_widths), 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Add data for sheet 2
    absent_maternity = stats.get('absent_maternity', [])
    for idx, emp in enumerate(absent_maternity, 1):
        row_data = [
            idx,
            formatted_date,
            emp.get('attendance_device_id') or '',
            emp.get('employee_code') or '',
            emp.get('employee_name') or '',
            emp.get('department') or '',
            emp.get('custom_group') or '',
            emp.get('shift') or '',
            emp.get('designation') or '',
            emp.get('status_info') or ''
        ]
        ws2.append(row_data)

        # Apply styles to data cells
        for col_num in range(1, len(headers) + 1):
            cell = ws2.cell(row=idx + 1, column=col_num)
            cell.font = cell_font
            # Center align: STT (1), Ngày (2)
            cell.alignment = center_alignment if col_num in [1, 2] else cell_alignment
            cell.border = border

    # Add Table format for sheet 2
    if len(absent_maternity) > 0:
        table2 = Table(displayName="maternity_leave", ref=f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(absent_maternity) + 1}")
        style2 = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table2.tableStyleInfo = style2
        ws2.add_table(table2)

    # Sheet 3: Danh sách nhân viên chấm công thiếu
    # Note: Sheet names cannot contain: / \ ? * [ ]
    # Replace / with - in date format
    from frappe.utils import get_first_day, add_months
    current_month_first = get_first_day(report_date)
    prev_month_26 = add_days(add_months(current_month_first, -1), 25)
    yesterday = add_days(report_date, -1)
    sheet3_date_from = formatdate(prev_month_26, "dd-MM")
    sheet3_date_to = formatdate(yesterday, "dd-MM")
    ws3 = wb.create_sheet(f"Missing {sheet3_date_from} to {sheet3_date_to}")
    headers_incomplete = ['STT', 'Ngày', 'Att ID', 'Employee', 'Employee Name', 'Department', 'Group', 'Shift', 'Designation', 'Check-in', 'Check-out', 'Số lần chấm', 'Đã xử lý', 'Reason', 'Other Reason']
    column_widths_incomplete = [8, 12, 12, 15, 25, 20, 20, 15, 20, 20, 20, 15, 15, 15, 20]
    ws3.append(headers_incomplete)

    # Apply header styles for sheet 3
    for col_num, (header, width) in enumerate(zip(headers_incomplete, column_widths_incomplete), 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Add data for sheet 3
    incomplete_checkins = stats.get('incomplete_checkins', [])
    for idx, emp in enumerate(incomplete_checkins, 1):
        checkin_count = emp.get('checkin_count') or 0

        first_checkin = ""
        last_checkout = ""

        if checkin_count == 1:
            # Rule 1: Single checkin - determine if it's check-in or check-out based on distance to shift times
            single_time = emp.get("first_check_in")
            if single_time:
                if isinstance(single_time, str):
                    single_time = get_datetime(single_time)
                formatted_time = single_time.strftime("%H:%M:%S %d/%m/%Y")
                single_time_only = single_time.time()

                # Get shift begin and end times
                begin_time = emp.get('begin_time')
                end_time = emp.get('end_time')

                # Convert timedelta to time
                if begin_time and isinstance(begin_time, timedelta):
                    total_seconds = int(begin_time.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    begin_time = time_obj(hours, minutes, seconds)

                if end_time and isinstance(end_time, timedelta):
                    total_seconds = int(end_time.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    end_time = time_obj(hours, minutes, seconds)

                # Calculate distance from begin and end time
                if begin_time and end_time:
                    # Convert time to seconds for easier calculation
                    single_seconds = single_time_only.hour * 3600 + single_time_only.minute * 60 + single_time_only.second
                    begin_seconds = begin_time.hour * 3600 + begin_time.minute * 60 + begin_time.second
                    end_seconds = end_time.hour * 3600 + end_time.minute * 60 + end_time.second

                    # Calculate absolute distance
                    distance_to_begin = abs(single_seconds - begin_seconds)
                    distance_to_end = abs(single_seconds - end_seconds)

                    # If closer to begin time, it's a check-in; if closer to end time, it's a check-out
                    if distance_to_begin < distance_to_end:
                        first_checkin = formatted_time
                        last_checkout = ""
                    else:
                        first_checkin = ""
                        last_checkout = formatted_time
                else:
                    # No shift info - use simple time-based logic (before 12:00 = check-in, after 12:00 = check-out)
                    noon_seconds = 12 * 3600  # 12:00:00
                    single_seconds = single_time_only.hour * 3600 + single_time_only.minute * 60 + single_time_only.second

                    if single_seconds < noon_seconds:
                        first_checkin = formatted_time
                        last_checkout = ""
                    else:
                        first_checkin = ""
                        last_checkout = formatted_time
        else:
            # Rule 2 & 3: Multiple checkins (>=2) - always show first as check-in and last as check-out
            first_checkin_time = emp.get("first_check_in")
            if first_checkin_time:
                if isinstance(first_checkin_time, str):
                    first_checkin_time = get_datetime(first_checkin_time)
                first_checkin = first_checkin_time.strftime("%H:%M:%S %d/%m/%Y")

            last_checkout_time = emp.get("last_check_out")
            if last_checkout_time:
                if isinstance(last_checkout_time, str):
                    last_checkout_time = get_datetime(last_checkout_time)
                last_checkout = last_checkout_time.strftime("%H:%M:%S %d/%m/%Y")

        # Get checkin date and format it
        checkin_date = emp.get('checkin_date')
        checkin_date_formatted = formatdate(checkin_date, "dd/MM/yyyy") if checkin_date else ""

        # Get manual check-in info
        manual_checkins = emp.get('manual_checkins', '')
        reason_for_manual = emp.get('reason_for_manual', '')
        other_reason_for_manual = emp.get('other_reason_for_manual', '')

        row_data = [
            idx,
            checkin_date_formatted,
            emp.get('attendance_device_id') or '',
            emp.get('employee_code') or '',
            emp.get('employee_name') or '',
            emp.get('department') or '',
            emp.get('custom_group') or '',
            emp.get('shift') or '',
            emp.get('designation') or '',
            first_checkin,
            last_checkout,
            checkin_count,
            manual_checkins,
            reason_for_manual,
            other_reason_for_manual
        ]
        ws3.append(row_data)

        # Apply styles to data cells
        for col_num in range(1, len(headers_incomplete) + 1):
            cell = ws3.cell(row=idx + 1, column=col_num)
            cell.font = cell_font
            # Center align: STT (1), Ngày (2), Số lần chấm (12), Đã xử lý (13)
            cell.alignment = center_alignment if col_num in [1, 2, 12, 13] else cell_alignment
            cell.border = border

    # Add Table format for sheet 3
    if len(incomplete_checkins) > 0:
        table3 = Table(displayName="missing_checkin", ref=f"A1:{openpyxl.utils.get_column_letter(len(headers_incomplete))}{len(incomplete_checkins) + 1}")
        style3 = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table3.tableStyleInfo = style3
        ws3.add_table(table3)

    # Save to temporary file
    temp_dir = tempfile.gettempdir()
    file_name = f"Attendance Report {formatdate(report_date, 'dd-MM-yyyy')}.xlsx"
    file_path = os.path.join(temp_dir, file_name)
    wb.save(file_path)

    return file_path, file_name
def get_current_frappe_site_name():
    """
    Returns the name of the current Frappe site.
    """
    if hasattr(frappe.local, 'site'):
        return frappe.local.site
    else:
        return None
# cmd test from web console
'''
  frappe.call({
      method: "customize_erpnext.customize_erpnext.report.daily_check_in_report.scheduler.send_daily_check_in_report",
      callback: function(r) {
          if (!r.exc) {
              frappe.msgprint("Email đã được gửi thành công!");
          } else {
              frappe.msgprint("Có lỗi xảy ra khi gửi email.");
          }
      }
  });

'''
#    Code structure 
'''
  1. Imports 
  2. send_daily_check_in_report() - Main function
  3. get_incomplete_checkins() - 3 rules logic kiểm tra
  4. calculate_statistics() - Tính thống kê
  5. generate_email_content() - HTML email với 3 tables
  6. generate_excel_report() - Excel với 3 sheets + Table format
'''