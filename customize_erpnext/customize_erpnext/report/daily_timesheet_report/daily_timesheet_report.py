# Copyright (c) 2025, IT Team - TIQN and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.utils import getdate, add_days, get_first_day, get_last_day, formatdate
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import calendar
import tempfile
import os

# Constant for decimal rounding precision
DECIMAL_ROUND_NUMBER = 2

def decimal_round(value, places=DECIMAL_ROUND_NUMBER):
	"""Làm tròn số sử dụng decimal để đồng nhất với JavaScript"""
	if value is None:
		return 0.0
	try:
		decimal_value = Decimal(str(value))
		# Tạo pattern cho số chữ số thập phân
		if places == 1:
			pattern = Decimal('0.1')
		elif places == 2:
			pattern = Decimal('0.01')
		else:
			pattern = Decimal('0.' + '0' * (places - 1) + '1')
		
		return float(decimal_value.quantize(pattern, rounding=ROUND_HALF_UP))
	except (ValueError, TypeError):
		return 0.0


def execute(filters=None):
	columns = get_columns(filters)
	data = get_data(filters)

	# Disable prepared report since this report only queries from DB
	return columns, data, None, None, None, True


def get_columns(filters=None):
	is_summary = filters and filters.get("summary")
	show_detail_columns = filters and filters.get("detail_columns")
	
	# Debug: Check if we're in single date mode and not summary mode
	is_single_date = filters and filters.get("date_type") == "Single Date"
	
	# Base columns - always shown
	columns = [
		{
			"fieldname": "employee",
			"label": _("Employee"),
			"fieldtype": "Link",
			"options": "Employee",
			"width": 110
		},
		{
			"fieldname": "employee_name",
			"label": _("Employee Name"),
			"fieldtype": "Data",
			"width": 170
		},
		{
			"fieldname": "custom_group",
			"label": _("Group"),
			"fieldtype": "Data",
			"width": 110
		}
	]
	
	# Add date column only if not summary
	if not is_summary:
		columns.extend([
			{
				"fieldname": "attendance_date",
				"label": _("Attendance Date"),
				"fieldtype": "Date",
				"width": 100
			},
			{
				"fieldname": "shift",
				"label": _("Shift"),
				"fieldtype": "Link",
				"options": "Shift Type",
				"width": 70
			}
		])
	
	# Department, Section columns - only when Detail Columns checked
	if show_detail_columns:
		columns.extend([
			{
				"fieldname": "department",
				"label": _("Department"),
				"fieldtype": "Link",
				"options": "Department",
				"width": 120
			},
			{
				"fieldname": "custom_section",
				"label": _("Section"),
				"fieldtype": "Data",
				"width": 100
			}
		])
	
	# Check in/out columns - always shown when not summary mode
	if not is_summary:
		columns.extend([
			{
				"fieldname": "check_in",
				"label": _("Check In"),
				"fieldtype": "Data",
				"width": 90
			},
			{
				"fieldname": "check_out",
				"label": _("Check Out"),
				"fieldtype": "Data",
				"width": 90
			}
		])
	
	# Core hour columns - always shown
	columns.extend([
		{
			"fieldname": "working_hours",
			"label": _("Working Hours"),
			"fieldtype": "Float",
			"precision": DECIMAL_ROUND_NUMBER,
			"width": 100
		}
	])
	
	# Working Day column - only in summary mode or when Detail Columns is checked
	if is_summary or show_detail_columns:
		columns.append({
			"fieldname": "working_day",
			"label": _("Working Days"),
			"fieldtype": "Float",
			"precision": 2,
			"width": 100,
			"description": _("Working Hours / 8")
		})
	
	# Basic overtime columns - always shown
	columns.extend([
		{
			"fieldname": "actual_overtime",
			"label": _("Actual OT"),
			"fieldtype": "Float",
			"precision": DECIMAL_ROUND_NUMBER,
			"width": 90
		},
		{
			"fieldname": "approved_overtime",
			"label": _("Registered OT"),
			"fieldtype": "Float",
			"precision": DECIMAL_ROUND_NUMBER,
			"width": 100
		},
		{
			"fieldname": "overtime_hours",
			"label": _("Final OT"),
			"fieldtype": "Float",
			"precision": DECIMAL_ROUND_NUMBER,
			"width": 90
		}
	])
	
	# Detail columns controlled by Detail Columns checkbox
	if show_detail_columns:
		columns.extend([
			{
				"fieldname": "overtime_coefficient",
				"label": _("Overtime Coefficient"),
				"fieldtype": "Float",
				"precision": DECIMAL_ROUND_NUMBER,
				"width": 130
			},
			{
				"fieldname": "final_ot_with_coefficient",
				"label": _("Final OT - With Coefficient"),
				"fieldtype": "Float",
				"precision": DECIMAL_ROUND_NUMBER,
				"width": 160
			}
		])
		
		# Status columns - only when Detail Columns checked and not summary
		if not is_summary:
			columns.extend([
				{
					"fieldname": "late_entry",
					"label": _("Late Entry"),
					"fieldtype": "Check",
					"width": 80
				},
				{
					"fieldname": "early_exit",
					"label": _("Early Exit"),
					"fieldtype": "Check",
					"width": 80
				},
				{
					"fieldname": "maternity_benefit",
					"label": _("Maternity Benefit"),
					"fieldtype": "Check",
					"width": 110
				}
			])
	
	# Other detail columns - only when Detail Columns checked
	if show_detail_columns and not is_summary:
		columns.extend([
			{
				"fieldname": "shift_determined_by",
				"label": _("Shift Determined By"),
				"fieldtype": "Data",
				"width": 100
			}
		])
	
	# Employee date columns - only if detail_columns=true
	if show_detail_columns:
		columns.extend([
			{
				"fieldname": "date_of_joining",
				"label": _("Date of Joining"),
				"fieldtype": "Date",
				"width": 100
			},
			{
				"fieldname": "relieving_date",
				"label": _("Relieving Date"),
				"fieldtype": "Date",
				"width": 100
			}
		])
	
	# Status column - always shown  
	if not is_summary:
		columns.append({
			"fieldname": "status",
			"label": _("Status"),
			"fieldtype": "Data",
			"width": 110
		})
	
	# ID column - always shown at the end
	if not is_summary:
		columns.append({
			"fieldname": "name",
			"label": _("ID"),
			"fieldtype": "Link",
			"options": "Daily Timesheet",
			"width": 120
		})
	
	return columns


def get_data(filters):
	is_summary = filters and filters.get("summary")
	show_detail_columns = filters and filters.get("detail_columns")

	# Get date range conditions for filtering active employees
	date_conditions = get_date_range_conditions(filters)
	employee_conditions = get_employee_filter_conditions(filters)
	# Note: Status filter will be applied AFTER loading data, not in SQL query
	# because filtering in LEFT JOIN ON clause doesn't work as expected
	
	if is_summary:
		# Summary mode - Always show all active employees in period
		employee_fields = ""
		if show_detail_columns:
			employee_fields = ", emp.date_of_joining, emp.relieving_date, emp.designation"
		
		# Calculate period dates
		if filters.get("date_type") == "Single Date" and filters.get("single_date"):
			period_start = filters.get('single_date')
			period_end = filters.get('single_date')
		elif filters.get("date_type") == "Date Range":
			period_start = filters.get("from_date", "1900-01-01")
			period_end = filters.get("to_date", "2099-12-31")
		elif filters.get("date_type") == "Monthly" and filters.get("month") and filters.get("year"):
			month = int(filters.get("month"))
			year = int(filters.get("year"))
			if month == 1:
				prev_month = 12
				prev_year = year - 1
			else:
				prev_month = month - 1
				prev_year = year
			period_start = f"{prev_year}-{prev_month:02d}-26"
			period_end = f"{year}-{month:02d}-25"
		else:
			period_start = "1900-01-01"
			period_end = "2099-12-31"
		
		# Get all employees who were active during the period with LEFT JOIN to Daily Timesheet
		# Note: Don't filter by status here - will filter after loading
		data = frappe.db.sql(f"""
			SELECT
				emp.name as employee,
				emp.employee_name,
				emp.department,
				emp.custom_section,
				emp.custom_group,
				COALESCE(SUM(dt.working_hours), 0) as working_hours,
				COALESCE(SUM(dt.actual_overtime), 0) as actual_overtime,
				COALESCE(SUM(dt.approved_overtime), 0) as approved_overtime,
				COALESCE(SUM(dt.overtime_hours), 0) as overtime_hours,
				COALESCE(SUM(dt.working_hours + dt.overtime_hours), 0) as total_hours
				{employee_fields}
			FROM `tabEmployee` emp
			LEFT JOIN `tabDaily Timesheet` dt ON emp.name = dt.employee
				AND ({date_conditions if date_conditions else '1=1'})
			WHERE (
				-- Show employees who were active during any part of the period
				((emp.date_of_joining IS NULL OR emp.date_of_joining <= '{period_end}')
				 AND (emp.relieving_date IS NULL OR emp.relieving_date >= '{period_start}'))
			)
			-- Exclude Inactive employees
			AND emp.status IN ('Active', 'Left')
			{employee_conditions}
			GROUP BY emp.name, emp.employee_name, emp.department, emp.custom_section, emp.custom_group
			{', emp.date_of_joining, emp.relieving_date, emp.designation' if show_detail_columns else ''}
			ORDER BY emp.name
		""", filters, as_dict=1)
		
		# Apply decimal rounding to summary data and calculate working_day + coefficients
		for row in data:
			working_hours = decimal_round(row.get('working_hours'))
			row['working_hours'] = working_hours
			row['working_day'] = decimal_round(working_hours / 8, 2) if working_hours else 0.0
			row['actual_overtime'] = decimal_round(row.get('actual_overtime'))
			row['approved_overtime'] = decimal_round(row.get('approved_overtime'))
			row['overtime_hours'] = decimal_round(row.get('overtime_hours'))
			row['total_hours'] = decimal_round(row.get('total_hours'))
			
			# Calculate final_ot_with_coefficient by summing individual daily calculations
			employee = row.get('employee')
			if employee:
				# Get all records for this employee in the date range to calculate coefficient total
				individual_records = frappe.db.sql(f"""
					SELECT dt.overtime_hours, dt.overtime_coefficient, dt.final_ot_with_coefficient
					FROM `tabDaily Timesheet` dt
					WHERE dt.employee = %(employee)s AND ({date_conditions if date_conditions else '1=1'})
				""", dict(filters, employee=employee), as_dict=1)
				
				total_final_ot_with_coeff = sum([
					decimal_round((r.get('overtime_hours') or 0) * (r.get('overtime_coefficient') or 1.5))
					for r in individual_records
				])
				
				row['final_ot_with_coefficient'] = decimal_round(total_final_ot_with_coeff)
				
				# For coefficient, show average (for display purposes)
				if individual_records:
					avg_coefficient = sum([r.get('overtime_coefficient') or 1.5 for r in individual_records]) / len(individual_records)
					row['overtime_coefficient'] = decimal_round(avg_coefficient)
				else:
					row['overtime_coefficient'] = 1.5
	else:
		# Detail mode - Always show all active employees with individual daily records
		
		# Calculate period dates (same logic as summary mode)
		if filters.get("date_type") == "Single Date" and filters.get("single_date"):
			period_start = filters.get('single_date')
			period_end = filters.get('single_date')
		elif filters.get("date_type") == "Date Range":
			period_start = filters.get("from_date", "1900-01-01")
			period_end = filters.get("to_date", "2099-12-31")
		elif filters.get("date_type") == "Monthly" and filters.get("month") and filters.get("year"):
			month = int(filters.get("month"))
			year = int(filters.get("year"))
			if month == 1:
				prev_month = 12
				prev_year = year - 1
			else:
				prev_month = month - 1
				prev_year = year
			period_start = f"{prev_year}-{prev_month:02d}-26"
			period_end = f"{year}-{month:02d}-25"
		else:
			period_start = "1900-01-01"
			period_end = "2099-12-31"
		
		# Get all employees who were active during the period with all their Daily Timesheet records
		# Use LEFT JOIN to show employees even if they don't have timesheet data for some dates
		# Note: Don't filter by status in query - will filter after loading
		data = frappe.db.sql(f"""
			SELECT
				emp.name as employee,
				emp.employee_name,
				emp.department,
				emp.custom_section,
				emp.custom_group,
				dt.attendance_date,
				dt.shift,
				dt.shift_determined_by,
				dt.check_in,
				dt.check_out,
				COALESCE(dt.working_hours, 0) as working_hours,
				COALESCE(dt.actual_overtime, 0) as actual_overtime,
				COALESCE(dt.approved_overtime, 0) as approved_overtime,
				COALESCE(dt.overtime_hours, 0) as overtime_hours,
				COALESCE(dt.overtime_coefficient, 1.5) as overtime_coefficient,
				COALESCE(dt.final_ot_with_coefficient, 0) as final_ot_with_coefficient,
				COALESCE(dt.working_hours + dt.overtime_hours, 0) as total_hours,
				dt.late_entry,
				dt.early_exit,
				dt.maternity_benefit,
				dt.status,
				emp.date_of_joining,
				emp.relieving_date,
				emp.designation,
				dt.name
			FROM `tabEmployee` emp
			LEFT JOIN `tabDaily Timesheet` dt ON emp.name = dt.employee
				AND (dt.docstatus <= 1)
				AND ({date_conditions if date_conditions else '1=1'})
			WHERE (
				-- Show employees who were active during any part of the period
				((emp.date_of_joining IS NULL OR emp.date_of_joining <= '{period_end}')
				 AND (emp.relieving_date IS NULL OR emp.relieving_date >= '{period_start}'))
			)
			-- Exclude Inactive employees
			AND emp.status IN ('Active', 'Left')
			{employee_conditions}
			-- Only show records where there's either timesheet data OR we want to show all active employees
			AND (dt.attendance_date IS NOT NULL)
			ORDER BY dt.attendance_date, emp.name
		""", filters, as_dict=1)
		
		# Add missing employee-date combinations for active employees without timesheet records
		if data:
			# Get all active employees in the period
			active_employees = frappe.db.sql(f"""
				SELECT DISTINCT emp.name as employee,
					emp.employee_name,
					emp.department,
					emp.custom_section,
					emp.custom_group,
					emp.date_of_joining,
					emp.relieving_date,
					emp.designation
				FROM `tabEmployee` emp
				WHERE (
					-- Show employees who were active during any part of the period
					((emp.date_of_joining IS NULL OR emp.date_of_joining <= '{period_end}')
					 AND (emp.relieving_date IS NULL OR emp.relieving_date >= '{period_start}'))
				)
				-- Exclude Inactive employees
				AND emp.status IN ('Active', 'Left')
				{employee_conditions}
			""", filters, as_dict=1)
			
			# Generate all dates in the period
			from datetime import datetime, timedelta
			if filters.get("date_type") == "Single Date" and filters.get("single_date"):
				start_date = datetime.strptime(period_start, '%Y-%m-%d').date()
				end_date = datetime.strptime(period_end, '%Y-%m-%d').date()
			elif filters.get("date_type") == "Date Range":
				start_date = datetime.strptime(period_start, '%Y-%m-%d').date()
				end_date = datetime.strptime(period_end, '%Y-%m-%d').date()
			elif filters.get("date_type") == "Monthly":
				start_date = datetime.strptime(period_start, '%Y-%m-%d').date()
				end_date = datetime.strptime(period_end, '%Y-%m-%d').date()
			else:
				start_date = datetime.strptime(period_start, '%Y-%m-%d').date()
				end_date = datetime.strptime(period_end, '%Y-%m-%d').date()
			
			all_dates = []
			current_date = start_date
			while current_date <= end_date:
				all_dates.append(current_date)
				current_date += timedelta(days=1)
			
			# Check which employee-date combinations are missing
			existing_records = set()
			for row in data:
				if row.get('attendance_date'):
					emp_date_key = f"{row['employee']}-{row['attendance_date']}"
					existing_records.add(emp_date_key)
			
			# Add missing records with empty timesheet data
			missing_records = []
			for emp in active_employees:
				# Parse employee dates ONCE per employee (not per date iteration) - Performance optimization
				emp_start = emp.get('date_of_joining')
				emp_end = emp.get('relieving_date')

				# Convert to date objects if they're strings
				if emp_start and isinstance(emp_start, str):
					emp_start = datetime.strptime(emp_start, '%Y-%m-%d').date()
				if emp_end and isinstance(emp_end, str):
					emp_end = datetime.strptime(emp_end, '%Y-%m-%d').date()

				for date_obj in all_dates:
					emp_date_key = f"{emp['employee']}-{date_obj}"
					if emp_date_key not in existing_records:
						# Check if employee was active on this date
						is_active_on_date = True
						if emp_start and date_obj < emp_start:
							is_active_on_date = False
						if emp_end and date_obj > emp_end:
							is_active_on_date = False
						
						if is_active_on_date:
							missing_records.append({
								'employee': emp['employee'],
								'employee_name': emp['employee_name'],
								'department': emp['department'],
								'custom_section': emp['custom_section'],
								'custom_group': emp['custom_group'],
								'attendance_date': date_obj,
								'shift': None,
								'shift_determined_by': None,
								'check_in': None,
								'check_out': None,
								'working_hours': 0,
								'actual_overtime': 0,
								'approved_overtime': 0,
								'overtime_hours': 0,
								'overtime_coefficient': 1.5,
								'final_ot_with_coefficient': 0,
								'total_hours': 0,
								'late_entry': None,
								'early_exit': None,
								'maternity_benefit': None,
								'status': None,  # Leave empty for missing records
								'date_of_joining': emp['date_of_joining'],
								'relieving_date': emp['relieving_date'],
								'designation': emp['designation'],
								'name': None
							})
			
			# Add missing records to data
			data.extend(missing_records)
			
			# Sort again by date and employee
			data.sort(key=lambda x: (x.get('attendance_date') or datetime.min.date(), x.get('employee') or ''))
		
		# Apply decimal rounding to detail data and format time
		for row in data:
			working_hours = decimal_round(row.get('working_hours'))
			row['working_hours'] = working_hours
			row['actual_overtime'] = decimal_round(row.get('actual_overtime'))
			row['approved_overtime'] = decimal_round(row.get('approved_overtime'))
			row['overtime_hours'] = decimal_round(row.get('overtime_hours'))
			row['overtime_coefficient'] = decimal_round(row.get('overtime_coefficient'))
			
			# Calculate working_day when Detail Columns is checked
			if show_detail_columns:
				row['working_day'] = decimal_round(working_hours / 8, 2) if working_hours else 0.0
			
			# Calculate final_ot_with_coefficient for individual records
			overtime_hours = row.get('overtime_hours') or 0
			overtime_coefficient = row.get('overtime_coefficient') or 1.5
			row['final_ot_with_coefficient'] = decimal_round(overtime_hours * overtime_coefficient)
			
			row['total_hours'] = decimal_round(row.get('total_hours'))
			
			# Format check_in and check_out to show only time (HH:MM:SS)
			if row.get('check_in'):
				try:
					if hasattr(row['check_in'], 'strftime'):
						row['check_in'] = row['check_in'].strftime('%H:%M:%S')
					else:
						# Handle string datetime format
						check_in_str = str(row['check_in'])
						if ' ' in check_in_str:
							row['check_in'] = check_in_str.split(' ')[1]  # Get time part
						elif ':' in check_in_str:
							row['check_in'] = check_in_str  # Already time format
				except:
					pass  # Keep original value if formatting fails
					
			if row.get('check_out'):
				try:
					if hasattr(row['check_out'], 'strftime'):
						row['check_out'] = row['check_out'].strftime('%H:%M:%S')
					else:
						# Handle string datetime format
						check_out_str = str(row['check_out'])
						if ' ' in check_out_str:
							row['check_out'] = check_out_str.split(' ')[1]  # Get time part
						elif ':' in check_out_str:
							row['check_out'] = check_out_str  # Already time format
				except:
					pass  # Keep original value if formatting fails

	return data






def get_date_range_conditions(filters):
	"""Get SQL conditions for date range based on filter type"""
	conditions = []

	if filters.get("date_type") == "Single Date":
		if filters.get("single_date"):
			return f"dt.attendance_date = '{filters.get('single_date')}'"
	elif filters.get("date_type") == "Date Range":
		if filters.get("from_date"):
			conditions.append(f"dt.attendance_date >= '{filters.get('from_date')}'")
		if filters.get("to_date"):
			conditions.append(f"dt.attendance_date <= '{filters.get('to_date')}'")
		return " AND ".join(conditions) if conditions else "1=1"
	elif filters.get("date_type") == "Monthly":
		if filters.get("month") and filters.get("year"):
			month = int(filters.get("month"))
			year = int(filters.get("year"))

			# Calculate previous month for monthly range (26th to 25th)
			if month == 1:
				prev_month = 12
				prev_year = year - 1
			else:
				prev_month = month - 1
				prev_year = year

			from_date = f"{prev_year}-{prev_month:02d}-26"
			to_date = f"{year}-{month:02d}-25"

			return f"dt.attendance_date >= '{from_date}' AND dt.attendance_date <= '{to_date}'"

	return None


def get_employee_filter_conditions(filters):
	"""Get SQL conditions for employee filtering"""
	conditions = []

	# Always exclude specific departments
	conditions.append("emp.department NOT IN ('Head of Branch - TIQN', 'Operations Manager - TIQN')")

	if filters.get("department"):
		if isinstance(filters.get("department"), list):
			dept_list = "', '".join(filters.get("department"))
			conditions.append(f"emp.department IN ('{dept_list}')")
		else:
			conditions.append(f"emp.department = '{filters.get('department')}'")

	if filters.get("custom_section"):
		section_value = filters.get("custom_section")
		# MultiSelectList returns comma-separated string, not list
		if isinstance(section_value, str) and ',' in section_value:
			# Multiple values selected
			section_list = [s.strip() for s in section_value.split(',') if s.strip() and s.strip() != 'undefined']
			if section_list:
				section_str = "', '".join(section_list)
				conditions.append(f"emp.custom_section IN ('{section_str}')")
		elif isinstance(section_value, list):
			# Just in case it's a list
			section_list = [s for s in section_value if s and s != 'undefined']
			if section_list:
				section_str = "', '".join(section_list)
				conditions.append(f"emp.custom_section IN ('{section_str}')")
		else:
			# Single value
			if section_value and section_value != 'undefined':
				conditions.append(f"emp.custom_section = '{section_value}'")

	if filters.get("custom_group"):
		group_value = filters.get("custom_group")
		# MultiSelectList returns comma-separated string, not list
		if isinstance(group_value, str) and ',' in group_value:
			# Multiple values selected
			group_list = [g.strip() for g in group_value.split(',') if g.strip() and g.strip() != 'undefined']
			if group_list:
				group_str = "', '".join(group_list)
				conditions.append(f"emp.custom_group IN ('{group_str}')")
		elif isinstance(group_value, list):
			# Just in case it's a list
			group_list = [g for g in group_value if g and g != 'undefined']
			if group_list:
				group_str = "', '".join(group_list)
				conditions.append(f"emp.custom_group IN ('{group_str}')")
		else:
			# Single value
			if group_value and group_value != 'undefined':
				conditions.append(f"emp.custom_group = '{group_value}'")

	if filters.get("employee"):
		conditions.append(f"emp.name = '{filters.get('employee')}'")

	return " AND " + " AND ".join(conditions) if conditions else ""





def get_chart_data(data, filters):
	"""Generate chart data for the report"""
	if not data:
		return None
	
	# Group by department for chart
	dept_data = {}
	for row in data:
		dept = row.get("department") or "No Department"
		if dept not in dept_data:
			dept_data[dept] = {
				"working_hours": 0,
				"overtime_hours": 0,
				"total_hours": 0
			}
		
		dept_data[dept]["working_hours"] += row.get("working_hours", 0)
		dept_data[dept]["overtime_hours"] += row.get("overtime_hours", 0)
		dept_data[dept]["total_hours"] += row.get("total_hours", 0)
	
	# Apply decimal rounding to chart data
	for dept in dept_data:
		dept_data[dept]["working_hours"] = decimal_round(dept_data[dept]["working_hours"])
		dept_data[dept]["overtime_hours"] = decimal_round(dept_data[dept]["overtime_hours"])
		dept_data[dept]["total_hours"] = decimal_round(dept_data[dept]["total_hours"])
	
	return {
		"data": {
			"labels": list(dept_data.keys()),
			"datasets": [
				{
					"name": "Working Hours",
					"values": [dept_data[dept]["working_hours"] for dept in dept_data.keys()]
				},
				{
					"name": "Overtime Hours", 
					"values": [dept_data[dept]["overtime_hours"] for dept in dept_data.keys()]
				}
			]
		},
		"type": "bar",
		"height": 150
	}

@frappe.whitelist()
def export_timesheet_excel(filters=None, report_data=None):
	"""Export timesheet data to Excel format similar to the template"""
	import json
	if isinstance(filters, str):
		filters = json.loads(filters)
	if isinstance(report_data, str) and report_data:
		report_data = json.loads(report_data)
	
	# Always use fresh data from the report to ensure consistency
	# Force summary = 0 for Excel export to get daily detail data
	export_filters = filters.copy() if filters else {}
	export_filters['summary'] = 0
	columns, data, message, chart, skip_total_row, disable_prepared_report = execute(export_filters)
	employee_data = convert_report_data_to_excel_format(data, export_filters)
	
	# Create Excel file
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Timesheet"
	
	# Get date range based on current filters
	date_range = get_export_date_range(filters)
	
	# Setup Excel headers and data
	setup_excel_headers(ws, date_range, filters)
	all_employees, current_row, first_employee_row = populate_employee_data(ws, employee_data, date_range)
	
	# Add total row (only if we have employees)
	if first_employee_row is not None:
		add_total_row(ws, all_employees, current_row, date_range, first_employee_row)
	total_row_end = current_row + 1
	
	# Add footer with signatures and notes
	add_excel_footer(ws, total_row_end, date_range)
	
	# Apply formatting (account for department headers + total row)
	total_rows = len(all_employees) + len(employee_data) + 1  # employees + dept headers + total row
	apply_excel_formatting(ws, total_rows, len(date_range['dates']), date_range)
	
	# Save to temporary file
	temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
	wb.save(temp_file.name)
	temp_file.close()
	
	# Read file content and encode as base64
	import base64
	with open(temp_file.name, 'rb') as f:
		file_content = base64.b64encode(f.read()).decode('utf-8')
	
	# Clean up
	os.unlink(temp_file.name)
	
	# Create filename
	filename = f"Timesheet {date_range['period_name']}.xlsx"
	
	return {
		'filecontent': file_content,
		'filename': filename,
		'type': 'binary'
	}

def get_export_date_range(filters):
	"""Get date range based on current report filters"""
	if filters.get("date_type") == "Single Date":
		if filters.get("single_date"):
			single_date = getdate(filters.get("single_date"))
			from_date = single_date
			to_date = single_date
		else:
			# Default to today
			from_date = to_date = getdate()
	elif filters.get("date_type") == "Date Range":
		from_date = getdate(filters.get("from_date")) if filters.get("from_date") else getdate()
		to_date = getdate(filters.get("to_date")) if filters.get("to_date") else getdate()
	elif filters.get("date_type") == "Monthly":
		month = int(filters.get("month", 1))
		year = int(filters.get("year", 2025))
		
		# Calculate previous month for salary period (26th to 25th)
		if month == 1:
			prev_month = 12
			prev_year = year - 1
		else:
			prev_month = month - 1
			prev_year = year
		
		from_date = datetime(prev_year, prev_month, 26).date()
		to_date = datetime(year, month, 25).date()
	else:
		# Default to current month
		from_date = get_first_day(getdate())
		to_date = get_last_day(getdate())
	
	# Generate all dates in range
	dates = []
	current_date = from_date
	while current_date <= to_date:
		dates.append(current_date)
		current_date = add_days(current_date, 1)
	
	# Create display info with bilingual format
	if filters.get("date_type") == "Monthly":
		month_num = int(filters.get("month", 1))
		year = int(filters.get("year", 2025))
		month_name_en = calendar.month_name[month_num][:3]  # Short English month name
		vietnamese_months = {
			1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",
			7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"
		}
		month_vn = vietnamese_months[month_num]
		period_name = f"{month_name_en} {year}/ Tháng {month_vn} năm {year}"
	else:
		period_name = f"{formatdate(from_date)} - {formatdate(to_date)}"
	
	return {
		'from_date': from_date,
		'to_date': to_date,
		'dates': dates,
		'period_name': period_name,
		'date_type': filters.get("date_type", "Date Range")
	}


def clean_department_name(dept_name):
	"""Remove '- TIQN' suffix from department names"""
	if dept_name and dept_name.endswith(' - TIQN'):
		return dept_name.replace(' - TIQN', '')
	return dept_name or ''

def convert_report_data_to_excel_format(report_data, filters):
	"""Convert report data to Excel format for timesheet export"""
	if not report_data:
		return []
	
	# Get date range for columns
	date_range = get_export_date_range(filters)
	
	# Group data by employee
	employee_data = {}
	for row in report_data:
		employee_id = row.get('employee')
		if not employee_id:
			continue
			
		if employee_id not in employee_data:
			employee_data[employee_id] = {
				'employee_id': employee_id,
				'employee_name': row.get('employee_name', '') or '',
				'date_of_joining': row.get('date_of_joining'),
				'relieving_date': row.get('relieving_date'),
				'custom_group': row.get('custom_group', '') or '',
				'custom_section': row.get('custom_section', '') or '',
				'department': clean_department_name(row.get('department', '') or ''),
				'designation': row.get('designation', '') or '',
				'daily_data': {}
			}
		
		# Add daily attendance data
		attendance_date = row.get('attendance_date')
		if attendance_date:
			# Handle different date formats
			if hasattr(attendance_date, 'strftime'):
				date_key = attendance_date.strftime('%Y-%m-%d')
			else:
				date_key = str(attendance_date)
			
			# Convert working hours to working days
			working_hours = row.get('working_hours', 0) or 0
			status = row.get('status', '')
			
			# Always show all data including zeros (show_zero filter removed)
			# Use decimal rounding to match report calculations
			if working_hours > 0:
				working_days = decimal_round(working_hours / 8, 2)
				employee_data[employee_id]['daily_data'][date_key] = working_days
			elif status == 'Present':
				# Present but no working hours - show 0
				employee_data[employee_id]['daily_data'][date_key] = 0
			else:
				# No working day - leave blank for other statuses or missing data
				employee_data[employee_id]['daily_data'][date_key] = ''
	
	# Ensure all active employees in the period are included (for summary mode or missing employees)
	# Get all active employees for the period
	if filters.get("date_type") == "Single Date" and filters.get("single_date"):
		period_start = filters.get('single_date')
		period_end = filters.get('single_date')
	elif filters.get("date_type") == "Date Range":
		period_start = filters.get("from_date", "1900-01-01")
		period_end = filters.get("to_date", "2099-12-31")
	elif filters.get("date_type") == "Monthly" and filters.get("month") and filters.get("year"):
		month = int(filters.get("month"))
		year = int(filters.get("year"))
		if month == 1:
			prev_month = 12
			prev_year = year - 1
		else:
			prev_month = month - 1
			prev_year = year
		period_start = f"{prev_year}-{prev_month:02d}-26"
		period_end = f"{year}-{month:02d}-25"
	else:
		period_start = "1900-01-01"
		period_end = "2099-12-31"
	
	# Get employee conditions for filtering
	employee_conditions = get_employee_filter_conditions(filters)

	# Get all active employees that should be included
	all_active_employees = frappe.db.sql(f"""
		SELECT DISTINCT emp.name as employee,
			emp.employee_name,
			emp.department,
			emp.custom_section,
			emp.custom_group,
			emp.date_of_joining,
			emp.relieving_date,
			emp.designation
		FROM `tabEmployee` emp
		WHERE (
			-- Show employees who were active during any part of the period
			((emp.date_of_joining IS NULL OR emp.date_of_joining <= '{period_end}')
			 AND (emp.relieving_date IS NULL OR emp.relieving_date >= '{period_start}'))
		)
		-- Exclude Inactive employees
		AND emp.status IN ('Active', 'Left')
		{employee_conditions}
	""", filters, as_dict=1)
	
	# Add missing employees to employee_data
	for emp in all_active_employees:
		if emp.employee not in employee_data:
			employee_data[emp.employee] = {
				'employee_id': emp.employee,
				'employee_name': emp.employee_name or '',
				'date_of_joining': emp.date_of_joining,
				'relieving_date': emp.relieving_date,
				'custom_group': emp.custom_group or '',
				'custom_section': emp.custom_section or '',
				'department': clean_department_name(emp.department or ''),
				'designation': emp.designation or '',
				'daily_data': {}
			}

	# Convert to list and sort by department, section, group, name
	result = list(employee_data.values())
	result.sort(key=lambda x: (
		x.get('department', '') or '',
		x.get('custom_section', '') or '',
		x.get('custom_group', '') or '',
		x.get('employee_name', '') or ''
	))
	
	# Group by department for Excel export
	grouped_result = []
	current_dept = None
	dept_employees = []
	
	for emp in result:
		# Ensure all employees have entries for all dates in range
		for date_obj in date_range['dates']:
			date_key = date_obj.strftime('%Y-%m-%d')
			if date_key not in emp['daily_data']:
				emp['daily_data'][date_key] = ''
		
		emp_dept = emp.get('department', '') or 'No Department'
		
		if current_dept != emp_dept:
			# Add previous department group
			if current_dept is not None and dept_employees:
				grouped_result.append({
					'type': 'department_header',
					'department': current_dept,
					'employees': dept_employees
				})
			
			# Start new department group
			current_dept = emp_dept
			dept_employees = [emp]
		else:
			dept_employees.append(emp)
	
	# Add the last department group
	if current_dept is not None and dept_employees:
		grouped_result.append({
			'type': 'department_header',
			'department': current_dept,
			'employees': dept_employees
		})
	
	return grouped_result

def setup_excel_headers(ws, date_range, filters):
	"""Setup Excel headers similar to template"""
	# Calculate total columns needed - Fixed calculation
	total_cols = 8 + len(date_range['dates']) + 2  # Basic columns (8) + dates + total + confirmation
	
	# Row 1: Main title (merge and center)
	ws.cell(row=1, column=1, value="EMPLOYEE TIMESHEET")
	ws.cell(row=1, column=1).font = Font(name='Times New Roman', size=12, bold=True)
	ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
	
	# Row 2: Vietnamese title (merge and center)
	ws.cell(row=2, column=1, value="BẢNG CHẤM CÔNG NHÂN VIÊN")
	ws.cell(row=2, column=1).font = Font(name='Times New Roman', size=12, bold=True)
	ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
	
	# Row 3: Period and date range (merge and center)
	from_date_str = date_range['from_date'].strftime('%d %b %y')
	to_date_str = date_range['to_date'].strftime('%d %b %y')
	ws.cell(row=3, column=1, value=f"{date_range['period_name']} ({from_date_str} - {to_date_str})")
	ws.cell(row=3, column=1).font = Font(name='Times New Roman', size=10)
	ws.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')
	ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
	
	# Row 5: Column headers
	headers = [
		"No./ STT", 
		"Employee ID /Mã NV", 
		"Full name/Họ tên", 
		"Ngày kí hợp đồng",
		"Resign on/ Nghỉ việc", 
		"Line/ Chuyền", 
		"Section/ Bộ phận", 
		"Position/Chức vụ",
		"DATE IN THE MONTH/ NGÀY TRONG THÁNG",
		"", # Placeholder for merged cells
		"Xác nhận",
		"Total working days/Tổng ngày công"
	]
	
	# Set fixed headers (columns 1-8) and merge with rows 6-7
	for i, header in enumerate(headers[:8], 1):
		cell = ws.cell(row=5, column=i, value=header)
		cell.font = Font(name='Times New Roman', bold=True, size=10)
		cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		# Merge each header cell with the 2 rows below (rows 6 and 7)
		ws.merge_cells(start_row=5, start_column=i, end_row=7, end_column=i)
	
	# Set last two headers (swapped: total first, then confirmation)
	total_col = 9 + len(date_range['dates'])
	confirmation_col = total_col + 1
	
	ws.cell(row=5, column=total_col, value="Total working days/Tổng ngày công")
	ws.cell(row=5, column=total_col).font = Font(name='Times New Roman', bold=True, size=10)
	ws.cell(row=5, column=total_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
	# Merge total column with rows 6-7
	ws.merge_cells(start_row=5, start_column=total_col, end_row=7, end_column=total_col)
	
	ws.cell(row=5, column=confirmation_col, value="Xác nhận")
	ws.cell(row=5, column=confirmation_col).font = Font(name='Times New Roman', bold=True, size=10)
	ws.cell(row=5, column=confirmation_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
	# Merge confirmation column with rows 6-7
	ws.merge_cells(start_row=5, start_column=confirmation_col, end_row=7, end_column=confirmation_col)
	
	# Merge and set "DATE IN THE MONTH/ NGÀY TRONG THÁNG" header
	if len(date_range['dates']) > 0:
		date_start_col = 9
		date_end_col = 8 + len(date_range['dates'])
		ws.cell(row=5, column=date_start_col, value="DATE IN THE MONTH/ NGÀY TRONG THÁNG")
		ws.cell(row=5, column=date_start_col).font = Font(name='Times New Roman', bold=True, size=10)
		ws.cell(row=5, column=date_start_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		ws.merge_cells(start_row=5, start_column=date_start_col, end_row=5, end_column=date_end_col)
	
	# Row 6: Dates (only day number)
	for i, date_obj in enumerate(date_range['dates']):
		col = 9 + i  # Start from column 9
		ws.cell(row=6, column=col, value=date_obj.day)  # Only day number
		cell = ws.cell(row=6, column=col)
		cell.font = Font(name='Times New Roman', size=9)
		cell.alignment = Alignment(horizontal='center')
	
	# Row 7: Day of week
	for i, date_obj in enumerate(date_range['dates']):
		col = 9 + i
		day_of_week = date_obj.strftime('%a')[:1]  # First letter of day
		ws.cell(row=7, column=col, value=day_of_week)
		cell = ws.cell(row=7, column=col)
		cell.font = Font(name='Times New Roman', size=9, bold=True)
		cell.alignment = Alignment(horizontal='center')

def populate_employee_data(ws, employee_data, date_range):
	"""Populate employee data into Excel with department grouping"""
	start_row = 8
	current_row = start_row
	total_columns = 9 + len(date_range['dates']) + 2  # Basic + dates + total + confirmation
	
	# Create light gray fill for department headers
	dept_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
	
	all_employees = []  # Keep track of all employees for total calculation
	stt_counter = 1  # Continuous numbering across departments
	first_employee_row = None  # Track first employee row for SUM formulas
	special_rows = []  # Track department header rows (no borders)
	
	for dept_group in employee_data:
		if dept_group.get('type') == 'department_header':
			# Add department header row - only fill first cell, gray background for entire row
			dept_name = dept_group['department']
			ws.cell(row=current_row, column=1, value=dept_name)
			dept_cell = ws.cell(row=current_row, column=1)
			dept_cell.font = Font(name='Times New Roman', bold=True, size=10)
			dept_cell.alignment = Alignment(horizontal='left', vertical='center')
			
			# Apply gray background to entire row (no borders)
			for col in range(1, total_columns):
				cell = ws.cell(row=current_row, column=col)
				cell.fill = dept_fill
			
			current_row += 1
			
			# Add employees in this department
			for emp in dept_group['employees']:
				all_employees.append(emp)  # Track for totals
				# Track first employee row for formulas
				if first_employee_row is None:
					first_employee_row = current_row
				populate_single_employee_row(ws, emp, current_row, date_range, stt_counter)
				stt_counter += 1
				current_row += 1
	
	return all_employees, current_row, first_employee_row

def populate_single_employee_row(ws, emp, row, date_range, stt_number):
	"""Populate a single employee row"""
	from datetime import datetime
	
	# Basic employee info
	ws.cell(row=row, column=1, value=stt_number)  # STT
	ws.cell(row=row, column=2, value=emp['employee_id'])
	ws.cell(row=row, column=3, value=emp['employee_name'])
	
	# Format dates as dd/mm/yyyy
	if emp['date_of_joining']:
		if hasattr(emp['date_of_joining'], 'strftime'):
			ws.cell(row=row, column=4, value=emp['date_of_joining'].strftime('%d/%m/%Y'))
		else:
			# Parse string date and format
			try:
				date_obj = datetime.strptime(str(emp['date_of_joining']), '%Y-%m-%d')
				ws.cell(row=row, column=4, value=date_obj.strftime('%d/%m/%Y'))
			except:
				ws.cell(row=row, column=4, value=str(emp['date_of_joining']))
	else:
		ws.cell(row=row, column=4, value='')
		
	if emp['relieving_date']:
		if hasattr(emp['relieving_date'], 'strftime'):
			ws.cell(row=row, column=5, value=emp['relieving_date'].strftime('%d/%m/%Y'))
		else:
			# Parse string date and format
			try:
				date_obj = datetime.strptime(str(emp['relieving_date']), '%Y-%m-%d')
				ws.cell(row=row, column=5, value=date_obj.strftime('%d/%m/%Y'))
			except:
				ws.cell(row=row, column=5, value=str(emp['relieving_date']))
	else:
		ws.cell(row=row, column=5, value='')
		
	ws.cell(row=row, column=6, value=emp['custom_group'])
	ws.cell(row=row, column=7, value=emp['custom_section']) 
	ws.cell(row=row, column=8, value=emp['designation'])  # Position/Chức vụ
	
	# Set font and alignment for employee data (size 8)
	for col in range(1, 9):  # Columns 1-8
		cell = ws.cell(row=row, column=col)
		cell.font = Font(name='Times New Roman', size=8)
		# Left-align these specific columns
		cell.alignment = Alignment(horizontal='left', vertical='center')
	
	# Daily attendance data
	daily_total = 0
	for j, date_obj in enumerate(date_range['dates']):
		col = 9 + j
		date_key = date_obj.strftime('%Y-%m-%d')
		value = emp['daily_data'].get(date_key, '')
		ws.cell(row=row, column=col, value=value)
		
		# Center align attendance data with font
		cell = ws.cell(row=row, column=col)
		cell.font = Font(name='Times New Roman', size=8)
		cell.alignment = Alignment(horizontal='center', vertical='center')
		
		# Sum for total working days
		if isinstance(value, (int, float)) and value > 0:
			daily_total += value
	
	# Add total working days column with SUM formula (now comes first)
	total_col = 9 + len(date_range['dates'])
	# Create SUM formula for this row's daily attendance data
	from openpyxl.utils import get_column_letter
	start_col_letter = get_column_letter(9)  # First date column
	end_col_letter = get_column_letter(8 + len(date_range['dates']))  # Last date column
	sum_formula = f"=SUM({start_col_letter}{row}:{end_col_letter}{row})"
	cell = ws.cell(row=row, column=total_col, value=sum_formula)
	cell.font = Font(name='Times New Roman', size=8)
	cell.alignment = Alignment(horizontal='center', vertical='center')
	
	# Add confirmation column (empty, now comes second)
	confirmation_col = total_col + 1
	cell = ws.cell(row=row, column=confirmation_col, value='')
	cell.font = Font(name='Times New Roman', size=8)
	cell.alignment = Alignment(horizontal='center', vertical='center')

def add_total_row(ws, all_employees, total_row, date_range, first_employee_row):
	"""Add total row at the end with SUM formulas"""
	from openpyxl.utils import get_column_letter
	from openpyxl.styles import PatternFill, Border, Side
	
	# Create light gray fill for total row
	total_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
	total_columns = 9 + len(date_range['dates']) + 2  # Basic + dates + total + confirmation
	
	ws.cell(row=total_row, column=1, value="TOTAL")
	ws.cell(row=total_row, column=1).font = Font(name='Times New Roman', bold=True, size=8)
	ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
	
	# Add SUM formulas for each date column
	last_employee_row = total_row - 1  # Row just before TOTAL row
	for j, date_obj in enumerate(date_range['dates']):
		col = 9 + j
		col_letter = get_column_letter(col)
		# Create SUM formula from first employee row to last employee row
		sum_formula = f"=SUM({col_letter}{first_employee_row}:{col_letter}{last_employee_row})"
		ws.cell(row=total_row, column=col, value=sum_formula)
		
		ws.cell(row=total_row, column=col).font = Font(name='Times New Roman', bold=True, size=8)
		ws.cell(row=total_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
	
	# Add grand total SUM formula in Total working days column
	total_col = 9 + len(date_range['dates'])
	total_col_letter = get_column_letter(total_col)
	# Sum all individual employee total working days
	grand_total_formula = f"=SUM({total_col_letter}{first_employee_row}:{total_col_letter}{last_employee_row})"
	ws.cell(row=total_row, column=total_col, value=grand_total_formula)
	ws.cell(row=total_row, column=total_col).font = Font(name='Times New Roman', bold=True, size=8)
	ws.cell(row=total_row, column=total_col).alignment = Alignment(horizontal='center', vertical='center')
	
	# Add gray background to entire TOTAL row (no borders)
	for col in range(1, total_columns):
		cell = ws.cell(row=total_row, column=col)
		cell.fill = total_fill

def add_excel_footer(ws, current_row, date_range):
	"""Add footer with signatures and notes exactly like the template"""
	from datetime import datetime, date
	
	total_columns = 8 + len(date_range['dates']) + 2  # Basic columns + dates + total + confirmation
	
	# Add some space before footer
	footer_start_row = current_row + 2
	
	# Add location and date (like template row 782-783)
	date_row = footer_start_row
	current_date = date.today()
	date_str_en = f"Quang Ngai, {current_date.strftime('%d %b %Y')}"
	date_str_vn = f"Quảng Ngãi, ngày {current_date.day} tháng {current_date.month:02d} năm {current_date.year}"
	
	# Position date in column 29 area (right side)
	date_col = min(total_columns - 10, 25)  # Adjust based on total columns
	ws.cell(row=date_row, column=date_col, value=date_str_en)
	ws.cell(row=date_row, column=date_col).font = Font(name='Times New Roman', size=10)
	ws.cell(row=date_row, column=date_col).alignment = Alignment(horizontal='center', vertical='center')
	
	date_row_vn = date_row + 1
	ws.cell(row=date_row_vn, column=date_col, value=date_str_vn)
	ws.cell(row=date_row_vn, column=date_col).font = Font(name='Times New Roman', size=10)
	ws.cell(row=date_row_vn, column=date_col).alignment = Alignment(horizontal='center', vertical='center')
	
	# Signature section (like template row 784)
	signature_row = date_row_vn + 1
	
	# Three signature columns as per template
	# Left signature (Col 3 area) - Prepared by
	ws.cell(row=signature_row, column=3, value="Prepared by")
	ws.cell(row=signature_row, column=3).font = Font(name='Times New Roman', bold=True, size=10)
	ws.cell(row=signature_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
	
	# Middle signature (Col 15 area) - Verified by 
	middle_col = min(15, total_columns // 2)
	ws.cell(row=signature_row, column=middle_col, value="Verified by")
	ws.cell(row=signature_row, column=middle_col).font = Font(name='Times New Roman', bold=True, size=10)
	ws.cell(row=signature_row, column=middle_col).alignment = Alignment(horizontal='center', vertical='center')
	
	# Right signature (Col 29 area) - Approved by
	right_col = min(total_columns - 5, date_col)
	ws.cell(row=signature_row, column=right_col, value="Approved by")
	ws.cell(row=signature_row, column=right_col).font = Font(name='Times New Roman', bold=True, size=10)
	ws.cell(row=signature_row, column=right_col).alignment = Alignment(horizontal='center', vertical='center')
	
	# Job titles (like template row 785)
	title_row = signature_row + 1
	ws.cell(row=title_row, column=3, value="C&B Executive")
	ws.cell(row=title_row, column=3).font = Font(name='Times New Roman', size=10)
	ws.cell(row=title_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
	
	ws.cell(row=title_row, column=middle_col, value="Office Department Manager")
	ws.cell(row=title_row, column=middle_col).font = Font(name='Times New Roman', size=10)
	ws.cell(row=title_row, column=middle_col).alignment = Alignment(horizontal='center', vertical='center')
	
	ws.cell(row=title_row, column=right_col, value="Head of Branch")
	ws.cell(row=title_row, column=right_col).font = Font(name='Times New Roman', size=10)
	ws.cell(row=title_row, column=right_col).alignment = Alignment(horizontal='center', vertical='center')
	
	# Names (like template row 790) - skip for now, leave blank for manual signing
	name_row = signature_row + 6
	ws.cell(row=name_row, column=3, value="Phạm Thị Kim Loan")
	ws.cell(row=name_row, column=3).font = Font(name='Times New Roman', size=10)
	ws.cell(row=name_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
	
	ws.cell(row=name_row, column=middle_col, value="Nguyễn Thị Yến Ni")
	ws.cell(row=name_row, column=middle_col).font = Font(name='Times New Roman', size=10)
	ws.cell(row=name_row, column=middle_col).alignment = Alignment(horizontal='center', vertical='center')
	
	ws.cell(row=name_row, column=right_col, value="KITAJIMA MOTOHARU")
	ws.cell(row=name_row, column=right_col).font = Font(name='Times New Roman', size=10)
	ws.cell(row=name_row, column=right_col).alignment = Alignment(horizontal='center', vertical='center')
	
	# Legend section (like template rows 792-805)
	legend_start_row = name_row + 2
	
	# Left column legend items (Col 3) - Basic abbreviations
	left_legend_items = [
		"O: Ốm hưởng BHXH",
		"CO: Con ốm hưởng BHXH", 
		"KL: Không lương",
		"P: Nghỉ 1 ngày phép năm",
		"P/2: Nghỉ 1/2 ngày phép năm",
		"F0: Covid",
		"NL: Nghỉ lễ",
		"MC: Nghỉ ma chay",
		"HS: Nghỉ hỉ sự",
		"NB: Nghỉ bù",
		"HL: Nghỉ hưởng lương",
		"HL/2: Nghỉ hưởng lương 1/2 ngày",
		"TS: Thai sản",
		"DS: Dưỡng sức"
	]
	
	# Right column legend items (Col 7-8) - Advanced abbreviations with explanations
	right_legend_items = [
		("O/2", "Ốm - Đi làm"),
		("CO/2", "Con ốm - Đi làm"),
		("OP/2", "Ốm - Phép năm"),
		("COP/2", "Con ốm - Phép năm"),
		("OL/2", "Ốm - Đi trễ về sớm (Số giờ nghỉ <=1 giờ)"),
		("COL/2", "Con ốm - Đi trễ về sớm (Số giờ nghỉ <=1 giờ)"),
		("OCO/2", "Ốm-Con ốm"),
		("OK/2", "1. Ốm - Không lương"),
		("OK/2", "2. Ốm - Đi trễ về sớm (1 giờ< số giờ nghỉ <4 giờ)"),
		("COK/2", "1. Con ốm - Không lương"),
		("COK/2", "2. Con ốm - Đi trễ về sớm (1 giờ< Số giờ nghỉ <4 giờ)")
	]
	
	# Add left column items (basic abbreviations)
	for i, text in enumerate(left_legend_items):
		legend_row = legend_start_row + i
		ws.cell(row=legend_row, column=3, value=text)
		ws.cell(row=legend_row, column=3).font = Font(name='Times New Roman', size=9)
		
	# Add right column items (advanced abbreviations with explanations)
	for i, (abbrev, explanation) in enumerate(right_legend_items):
		legend_row = legend_start_row + i
		# Abbreviation in column 7
		ws.cell(row=legend_row, column=7, value=abbrev)
		ws.cell(row=legend_row, column=7).font = Font(name='Times New Roman', size=9)
		# Explanation in column 8
		ws.cell(row=legend_row, column=8, value=explanation)
		ws.cell(row=legend_row, column=8).font = Font(name='Times New Roman', size=9)

def apply_excel_formatting(ws, num_employees, num_dates, date_range):
	"""Apply formatting to Excel worksheet"""
	# Set column widths
	ws.column_dimensions['A'].width = 8   # STT
	ws.column_dimensions['B'].width = 12  # Employee ID  
	ws.column_dimensions['C'].width = 25  # Name
	ws.column_dimensions['D'].width = 12  # Join date
	ws.column_dimensions['E'].width = 12  # Resign date
	ws.column_dimensions['F'].width = 12  # Group
	ws.column_dimensions['G'].width = 15  # Section
	ws.column_dimensions['H'].width = 15  # Position
	
	# Set width for date columns
	for i in range(num_dates):
		col_letter = get_column_letter(9 + i)
		ws.column_dimensions[col_letter].width = 4
	
	# Set width for total and confirmation columns (swapped order)
	total_col = get_column_letter(9 + num_dates)
	confirmation_col = get_column_letter(9 + num_dates + 1)
	ws.column_dimensions[total_col].width = 9  # Total working days (50% of original 18)
	ws.column_dimensions[confirmation_col].width = 12  # Xác nhận
	
	# Set row heights
	ws.row_dimensions[5].height = 30  # Header row
	ws.row_dimensions[6].height = 20  # Date row
	ws.row_dimensions[7].height = 15  # Day row
	
	# Apply borders to data area
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'), 
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	
	# Header borders and background (light blue-gray)
	header_fill = PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid')  # Light steel blue
	sunday_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Gray for Sundays
	
	# Calculate total columns including new ones
	total_columns = 9 + num_dates + 2  # Basic + dates + confirmation + total
	
	# Apply to headers (rows 5-7)
	for row in range(5, 8):
		for col in range(1, total_columns):
			cell = ws.cell(row=row, column=col)
			cell.border = thin_border
			cell.fill = header_fill
	
	# Apply to data rows (including total row)
	for row in range(8, 8 + num_employees):  # num_employees already includes total row
		for col in range(1, total_columns):
			cell = ws.cell(row=row, column=col)
			
			# Apply border to regular employee rows only (not department/total rows)
			cell.border = thin_border
			
			# Check if this is a Sunday column
			if col >= 9 and col < 9 + num_dates:  # Date columns only
				date_index = col - 9
				if date_index < len(date_range['dates']):
					date_obj = date_range['dates'][date_index]
					if date_obj.weekday() == 6:  # Sunday = 6 in Python weekday()
						# Only apply Sunday fill if not already a department/total row
						if not cell.fill or cell.fill.start_color.rgb != 'FFE0E0E0':
							cell.fill = sunday_fill