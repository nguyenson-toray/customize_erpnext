"""
Standard-app Excel export — faithful Python replica of the Flutter timesheet
app's `TimesheetFunctions.exportTimesheets()` (see flutter_app_chuẩn/
timesheetFunctions.dart and App_Chuẩn_timesheetAlgorithm.md).

Produces exactly 6 sheets, in order:
    Important Note | Detail | Summary | Timesheet | Overtime | Shift

Data source is the Attendance table (values computed by the optimized
attendance engine); the row universe mirrors the app:
  - only dates on which at least one attendance record exists ("skip date
    if nobody checked in", e.g. idle Sundays)
  - on each such date: every employee active that day, INCLUDING absent
    ones (blank in/out, zeros) — joiners appear from joining date, leavers
    disappear from relieving date unless they still have attendance.

Layout, formats and sheet set follow the Dart source exactly. One content
addition kept at the user's request: leave applications appear in Note
Checkin as "Phép: {abbr}" (the app has no leave data).
"""

import re

import frappe
from frappe.utils import getdate, now_datetime

from customize_erpnext.overrides.shift_type.leave_hour_cap import is_suspicious
from datetime import timedelta, datetime, date as date_type
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Constants (mirroring the Dart implementation) ───────────────────────────

DETAIL_HDRS = [
    "No", "Date", "Employee ID", "Finger ID", "Full name", "Department",
    "Section", "Group", "Shift", "First In", "Last Out",
    "Working (hour)", "Actual (hour)", "Working (day)", "OT Actual (hours)",
    "OT Approved (hours)", "OT Final", "Note Checkin", "Note Sunday",
    "Joining Date", "Resign Date",
]
DETAIL_WIDTHS = [4, 11, 10, 6, 24, 10, 14, 18, 7, 8, 8, 8, 9, 8, 8, 8, 8, 20, 20, 11, 11]

EMP_FIXED_HDRS = ["No", "Employee ID", "Full name", "Joining Date",
                  "Resign Date", "Group", "Section", "Position"]
EMP_FIXED_WIDTHS = [4, 12, 24, 11, 11, 18, 14, 14]

# Độ rộng cột NGÀY (dd/mm) trên các sheet pivot — nới +10% so với bản gốc (6 và 8) cho dễ đọc.
DATE_COL_WIDTH_PIVOT = 6.6      # Timesheet · Overtime
DATE_COL_WIDTH_SHIFT = 8.8      # Shift (ô chứa chữ "Shift 1"/"Shift 2")
EMP_FIXED = 8

SUM_EXTRA_HDRS = [
    "Total Working (hours)", "Total Working (days)", "Total OT Actual (hours)",
    "Total OT Approved (hours)", "Total OT Final (hours)",
]

SUNDAY_GRAY = "D9D9D9"
SHIFT1_COLOR = "C55A11"  # dark orange
SHIFT2_COLOR = "1F4E79"  # dark blue

_CENTER = Alignment(horizontal="center")
_WRAP = Alignment(wrap_text=True)
_THIN = Side(style="thin")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _r1(v):
    return round(float(v or 0), 1)


def _r2(v):
    return round(float(v or 0), 2)


# ── Data loading ─────────────────────────────────────────────────────────────

def load_export_universe(from_date, to_date, department=None, only_resigned=False):
    """Employees + attendance + maternity periods for the range.

    Returns dict(employees=[{...}], att={(emp, date): row}, dates=[date],
    maternity={emp: [(kind, from, to)]}).

    `only_resigned`: chỉ lấy người **nghỉ việc TRONG kỳ** — `relieving_date` nằm trong khoảng
    from..to. Dùng khi HR làm thủ tục chốt lương cho người thôi việc. Không phải "status = Left":
    người nghỉ từ năm ngoái cũng là `Left` nhưng không liên quan kỳ này.
    """
    from customize_erpnext.customize_erpnext.doctype.attendance_calculation_setting.attendance_calculation_setting import get_attendance_settings

    from_date, to_date = getdate(from_date), getdate(to_date)
    params = {"from_date": from_date, "to_date": to_date}

    conds = ""
    emp_prefix = (get_attendance_settings().employee_id_prefix or "").strip()
    if emp_prefix:
        conds += " AND emp.name LIKE %(prefix)s"
        params["prefix"] = f"{emp_prefix}%"
    if department:
        dept_list = department if isinstance(department, list) else [department]
        conds += " AND emp.department IN %(departments)s"
        params["departments"] = tuple(dept_list)
    # `only_resigned` THAY THẾ điều kiện trạng thái mặc định, không cộng thêm vào.
    #
    # ⚠ Điều kiện gốc dùng `relieving_date > from_date` (lớn hơn HẲN) nên loại luôn người nghỉ
    # đúng ngày đầu kỳ. Nếu chỉ AND thêm `BETWEEN from AND to` thì option hứa "nghỉ việc trong
    # kỳ" nhưng lại rơi mất họ — đo tháng 6/2026: 57/59, thiếu TIQN-1653 và TIQN-2144 đều nghỉ
    # đúng 01/06.
    status_cond = """(
            (emp.status = 'Active'
             AND (emp.date_of_joining IS NULL OR emp.date_of_joining <= %(to_date)s))
            OR
            (emp.status = 'Left'
             AND (emp.date_of_joining IS NULL OR emp.date_of_joining <= %(to_date)s)
             AND (emp.relieving_date IS NULL OR emp.relieving_date > %(from_date)s))
        )"""
    if only_resigned:
        status_cond = ("(emp.relieving_date IS NOT NULL"
                       " AND emp.relieving_date BETWEEN %(from_date)s AND %(to_date)s)")

    employees = frappe.db.sql(f"""
        SELECT emp.name, emp.employee_name, emp.attendance_device_id,
               emp.department, emp.custom_section, emp.custom_group,
               emp.designation, emp.date_of_joining, emp.relieving_date,
               emp.status, emp.default_shift
        FROM `tabEmployee` emp
        WHERE {status_cond} {conds}
        ORDER BY emp.name
    """, params, as_dict=True)

    # Same exclusion list as the attendance engine (setting: exclude_employee_ids)
    from customize_erpnext.customize_erpnext.doctype.attendance_calculation_setting.attendance_calculation_setting import get_excluded_employee_ids
    excluded = get_excluded_employee_ids() or set()
    if excluded:
        employees = [e for e in employees if e.name not in excluded]
    emp_ids = [e.name for e in employees]

    att = {}
    dates = set()
    if emp_ids:
        att_rows = frappe.db.sql("""
            SELECT name, employee, attendance_date, shift, in_time, out_time,
                   working_hours, custom_actual_working_hours, actual_overtime_duration,
                   custom_approved_overtime_duration, custom_final_overtime_duration,
                   late_entry, early_exit, custom_note,
                   custom_leave_application_abbreviation, leave_application, status
            FROM `tabAttendance`
            WHERE employee IN %(employees)s
              AND attendance_date BETWEEN %(from_date)s AND %(to_date)s
              AND docstatus = 1
        """, {"employees": emp_ids, "from_date": from_date, "to_date": to_date},
            as_dict=True)
        # App rule: a date only enters the report when somebody actually
        # checked in that day (skips idle Sundays/holidays that still carry
        # engine-created Absent records, e.g. a full-absence Saturday)
        for r in att_rows:
            r.attendance_date = getdate(r.attendance_date)
            att[(r.employee, r.attendance_date)] = r
            if r.in_time:
                dates.add(r.attendance_date)

    # Maternity periods → per-date regime flags ("Chế độ mang thai/con nhỏ")
    maternity = defaultdict(list)
    if emp_ids:
        # NULL from_date = open start (app semantics — several HR-imported
        # records only carry the end date; without this the regime notes and
        # the [Ra 16-17h] anomaly suppression would miss those employees)
        open_start = date_type(1900, 1, 1)
        for m in frappe.db.sql("""
            SELECT employee, pregnant_from_date, pregnant_to_date,
                   estimated_due_date, youg_child_from_date, youg_child_to_date
            FROM `tabEmployee Maternity`
            WHERE employee IN %(employees)s
        """, {"employees": emp_ids}, as_dict=True):
            preg_to = m.pregnant_to_date or m.estimated_due_date
            if preg_to:
                maternity[m.employee].append(
                    ("Pregnant",
                     getdate(m.pregnant_from_date) if m.pregnant_from_date else open_start,
                     getdate(preg_to)))
            if m.youg_child_to_date:
                maternity[m.employee].append(
                    ("Young Child",
                     getdate(m.youg_child_from_date) if m.youg_child_from_date else open_start,
                     getdate(m.youg_child_to_date)))

    return {
        "employees": employees,
        "att": att,
        "dates": sorted(dates),
        "maternity": maternity,
        "from_date": from_date,
        "to_date": to_date,
    }


def _regime(universe, emp_id, d):
    """'Pregnant' / 'Young Child' / None for employee on date."""
    for kind, frm, to in universe["maternity"].get(emp_id, []):
        if frm <= d <= to:
            return kind
    return None


def _clean_department(name):
    if name and name.endswith(" - TIQN"):
        return name.replace(" - TIQN", "")
    return name or ""


# ── Row building (TimeSheetDate equivalent) ──────────────────────────────────

def _build_notes(att_row, regime):
    """(note_checkin, note_sunday) in the app's Vietnamese wording."""
    checkin_parts = []
    sunday_parts = []

    note_en = (att_row.get("custom_note") or "") if att_row else ""

    # Base state notes (app: replace, mutually exclusive)
    if "Only one check-in record" in note_en:
        checkin_parts.append("Chỉ có 1 lần chấm công")
    elif "No check-OUT" in note_en:
        checkin_parts.append("Không chấm công RA")
    elif "No check-IN" in note_en:
        checkin_parts.append("Không chấm công VÀO")

    if att_row:
        if att_row.get("late_entry"):
            checkin_parts.append("Vào trễ")
        if att_row.get("early_exit"):
            checkin_parts.append("Ra sớm")

    if "before shift without OT Registration" in note_en:
        checkin_parts.append("Vào sớm ≥1h, không có ĐK OT trước ca")
    if "after shift without OT Registration" in note_en:
        checkin_parts.append("Ra trễ ≥1h, không có ĐK OT sau ca")

    if regime == "Pregnant":
        checkin_parts.append("Chế độ mang thai")
    elif regime == "Young Child":
        checkin_parts.append("Chế độ con nhỏ")

    # Leave application info (user-requested addition — the app has no leave data)
    abbr = att_row.get("custom_leave_application_abbreviation") if att_row else None
    if abbr:
        checkin_parts.append(f"Phép: {abbr}")

    # Xin nghỉ phép nhưng vẫn đi làm — giờ thực tế vượt giờ được tính.
    # ⚠ Hàm này chỉ dịch những chuỗi ĐÃ BIẾT trong custom_note; note nào không khai ở đây sẽ
    # BIẾN MẤT khỏi cột Note Checkin. Thêm note mới ở engine thì phải thêm cả ở đây.
    # Chuỗi gốc do overrides/shift_type/leave_hour_cap.leave_hour_note() sinh.
    if "hours not counted" in note_en:
        checkin_parts.append("Nghỉ trọn ngày nhưng vẫn đi làm - không tính giờ")
    elif "capped to" in note_en:
        actual = ""
        m = re.search(r"worked ([\d.]+)h", note_en)
        if m:
            actual = f" {m.group(1)}h"
        checkin_parts.append(f"Nghỉ nửa ngày nhưng làm{actual} - chặn còn 4h")
    if "check if LA should be cancelled" in note_en:
        checkin_parts.append("→ XEM HUỶ ĐƠN NGHỈ")

    # Sunday notes (attNote3)
    if "Sunday work" in note_en:
        sunday_parts.append("OT ngày CN")
        if "meal allowance" in note_en:
            sunday_parts.append("Có phụ cấp cơm trưa")

    return " ; ".join(checkin_parts), " ; ".join(sunday_parts)


def build_export_rows(universe, leave_gap_minutes: int = 15):
    """One dict per (date × active employee) — the app's TimeSheetDate list.

    Also returns anomalies for the Important Note sheet.

    `leave_gap_minutes`: chênh lệch tối thiểu (phút) giữa giờ thực tế và giờ tính công để một
    ngày "có đơn nghỉ mà vẫn đi làm" được đưa vào sheet Important Note. `0` = báo tất.
    Mặc định 15 vì 95/312 ca bị chặn chỉ chênh dưới 15 phút — nhiễu làm tròn, không đáng báo.
    """
    rows = []
    anomalies = []

    for d in universe["dates"]:
        is_sunday = d.weekday() == 6
        for emp in universe["employees"]:
            if emp.date_of_joining and d < getdate(emp.date_of_joining):
                continue
            att_row = universe["att"].get((emp.name, d))
            if att_row is None and emp.status == "Left" and emp.relieving_date \
                    and d >= getdate(emp.relieving_date):
                continue

            regime = _regime(universe, emp.name, d)
            note_checkin, note_sunday = _build_notes(att_row, regime)

            if att_row:
                shift = att_row.shift or ("Day" if is_sunday else emp.default_shift or "Day")
                working_hours = _r2(att_row.working_hours)
                row = {
                    "date": d,
                    "emp_id": emp.name,
                    "finger_id": emp.attendance_device_id,
                    "name": emp.employee_name,
                    "department": _clean_department(emp.department),
                    "section": emp.custom_section or "",
                    "group": emp.custom_group or "",
                    "shift": shift,
                    "first_in": att_row.in_time,
                    "last_out": att_row.out_time,
                    "working_hours": working_hours,
                    "working_days": _r2(working_hours / 8.0),
                    # Giờ THỰC TẾ theo check in/out — không bị chặn theo đơn nghỉ phép.
                    # Xem overrides/shift_type/leave_hour_cap.py
                    "actual_hours": _r2(att_row.custom_actual_working_hours),
                    # Chỉ sheet Timesheet dùng — xem overrides/shift_attendance/timesheet_leave.py
                    "leave_abbr": (att_row.custom_leave_application_abbreviation or "").strip(),
                    "ot_actual": _r1(att_row.actual_overtime_duration),
                    "ot_approved": _r1(att_row.custom_approved_overtime_duration),
                    "ot_final": _r1(att_row.custom_final_overtime_duration),
                    "note_checkin": note_checkin,
                    "note_sunday": note_sunday,
                }

                note_en = att_row.custom_note or ""
                dat_str = d.strftime("%Y-%m-%d")
                emp_tag = f"{emp.name} {emp.employee_name}"
                # Một dòng của sheet Important Note. Giữ ngày/mã NV thành TRƯỜNG RIÊNG để
                # sắp xếp được — nhét vào chuỗi thì không sort theo Type/Date/Employee được.
                _i = att_row.in_time.strftime("%H:%M") if att_row.in_time else "?"
                _o = att_row.out_time.strftime("%H:%M") if att_row.out_time else "?"
                anom_cols = {
                    "date": d,
                    "employee": emp.name,
                    "employee_name": emp.employee_name,
                    "in_out": f"{_i}–{_o}",
                    "working_hours": working_hours,
                    "actual_hours": _r2(att_row.custom_actual_working_hours),
                    "abbr": (att_row.custom_leave_application_abbreviation or "").strip() or None,
                    "attendance": att_row.name,
                    "leave_application": att_row.leave_application or None,
                }
                if "Left employee" in note_en:
                    resign = getdate(emp.relieving_date).strftime("%Y-%m-%d") \
                        if emp.relieving_date else "?"
                    anomalies.append(
                        ("[Resigned + Att]", f"resigned on {resign}, has check-ins", anom_cols))
                # App suppresses [Ra 16-17h] for employees under a maternity
                # regime at that date (leaving 1h early is expected for them)
                # Xin nghỉ phép nhưng vẫn đi làm — HR cần biết để huỷ đơn.
                # Ngưỡng và cách nhận diện ở overrides/shift_type/leave_hour_cap.py
                _abbr = (att_row.custom_leave_application_abbreviation or "").strip()
                _actual = _r2(att_row.custom_actual_working_hours)
                # Chênh dưới ngưỡng = nhiễu làm tròn (vd 4,01h vs 4,00h = 36 giây), không phải
                # "đi làm dù có phép". Mặc định 15 phút; 0 = báo tất.
                _excess_min = (_actual - working_hours) * 60
                if _excess_min >= leave_gap_minutes and is_suspicious(
                        _actual, _abbr, att_row.status == "Half Day",
                        bool(att_row.custom_leave_application_abbreviation)):
                    _in = att_row.in_time.strftime("%H:%M") if att_row.in_time else "?"
                    _out = att_row.out_time.strftime("%H:%M") if att_row.out_time else "?"
                    # Cột Note để TRỐNG: các cột số + mã đơn đã nói đủ, HR tự quyết xử lý.
                    anomalies.append(("[Nghỉ phép + đi làm]", "", anom_cols))

                if "Female checkout" in note_en and not regime:
                    out_str = att_row.out_time.strftime("%H:%M") if att_row.out_time else "?"
                    anomalies.append(
                        ("[Ra 16-17h]", f"last out {out_str} (shift: {shift})", anom_cols))
            else:
                # Absent day — blank in/out, zeros; regime note still shown
                row = {
                    "date": d,
                    "emp_id": emp.name,
                    "finger_id": emp.attendance_device_id,
                    "name": emp.employee_name,
                    "department": _clean_department(emp.department),
                    "section": emp.custom_section or "",
                    "group": emp.custom_group or "",
                    "shift": "Day" if is_sunday else (emp.default_shift or "Day"),
                    "first_in": None,
                    "last_out": None,
                    "working_hours": 0.0,
                    "working_days": 0.0,
                    "actual_hours": 0.0,
                    "leave_abbr": "",
                    "ot_actual": 0.0,
                    "ot_approved": 0.0,
                    "ot_final": 0.0,
                    "note_checkin": note_checkin,
                    "note_sunday": note_sunday,
                }
            rows.append(row)

    return rows, anomalies


# ── Sheet writers ────────────────────────────────────────────────────────────

def _add_excel_table(ws, last_row, last_col, name):
    """Excel table with the app's style (Medium16, banded rows)."""
    if last_row < 2 or last_col < 1:
        return
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium16", showRowStripes=True,
        showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)


def _set_widths(ws, widths, start_col=1):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w


def _header_row(ws, headers, wrap=True, height=50):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        if wrap:
            cell.alignment = _WRAP
    if height:
        ws.row_dimensions[1].height = height


def _set_date(cell, d):
    if d:
        cell.value = d if isinstance(d, datetime) else datetime.combine(d, datetime.min.time())
        cell.number_format = "dd/MM/yyyy"


def _set_time(cell, t):
    if t:
        cell.value = t
        cell.number_format = "HH:mm"


def _joining_date(emp):
    if emp.date_of_joining and getdate(emp.date_of_joining).year > 1900:
        return getdate(emp.date_of_joining)
    return None


def _resign_date(emp):
    # App rule: only for resigned employees with a real (non-2099) date
    if emp.status == "Left" and emp.relieving_date \
            and getdate(emp.relieving_date).year < 2099:
        return getdate(emp.relieving_date)
    return None


def _write_fixed(ws, row, no, emp):
    """The 8 shared employee columns for Summary/pivot/Shift sheets."""
    ws.cell(row=row, column=1, value=no)
    ws.cell(row=row, column=2, value=emp.name)
    ws.cell(row=row, column=3, value=emp.employee_name)
    _set_date(ws.cell(row=row, column=4), _joining_date(emp))
    _set_date(ws.cell(row=row, column=5), _resign_date(emp))
    ws.cell(row=row, column=6, value=emp.custom_group or "")
    ws.cell(row=row, column=7, value=emp.custom_section or "")
    ws.cell(row=row, column=8, value=emp.designation or "")


NOTE_HDRS = ["Type", "Info", "Working Hour", "Working Hour Actual",
             "Leave Application Abbreviation", "Attendance", "Leave Application", "Note"]
NOTE_WIDTHS = [22, 46, 12, 16, 16, 14, 20, 34]


def add_important_note_sheet(ws, anomalies):
    """Sheet bất thường, định dạng Excel Table như các sheet khác.

    ⚠ `_add_excel_table` neo `ref` từ **A1**, nên tiêu đề cột phải ở **dòng 1** — không đặt được
    dòng tiêu đề "generated ..." phía trên như trước. Thời điểm xuất đã có trong TÊN FILE
    (`standard_export_filename`) nên không mất thông tin.

    Sắp xếp **Type → Date → Employee** (tăng dần). Vì vậy anomaly phải mang `date`/`employee`
    thành trường riêng; nhét vào chuỗi thì không sort được.
    """
    ws.title = "Important Note"
    _header_row(ws, NOTE_HDRS, height=30)

    rows = []
    for item in anomalies:
        a_type, a_note = item[0], item[1]
        x = item[2] if len(item) > 2 else {}
        rows.append({
            "type": a_type,
            "date": x.get("date"),
            "employee": x.get("employee") or "",
            "info": " · ".join(p for p in (
                x["date"].strftime("%d/%m/%Y") if x.get("date") else None,
                f"{x.get('employee', '')} {x.get('employee_name', '')}".strip() or None,
                x.get("in_out"),
            ) if p),
            "working_hours": x.get("working_hours"),
            "actual_hours": x.get("actual_hours"),
            "abbr": x.get("abbr"),
            "attendance": x.get("attendance"),
            "leave_application": x.get("leave_application"),
            "note": a_note,
        })

    # `date` có thể None với anomaly không gắn ngày -> dùng khoá thay thế cho khỏi vỡ so sánh
    rows.sort(key=lambda r: (r["type"], r["date"] or date_type.min, r["employee"]))

    row = 2
    if not rows:
        ws.cell(row=row, column=2, value="No anomalies detected.")
        row += 1
    else:
        for r in rows:
            ws.cell(row=row, column=1, value=r["type"])
            ws.cell(row=row, column=2, value=r["info"])
            for col, key in ((3, "working_hours"), (4, "actual_hours")):
                if r[key] is not None:
                    c = ws.cell(row=row, column=col, value=r[key])
                    c.number_format = "0.00"
                    c.alignment = _CENTER
            c = ws.cell(row=row, column=5, value=r["abbr"])
            c.alignment = _CENTER
            ws.cell(row=row, column=6, value=r["attendance"])
            ws.cell(row=row, column=7, value=r["leave_application"])
            ws.cell(row=row, column=8, value=r["note"])
            row += 1

    _add_excel_table(ws, row - 1, len(NOTE_HDRS), "TableImportantNote")
    _set_widths(ws, NOTE_WIDTHS)


def add_detail_sheet(wb, rows, emp_by_id):
    ws = wb.create_sheet("Detail")
    _header_row(ws, DETAIL_HDRS)

    data = sorted(rows, key=lambda r: (r["date"], r["emp_id"]))
    for no, r in enumerate(data, 1):
        row = no + 1
        emp = emp_by_id[r["emp_id"]]
        ws.cell(row=row, column=1, value=no)
        _set_date(ws.cell(row=row, column=2), r["date"])
        ws.cell(row=row, column=3, value=r["emp_id"])
        ws.cell(row=row, column=4, value=r["finger_id"])
        ws.cell(row=row, column=5, value=r["name"])
        ws.cell(row=row, column=6, value=r["department"])
        ws.cell(row=row, column=7, value=r["section"])
        ws.cell(row=row, column=8, value=r["group"])
        ws.cell(row=row, column=9, value=r["shift"])
        _set_time(ws.cell(row=row, column=10), r["first_in"])
        _set_time(ws.cell(row=row, column=11), r["last_out"])
        c = ws.cell(row=row, column=12, value=r["working_hours"]); c.number_format = "0.00"
        c = ws.cell(row=row, column=13, value=r["actual_hours"]); c.number_format = "0.00"
        c = ws.cell(row=row, column=14, value=r["working_days"]); c.number_format = "0.00"
        c = ws.cell(row=row, column=15, value=r["ot_actual"]); c.number_format = "0.0"
        c = ws.cell(row=row, column=16, value=r["ot_approved"]); c.number_format = "0.0"
        c = ws.cell(row=row, column=17, value=r["ot_final"]); c.number_format = "0.0"
        ws.cell(row=row, column=18, value=r["note_checkin"])
        ws.cell(row=row, column=19, value=r["note_sunday"])
        _set_date(ws.cell(row=row, column=20), _joining_date(emp))
        _set_date(ws.cell(row=row, column=21), _resign_date(emp))

    _add_excel_table(ws, len(data) + 1, len(DETAIL_HDRS), "TableDetail")
    _set_widths(ws, DETAIL_WIDTHS)


def add_summary_sheet(wb, rows, emp_by_id):
    ws = wb.create_sheet("Summary")
    _header_row(ws, EMP_FIXED_HDRS + SUM_EXTRA_HDRS)

    totals = {}
    for r in rows:
        t = totals.setdefault(r["emp_id"], [0.0, 0.0, 0.0, 0.0, 0.0])
        t[0] += r["working_hours"]
        t[1] += r["working_days"]   # Σ of per-day rounded values (app §11)
        t[2] += r["ot_actual"]
        t[3] += r["ot_approved"]
        t[4] += r["ot_final"]

    row = 2
    for no, emp_id in enumerate(sorted(totals), 1):
        t = totals[emp_id]
        _write_fixed(ws, row, no, emp_by_id[emp_id])
        for i, (val, fmt) in enumerate(zip(
                (_r2(t[0]), _r2(t[1]), _r1(t[2]), _r1(t[3]), _r1(t[4])),
                ("0.00", "0.00", "0.0", "0.0", "0.0"))):
            c = ws.cell(row=row, column=EMP_FIXED + 1 + i, value=val)
            c.number_format = fmt
        row += 1

    _add_excel_table(ws, row - 1, EMP_FIXED + 5, "TableSummary")
    _set_widths(ws, EMP_FIXED_WIDTHS + [8] * 5)


def _add_pivot_sheet(wb, sheet_name, table_name, number_fmt, pivot_values,
                     emp_order, emp_by_id, all_dates, skip_zero_rows=False,
                     display_values=None):
    """Numeric employee × date pivot (Timesheet / Overtime), app-identical.

    `display_values` (tuỳ chọn, chỉ Timesheet dùng): {emp_id: {date: "P"}} — ô nào có mã thì in
    CHỮ thay cho số, nhưng cột Total vẫn cộng số của `pivot_values`. Sheet Overtime không truyền
    tham số này nên hành vi giữ nguyên tuyệt đối.
    """
    ws = wb.create_sheet(sheet_name)
    date_cols = len(all_dates)
    total_col = EMP_FIXED + date_cols + 1
    _header_row(ws, EMP_FIXED_HDRS + [d.strftime("%d/%m") for d in all_dates] + ["Total"])

    def apply_cell(cell, val, text=None):
        if text is not None:
            cell.value = text          # mã nghỉ phép: để dạng chữ, KHÔNG gán number_format
        elif val > 0:
            cell.value = val
            cell.number_format = "0" if val == int(val) else number_fmt
        cell.alignment = _CENTER

    row = 2
    for no, emp_id in enumerate(emp_order, 1):
        day_map = pivot_values.get(emp_id, {})
        total = sum(day_map.values())
        if skip_zero_rows and total == 0:
            continue
        text_map = (display_values or {}).get(emp_id, {})
        _write_fixed(ws, row, no, emp_by_id[emp_id])
        for ci, d in enumerate(all_dates):
            apply_cell(ws.cell(row=row, column=EMP_FIXED + 1 + ci),
                       day_map.get(d, 0.0), text_map.get(d))
        apply_cell(ws.cell(row=row, column=total_col), total)
        row += 1

    # Fix numbering when rows were skipped (No must be 1..n consecutive)
    if skip_zero_rows:
        for i, r in enumerate(range(2, row), 1):
            ws.cell(row=r, column=1, value=i)

    _add_excel_table(ws, row - 1, total_col, table_name)
    _set_widths(ws, EMP_FIXED_WIDTHS + [DATE_COL_WIDTH_PIVOT] * date_cols + [8])

    # Sunday columns: gray header, keep data cells white
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    gray = PatternFill(start_color=SUNDAY_GRAY, end_color=SUNDAY_GRAY, fill_type="solid")
    for ci, d in enumerate(all_dates):
        if d.weekday() != 6:
            continue
        col = EMP_FIXED + 1 + ci
        ws.cell(row=1, column=col).fill = gray
        for r in range(2, row):
            ws.cell(row=r, column=col).fill = white


def add_shift_sheet(wb, rows, emp_by_id, all_dates):
    """Shift 1 / Shift 2 rotation matrix (only employees on rotating shifts)."""
    shift_map = defaultdict(dict)
    for r in rows:
        if r["shift"] in ("Shift 1", "Shift 2"):
            shift_map[r["emp_id"]][r["date"]] = r["shift"]
    if not shift_map:
        return

    ws = wb.create_sheet("Shift")
    date_cols = len(all_dates)
    last_col = EMP_FIXED + date_cols
    _header_row(ws, EMP_FIXED_HDRS + [d.strftime("%d/%m") for d in all_dates])

    row = 2
    for no, emp_id in enumerate(sorted(shift_map), 1):
        _write_fixed(ws, row, no, emp_by_id[emp_id])
        for ci, d in enumerate(all_dates):
            val = shift_map[emp_id].get(d, "")
            cell = ws.cell(row=row, column=EMP_FIXED + 1 + ci, value=val)
            cell.alignment = _CENTER
            if val == "Shift 1":
                cell.font = Font(bold=True, color=SHIFT1_COLOR)
            elif val == "Shift 2":
                cell.font = Font(bold=True, color=SHIFT2_COLOR)
            else:
                cell.font = Font(bold=True)
        row += 1

    _add_excel_table(ws, row - 1, last_col, "TableShift")
    _set_widths(ws, EMP_FIXED_WIDTHS + [DATE_COL_WIDTH_SHIFT] * date_cols)

    gray = PatternFill(start_color=SUNDAY_GRAY, end_color=SUNDAY_GRAY, fill_type="solid")
    for ci, d in enumerate(all_dates):
        if d.weekday() == 6:
            ws.cell(row=1, column=EMP_FIXED + 1 + ci).fill = gray


# ── Entry point ──────────────────────────────────────────────────────────────

ALL_SHEETS = ("Important Note", "Detail", "Summary", "Timesheet", "Overtime", "Shift")


def build_standard_workbook(from_date, to_date, department=None,
                            leave_gap_minutes: int = 15, sheets=None,
                            only_resigned: bool = False):
    """Build the standard-app workbook. Returns openpyxl Workbook.

    `sheets`: tên các sheet cần xuất; `None` = tất cả (`ALL_SHEETS`).
    `leave_gap_minutes`: xem `build_export_rows`.
    `only_resigned`: xem `load_export_universe`.

    ⚠ `wb.active` tạo sẵn sheet đầu tiên. Nếu KHÔNG xuất "Important Note" thì phải xoá sheet
    trống đó, nếu không file có một tab rỗng tên "Sheet".
    """
    wanted = set(sheets) if sheets else set(ALL_SHEETS)
    universe = load_export_universe(from_date, to_date, department, only_resigned=only_resigned)
    rows, anomalies = build_export_rows(universe, leave_gap_minutes=leave_gap_minutes)
    emp_by_id = {e.name: e for e in universe["employees"]}

    # Continuous date columns for the pivot sheets (full requested range,
    # like the app's dateRange input)
    all_dates = []
    d = universe["from_date"]
    while d <= universe["to_date"]:
        all_dates.append(d)
        d += timedelta(days=1)

    wb = Workbook()
    if "Important Note" in wanted:
        add_important_note_sheet(wb.active, anomalies)
    else:
        # wb.active là sheet mặc định "Sheet" — bỏ đi, nếu không file có tab rỗng
        wb.remove(wb.active)
    if "Detail" in wanted:
        add_detail_sheet(wb, rows, emp_by_id)
    if "Summary" in wanted:
        add_summary_sheet(wb, rows, emp_by_id)

    # Timesheet pivot: working days per day; every employee present in rows
    #
    # Sheet Timesheet in MÃ NGHỈ PHÉP (P, KL, O/2...) vào ô của ngày nghỉ và tính ngày công
    # theo mục 3 quy chế, thay vì giờ/8. Detail + Summary phía trên KHÔNG đổi — chúng vẫn dùng
    # r["working_days"]. Xem overrides/shift_attendance/timesheet_leave.py.
    from customize_erpnext.overrides.shift_attendance.timesheet_leave import (
        timesheet_cell_display, timesheet_working_days)

    emp_order = sorted({r["emp_id"] for r in rows})
    ts_pivot = defaultdict(lambda: defaultdict(float))
    ts_display = defaultdict(dict)
    ot_pivot = defaultdict(lambda: defaultdict(float))
    for r in rows:
        abbr = r.get("leave_abbr") or ""
        ts_pivot[r["emp_id"]][r["date"]] += timesheet_working_days(abbr, r["working_hours"])
        shown = timesheet_cell_display(abbr, r["working_hours"])
        if isinstance(shown, str):
            ts_display[r["emp_id"]][r["date"]] = shown
        if r["ot_final"] > 0:
            ot_pivot[r["emp_id"]][r["date"]] += r["ot_final"]

    if "Timesheet" in wanted:
        _add_pivot_sheet(wb, "Timesheet", "TableTimesheet", "0.##",
                         ts_pivot, emp_order, emp_by_id, all_dates,
                         display_values=ts_display)
    if "Overtime" in wanted:
        _add_pivot_sheet(wb, "Overtime", "TableOvertime", "0.#",
                         ot_pivot, emp_order, emp_by_id, all_dates,
                         skip_zero_rows=True)
    if "Shift" in wanted:
        add_shift_sheet(wb, rows, emp_by_id, all_dates)

    # Không sheet nào được chọn -> openpyxl không cho lưu workbook rỗng
    if not wb.sheetnames:
        wb.create_sheet("Empty")

    return wb


def standard_export_filename(from_date, to_date):
    """Timesheet_{yymmdd}_{yymmdd}_{yymmdd_HHMMSS}.xlsx — app naming."""
    ts = now_datetime().strftime("%y%m%d_%H%M%S")
    return (f"Timesheet_{getdate(from_date).strftime('%y%m%d')}_"
            f"{getdate(to_date).strftime('%y%m%d')}_{ts}.xlsx")
