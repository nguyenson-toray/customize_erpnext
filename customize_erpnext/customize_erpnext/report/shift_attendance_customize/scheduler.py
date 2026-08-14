import frappe
from frappe import _
from frappe.utils import today, formatdate, get_datetime, add_days, getdate
from datetime import datetime, time as time_obj, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import tempfile


def _get_employee_prefix():
	"""Employee ID prefix from Attendance Calculation Setting (cached doc, cheap to call).
	Empty prefix = no filtering (LIKE '%' / startswith(''))."""
	from customize_erpnext.customize_erpnext.doctype.attendance_calculation_setting.attendance_calculation_setting import get_attendance_settings
	return (get_attendance_settings().employee_id_prefix or "").strip()

def _is_holiday_or_sunday(date_str):
	"""Check if date is a Holiday (from default company Holiday List) or Sunday."""
	from frappe.utils import getdate

	check_date = getdate(date_str)

	# Check Sunday (weekday 6 = Sunday)
	if check_date.weekday() == 6:
		return True

	# Check Holiday List from default company
	default_company = frappe.defaults.get_defaults().get("company")
	if default_company:
		holiday_list = frappe.db.get_value("Company", default_company, "default_holiday_list")
		if holiday_list and frappe.db.exists("Holiday", {
			"parent": holiday_list,
			"holiday_date": check_date
		}):
			return True

	return False


def recalculate_attendance(report_date_str):
	"""Rebuild attendance for one day from the raw check-ins.

	Exposed on its own so a caller sending to several audiences can run it once
	up front instead of once per mail — it is the expensive part of the job.
	"""
	from frappe.utils import getdate

	from customize_erpnext.overrides.shift_type.shift_type_optimized import (
		_core_process_attendance_logic_optimized,
	)

	frappe.logger().info(f"[Daily Report] Force updating attendance for {report_date_str}")
	_core_process_attendance_logic_optimized(
		employees=[],
		days=[getdate(report_date_str)],
		from_date=report_date_str,
		to_date=report_date_str,
		fore_get_logs=True
	)
	frappe.logger().info(f"[Daily Report] Attendance recalculation complete for {report_date_str}")


def collect_daily_report_context(report_date_str, force_update_attendance=False):
	"""Gather everything the detailed report needs for one day.

	Split out of the send job so the Daily Attendance Report email can embed the
	same detail sections without duplicating this assembly — one place decides
	what "the detail" is, and the two mails cannot drift apart.

	Returns (data, stats).
	"""
	# Recalculate attendance if requested
	if force_update_attendance:
		recalculate_attendance(report_date_str)

	# Import the report get_data function
	from customize_erpnext.customize_erpnext.report.shift_attendance_customize.shift_attendance_customize import get_data

	# Prepare filters for single date report
	filters = {
		"from_date": report_date_str,
		"to_date": report_date_str,
		"summary": 0,
		"detail_join_resign_date": 1
	}

	# Get report data — only employees with the configured prefix
	emp_prefix = _get_employee_prefix()
	data = get_data(filters)
	data = [row for row in data if str(row.get('employee') or '').startswith(emp_prefix)]

	# Calculate statistics
	stats = calculate_attendance_statistics(report_date_str, data)

	# Get incomplete check-ins from day 26 of previous month to yesterday
	from frappe.utils import get_first_day, add_months
	current_month_first = get_first_day(report_date_str)
	prev_month_26 = add_days(add_months(current_month_first, -1), 25)
	yesterday = add_days(report_date_str, -1)
	incomplete_checkins = get_incomplete_checkins(prev_month_26, yesterday)

	# Add incomplete checkins to stats
	stats['incomplete_checkins'] = incomplete_checkins
	stats['incomplete_count'] = len(incomplete_checkins)
	stats['incomplete_processed'] = len([emp for emp in incomplete_checkins if emp.get('manual_checkins', '')])

	# Get employees with status 'Left' but still have checkins on report date
	left_with_checkins = get_left_employees_with_checkins(report_date_str)
	stats['left_with_checkins'] = left_with_checkins
	stats['left_with_checkins_count'] = len(left_with_checkins)

	# Get early-checkout Day-shift employees (7 ≤ working_hours < 8, checkout 16:xx)
	# These may be pregnant/nursing employees not yet registered in Employee Maternity
	early_checkout_list, early_checkout_date = get_early_checkout_day_shift(report_date_str)
	stats['early_checkout_list'] = early_checkout_list
	stats['early_checkout_count'] = len(early_checkout_list)
	stats['early_checkout_date'] = early_checkout_date

	return data, stats


def calculate_attendance_statistics(report_date, data):
	"""
	Calculate statistics for the report:
	- Total active employees
	- Total present
	- Total absent (excluding on leave)
	- Total on leave (On Leave, Half Day)
	- Working hours summary
	"""
	prefix = _get_employee_prefix()
	# Count employees who were active on report_date
	# (not current Active count — some may have left since then)
	total_employees = frappe.db.sql("""
		SELECT COUNT(*) FROM `tabEmployee`
		WHERE (date_of_joining IS NULL OR date_of_joining <= %(date)s)
		  AND (
		      status = 'Active'
		      OR (status = 'Left' AND relieving_date >= %(date)s)			   
		  )
		  AND employee LIKE %(prefix)s
	""", {"date": report_date, "prefix": f"{prefix}%"})[0][0]

	# Separate employees by status
	present_employees = []
	absent_employees = []
	on_leave_employees = []
	maternity_employees = []

	# Per-shift headcount + detail list for specific shifts (email sections)
	shift_stats = {}          # shift -> {"total", "present", "absent", "on_leave"}
	shift_detail_names = ("Shift 2",)
	shift_detail_employees = []

	total_working_hours = 0
	total_actual_overtime = 0
	total_approved_overtime = 0
	total_final_overtime = 0

	# Query Employee Maternity for employees on maternity leave on report_date
	# (maternity_from_date..maternity_to_date only — not pregnant/young_child)
	maternity_emp_set = set(frappe.db.sql_list("""
		SELECT employee FROM `tabEmployee Maternity`
		WHERE maternity_from_date <= %(date)s AND maternity_to_date >= %(date)s
		  AND employee LIKE %(prefix)s
	""", {"date": report_date, "prefix": f"{prefix}%"}))

	# Bulk-load employee details (designation, attendance_device_id) in ONE query
	all_employees = set(row.get('employee') for row in data if row.get('employee'))
	# Also include maternity employees who may have no attendance record
	all_employees.update(maternity_emp_set)
	emp_details_map = {}
	if all_employees:
		emp_details = frappe.db.sql("""
			SELECT name, employee_name, department, custom_group, designation, attendance_device_id
			FROM `tabEmployee`
			WHERE name IN %(employees)s AND status = 'Active'
		""", {"employees": list(all_employees)}, as_dict=True)
		for ed in emp_details:
			emp_details_map[ed.name] = ed

	# Group data by employee
	employee_data = {}
	for row in data:
		employee = row.get('employee')
		if employee not in employee_data:
			employee_data[employee] = row

			status = row.get('status')
			# Status may contain HTML tags, so extract the text
			if isinstance(status, str):
				if 'Present' in status or 'Sunday' in status:
					status_clean = 'Present'
				elif 'On Leave' in status or 'Half Day' in status:
					status_clean = 'On Leave'
				elif 'Absent' in status:
					status_clean = 'Absent'
				else:
					status_clean = status
			else:
				status_clean = status

			working_hours = row.get('working_hours', 0) or 0
			actual_overtime = row.get('actual_overtime_duration', 0) or 0
			approved_overtime = row.get('custom_approved_overtime_duration', 0) or 0
			final_overtime = row.get('final_overtime_duration', 0) or 0

			# Get employee details from pre-loaded map
			emp_doc = emp_details_map.get(employee, {})
			emp_info = {
				'employee': employee,
				'employee_name': row.get('employee_name'),
				'department': row.get('department'),
				'custom_group': row.get('custom_group'),
				'shift': row.get('shift'),
				'leave_type': row.get('leave_type'),
				'leave_application': row.get('leave_application'),
				'half_day_status': row.get('half_day_status'),
				'designation': emp_doc.get('designation') or '',
				'attendance_device_id': emp_doc.get('attendance_device_id') or ''
			}

			if status_clean == 'Present':
				present_employees.append(emp_info)
				total_working_hours += working_hours
				total_actual_overtime += actual_overtime
				total_approved_overtime += approved_overtime
				total_final_overtime += final_overtime
			elif status_clean == 'On Leave':
				on_leave_employees.append(emp_info)
			elif status_clean == 'Absent':
				absent_employees.append(emp_info)

			# Per-shift headcount
			shift_name = row.get('shift') or 'No Shift'
			ss = shift_stats.setdefault(shift_name, {"total": 0, "present": 0, "absent": 0, "on_leave": 0})
			ss["total"] += 1
			if status_clean == 'Present':
				ss["present"] += 1
			elif status_clean == 'Absent':
				ss["absent"] += 1
			elif status_clean == 'On Leave':
				ss["on_leave"] += 1

			# Detail rows for selected shifts (e.g. Shift 2 — starts 14:00, morning
			# email would otherwise only show them mixed into the absent list)
			if shift_name in shift_detail_names:
				in_time = row.get('in_time')
				shift_detail_employees.append({
					**emp_info,
					'status_clean': status_clean,
					'in_time': in_time.strftime('%H:%M') if hasattr(in_time, 'strftime') else (str(in_time)[11:16] if in_time else '')
				})

	# Add maternity employees from EM who have no attendance record (expected after simplification)
	# Also move any maternity employee who appeared as Absent into maternity list
	absent_emp_names = {e['employee'] for e in absent_employees}
	for emp in maternity_emp_set:
		emp_doc = emp_details_map.get(emp, {})
		if not emp_doc:
			continue  # Skip inactive or unknown employees
		mat_info = {
			'employee': emp,
			'employee_name': emp_doc.get('employee_name', ''),
			'department': emp_doc.get('department', ''),
			'custom_group': emp_doc.get('custom_group', ''),
			'shift': '',
			'leave_type': None,
			'leave_application': None,
			'half_day_status': None,
			'designation': emp_doc.get('designation') or '',
			'attendance_device_id': emp_doc.get('attendance_device_id') or ''
		}
		if emp not in employee_data:
			# No attendance record → add directly to maternity list
			maternity_employees.append(mat_info)
		elif emp in absent_emp_names:
			# Was marked Absent (old data before fix) → move to maternity list
			absent_employees = [e for e in absent_employees if e['employee'] != emp]
			maternity_employees.append(mat_info)

	# Sort lists by custom_group (A-Z)
	present_employees = sorted(present_employees, key=lambda x: (x.get("custom_group") or "").lower())
	absent_employees = sorted(absent_employees, key=lambda x: (x.get("custom_group") or "").lower())
	on_leave_employees = sorted(on_leave_employees, key=lambda x: (x.get("custom_group") or "").lower())
	maternity_employees = sorted(maternity_employees, key=lambda x: (x.get("custom_group") or "").lower())

	# Calculate counts
	total_present = len(present_employees)
	total_absent = len(absent_employees)
	on_leave_count = len(on_leave_employees)
	maternity_count = len(maternity_employees)

	return {
		"total_employees": total_employees,
		"total_present": total_present,
		"total_absent": total_absent,
		"on_leave_count": on_leave_count,
		"maternity_count": maternity_count,
		"total_working_hours": round(total_working_hours, 2),
		"total_actual_overtime": round(total_actual_overtime, 2),
		"total_approved_overtime": round(total_approved_overtime, 2),
		"total_final_overtime": round(total_final_overtime, 2),
		"present_employees": present_employees,
		"absent_employees": absent_employees,
		"on_leave_employees": on_leave_employees,
		"maternity_employees": maternity_employees,
		"shift_stats": shift_stats,
		"shift_detail_employees": sorted(shift_detail_employees, key=lambda x: (x.get("custom_group") or "").lower())
	}


def _timedelta_to_time(td):
	"""Convert timedelta to time object"""
	if isinstance(td, timedelta):
		total_seconds = int(td.total_seconds())
		return time_obj(total_seconds // 3600, (total_seconds % 3600) // 60, total_seconds % 60)
	return td


def get_incomplete_checkins(start_date, end_date):
	"""
	Get list of employees who have incomplete check-ins from start_date to end_date.
	Optimized: eliminates correlated subqueries and N+1 query patterns.
	"""
	from collections import defaultdict

	# Step 1: Pre-load Shift Assignments for the date range (1 query)
	# Shift Registration is deprecated/unused — shift overrides now come from
	# the standard Shift Assignment doctype (end_date NULL = open-ended).
	range_start = getdate(start_date)
	range_end = getdate(end_date)
	shift_reg_map = {}  # (employee, date) -> shift
	# No status filter — HRMS auto-marks expired assignments "Inactive";
	# they remain valid for their historical date range
	shift_regs = frappe.db.sql("""
		SELECT employee, shift_type AS shift, start_date, end_date
		FROM `tabShift Assignment`
		WHERE docstatus = 1
		  AND start_date <= %(end_date)s
		  AND (end_date IS NULL OR end_date >= %(start_date)s)
		ORDER BY start_date DESC, creation DESC
	""", {"start_date": start_date, "end_date": end_date}, as_dict=True)

	# Build map: (employee, date) -> shift (latest assignment wins).
	# Clamp each assignment to the report range; open-ended ones run to range_end.
	for sr in shift_regs:
		emp = sr.employee
		d = max(getdate(sr.start_date), range_start)
		sr_end = min(getdate(sr.end_date), range_end) if sr.end_date else range_end
		while d <= sr_end:
			key = (emp, d)
			if key not in shift_reg_map:
				shift_reg_map[key] = sr.shift
			d += timedelta(days=1)

	# Step 2: Pre-load shift begin/end times (1 query)
	shift_times = {}
	for sn in frappe.db.sql("""
		SELECT shift_name, begin_time, end_time FROM `tabShift Name`
	""", as_dict=True):
		shift_times[sn.shift_name] = {
			'begin_time': _timedelta_to_time(sn.begin_time),
			'end_time': _timedelta_to_time(sn.end_time)
		}

	# Step 3: Main query - simple aggregation, no correlated subqueries
	employees_with_checkins = frappe.db.sql("""
		SELECT
			e.attendance_device_id,
			e.name AS employee_code,
			e.employee_name,
			e.department,
			e.custom_group,
			e.designation,
			DATE(ec.time) AS checkin_date,
			MIN(ec.time) AS first_check_in,
			MAX(ec.time) AS last_check_out,
			COUNT(ec.name) AS checkin_count
		FROM `tabEmployee` e
		INNER JOIN `tabEmployee Checkin` ec
			ON e.name = ec.employee
			AND ec.time >= %(start)s
			AND ec.time < %(end)s
			AND ec.device_id IS NOT NULL
		WHERE e.status = 'Active'
		  AND e.name LIKE %(prefix)s
		GROUP BY e.name, DATE(ec.time)
		ORDER BY DATE(ec.time) DESC, e.name ASC
	""", {
		"start": f"{start_date} 00:00:00",
		"end": f"{end_date} 23:59:59",
		"prefix": f"{_get_employee_prefix()}%"
	}, as_dict=True)

	# Step 4: Resolve shift for each row using pre-loaded map
	for emp in employees_with_checkins:
		key = (emp.employee_code, emp.checkin_date)
		shift = shift_reg_map.get(key)
		if not shift:
			# Default shift is 'Day'; non-Day shifts (incl. Canteen) come from
			# Shift Assignment, not from custom_group.
			shift = 'Day'
		emp['shift'] = shift
		times = shift_times.get(shift, {})
		emp['begin_time'] = times.get('begin_time')
		emp['end_time'] = times.get('end_time')

	# Step 5: Pre-load ALL checkin times for candidates with >= 2 checkins (1 bulk query)
	# Also pre-load all manual checkins for the date range
	candidates = [(emp.get('employee_code'), emp.get('checkin_date')) for emp in employees_with_checkins]
	if not candidates:
		return []

	# Get unique employees
	candidate_employees = list(set(c[0] for c in candidates))

	# Bulk load all device checkins (for Rule 2/3 check)
	all_device_checkins = defaultdict(list)  # (employee, date) -> [time, ...]
	device_rows = frappe.db.sql("""
		SELECT employee, time
		FROM `tabEmployee Checkin`
		WHERE employee IN %(employees)s
		  AND time >= %(start)s
		  AND time < %(end)s
		  AND device_id IS NOT NULL
		ORDER BY employee, time
	""", {
		"employees": candidate_employees,
		"start": f"{start_date} 00:00:00",
		"end": f"{end_date} 23:59:59"
	}, as_dict=True)
	for row in device_rows:
		key = (row.employee, row.time.date())
		all_device_checkins[key].append(row.time)

	# Bulk load all manual checkins
	all_manual_checkins = defaultdict(list)  # (employee, date) -> [{...}, ...]
	manual_rows = frappe.db.sql("""
		SELECT employee, TIME(time) as checkin_time, time,
			custom_reason_for_manual_check_in,
			custom_other_reason_for_manual_check_in
		FROM `tabEmployee Checkin`
		WHERE employee IN %(employees)s
		  AND time >= %(start)s
		  AND time < %(end)s
		  AND device_id IS NULL
		ORDER BY employee, time
	""", {
		"employees": candidate_employees,
		"start": f"{start_date} 00:00:00",
		"end": f"{end_date} 23:59:59"
	}, as_dict=True)
	for row in manual_rows:
		key = (row.employee, row.time.date())
		all_manual_checkins[key].append(row)

	# Step 6: Filter incomplete check-ins using pre-loaded data
	incomplete_list = []
	for emp in employees_with_checkins:
		is_incomplete = False
		checkin_count = emp.get('checkin_count')
		checkin_date = emp.get('checkin_date')
		employee_code = emp.get('employee_code')

		# Rule 1: Only 1 check-in
		if checkin_count == 1:
			is_incomplete = True
		# Rule 2 & 3: >= 2 check-ins but all before begin_time OR all after end_time
		elif checkin_count >= 2:
			begin_time = emp.get('begin_time')
			end_time = emp.get('end_time')

			checkin_times = [t.time() for t in all_device_checkins.get((employee_code, checkin_date), [])]

			# Rule 2: All check-ins <= begin_time
			if begin_time and checkin_times and isinstance(begin_time, time_obj):
				if all(ct <= begin_time for ct in checkin_times):
					is_incomplete = True

			# Rule 3: All check-ins >= end_time
			if end_time and checkin_times and not is_incomplete and isinstance(end_time, time_obj):
				if all(ct >= end_time for ct in checkin_times):
					is_incomplete = True

		if is_incomplete:
			# Use pre-loaded manual checkins
			manual_checkins = all_manual_checkins.get((employee_code, checkin_date), [])

			manual_checkin_times = []
			reasons = []
			other_reasons = []

			for mc in manual_checkins:
				if mc.get('checkin_time'):
					manual_checkin_times.append(str(mc.get('checkin_time')))
				if mc.get('custom_reason_for_manual_check_in'):
					reasons.append(mc.get('custom_reason_for_manual_check_in'))
				if mc.get('custom_other_reason_for_manual_check_in'):
					other_reasons.append(mc.get('custom_other_reason_for_manual_check_in'))

			emp['manual_checkins'] = ', '.join(manual_checkin_times) if manual_checkin_times else ''
			emp['reason_for_manual'] = ', '.join(set(reasons)) if reasons else ''
			emp['other_reason_for_manual'] = ', '.join(set(other_reasons)) if other_reasons else ''

			incomplete_list.append(emp)

	return incomplete_list


def get_early_checkout_day_shift(report_date):
	"""
	Lấy các attendance ngày hôm qua (ca Day) có:
	  - last checkout 16:00–16:59
	  - 7 <= working_hours < 8
	Thứ Hai → lấy thứ Bảy tuần trước (bỏ qua CN).
	Trả về (list, date_str).
	"""
	from frappe.utils import getdate

	yesterday = add_days(report_date, -1)
	yesterday_date = getdate(yesterday)
	# Thứ Hai (weekday=0): hôm qua là CN → dùng thứ Bảy
	if yesterday_date.weekday() == 6:
		yesterday = add_days(yesterday, -1)
		yesterday_date = getdate(yesterday)

	yesterday_str = str(yesterday_date)

	try:
		rows = frappe.db.sql("""
			SELECT
				e.attendance_device_id,
				a.employee,
				e.employee_name,
				e.department,
				e.custom_group,
				e.designation,
				a.attendance_date,
				a.working_hours,
				a.shift,
				MAX(ec.time) AS last_checkout
			FROM `tabAttendance` a
			JOIN `tabEmployee` e ON e.name = a.employee
			LEFT JOIN `tabEmployee Checkin` ec
				ON ec.employee = a.employee
				AND DATE(ec.time) = a.attendance_date
			WHERE a.attendance_date = %(yesterday)s
			  AND a.shift LIKE 'Day%%'
			  AND a.working_hours >= 7
			  AND a.working_hours < 8
			  AND a.docstatus = 1
			  AND e.name LIKE %(prefix)s
			GROUP BY
				a.name, a.employee, a.attendance_date, a.working_hours, a.shift,
				e.attendance_device_id, e.employee_name, e.department,
				e.custom_group, e.designation
			HAVING TIME(MAX(ec.time)) >= '16:00:00'
			   AND TIME(MAX(ec.time)) <  '17:00:00'
			ORDER BY e.custom_group, e.employee_name
		""", {"yesterday": yesterday_str, "prefix": f"{_get_employee_prefix()}%"}, as_dict=True)
		return rows, yesterday_str
	except Exception as e:
		frappe.logger().error(f"Error in get_early_checkout_day_shift: {str(e)}")
		return [], yesterday_str


def get_left_employees_with_checkins(report_date):
	"""
	Get employees with status 'Left' who still have checkins on report_date.
	Used to alert HR to review these records.
	"""
	try:
		rows = frappe.db.sql("""
			SELECT
				e.attendance_device_id,
				e.name AS employee_code,
				e.employee_name,
				e.department,
				e.custom_group,
				e.designation,
				e.date_of_joining,
				e.relieving_date,
				MIN(ec.time) AS first_check_in,
				MAX(ec.time) AS last_check_out,
				COUNT(ec.name) AS checkin_count
			FROM `tabEmployee` e
			INNER JOIN `tabEmployee Checkin` ec
				ON e.name = ec.employee
				AND DATE(ec.time) = %(report_date)s
			WHERE e.status = 'Left'
			  AND e.name LIKE %(prefix)s
			GROUP BY e.name
			ORDER BY e.custom_group, e.employee_name
		""", {"report_date": report_date, "prefix": f"{_get_employee_prefix()}%"}, as_dict=True)
		return rows
	except Exception as e:
		frappe.logger().error(f"Error getting left employees with checkins: {str(e)}")
		return []


def _fmt_date(value):
	return formatdate(value, "dd/MM/yyyy") if value else ""


def _fmt_time(value):
	"""Datetimes come back from SQL as datetime, MIN()/MAX() sometimes as str."""
	if not value:
		return ""
	return value.strftime("%H:%M:%S") if hasattr(value, "strftime") else str(value)[:8]


def _add_anomaly_sheet(wb, title, headers, rows, header_font, header_fill, header_alignment, border):
	"""One flat sheet with a numbered header row; rows are tuples without the STT."""
	ws = wb.create_sheet(title[:31])

	for col, header in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col)
		cell.value = header
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = header_alignment
		cell.border = border

	for idx, row in enumerate(rows, 1):
		for col, value in enumerate((idx, *row), 1):
			cell = ws.cell(row=idx + 1, column=col)
			cell.value = value
			cell.border = border

	ws.column_dimensions["A"].width = 6
	for col in range(2, len(headers) + 1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

	return ws


def generate_excel_report(report_date, data, stats):
	"""
	Generate Excel file with 2 sheets:
	1. Absent-Maternity Leave-Present: All attendance data for the day
	2. Missing check-ins from day 26 of previous month to yesterday
	"""
	# Create a new workbook
	wb = openpyxl.Workbook()

	# Remove default sheet
	wb.remove(wb.active)

	# Define styles
	header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
	header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
	header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

	cell_font = Font(name='Arial', size=10)
	cell_alignment = Alignment(horizontal='left', vertical='center')
	center_alignment = Alignment(horizontal='center', vertical='center')

	border = Border(
		left=Side(style='thin', color='000000'),
		right=Side(style='thin', color='000000'),
		top=Side(style='thin', color='000000'),
		bottom=Side(style='thin', color='000000')
	)

	formatted_date = formatdate(report_date, "dd/MM/yyyy")

	# Sheet 1: All attendance data (Absent-OnLeave-Present)
	ws1 = wb.create_sheet("Absent-OnLeave-Present")

	# Combine all employee data sorted by status
	all_data = []
	all_data.extend(stats.get('absent_employees', []))
	all_data.extend(stats.get('maternity_employees', []))
	all_data.extend(stats.get('on_leave_employees', []))
	all_data.extend(stats.get('present_employees', []))

	# Add headers for Sheet 1
	headers1 = ["STT", "Ngày", "Att ID", "Employee", "Employee Name", "Department", "Group", "Shift", "Designation", "Leave Type", "Leave Application", "Status for Other Half", "Status"]
	for col_num, header in enumerate(headers1, 1):
		cell = ws1.cell(row=1, column=col_num)
		cell.value = header
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = header_alignment
		cell.border = border

	# Add data for Sheet 1
	for idx, emp in enumerate(all_data, 1):
		# Determine status
		if emp in stats.get('absent_employees', []):
			status = 'Absent'
		elif emp in stats.get('maternity_employees', []):
			status = 'Maternity Leave'
		elif emp in stats.get('on_leave_employees', []):
			status = 'On Leave'
		else:
			status = 'Present'

		row_data = [
			idx,
			formatted_date,
			emp.get('attendance_device_id') or '',
			emp.get('employee') or '',
			emp.get('employee_name') or '',
			emp.get('department') or '',
			emp.get('custom_group') or '',
			emp.get('shift') or '',
			emp.get('designation') or '',
			emp.get('leave_type') or '',
			emp.get('leave_application') or '',
			emp.get('half_day_status') or '',
			status
		]

		for col_num, value in enumerate(row_data, 1):
			cell = ws1.cell(row=idx + 1, column=col_num)
			cell.value = value
			cell.font = cell_font
			if col_num in [1, 2]:  # STT and Date columns
				cell.alignment = center_alignment
			else:
				cell.alignment = cell_alignment
			cell.border = border

	# Adjust column widths for Sheet 1
	ws1.column_dimensions['A'].width = 6   # STT
	ws1.column_dimensions['B'].width = 12  # Date
	ws1.column_dimensions['C'].width = 10  # Att ID
	ws1.column_dimensions['D'].width = 12  # Employee
	ws1.column_dimensions['E'].width = 25  # Employee Name
	ws1.column_dimensions['F'].width = 20  # Department
	ws1.column_dimensions['G'].width = 15  # Group
	ws1.column_dimensions['H'].width = 12  # Shift
	ws1.column_dimensions['I'].width = 20  # Designation
	ws1.column_dimensions['J'].width = 15  # Leave Type
	ws1.column_dimensions['K'].width = 20  # Leave Application
	ws1.column_dimensions['L'].width = 15  # Status

	# Sheet 2: Incomplete check-ins
	from frappe.utils import get_first_day, add_months
	current_month_first = get_first_day(report_date)
	prev_month_26 = add_days(add_months(current_month_first, -1), 25)
	yesterday = add_days(report_date, -1)
	from_date_str = formatdate(prev_month_26, "dd/MM/yyyy")
	to_date_str = formatdate(yesterday, "dd/MM/yyyy")

	# Sheet name cannot contain / and Excel refuses names over 31 characters —
	# the full "Missing dd-mm-yyyy to dd-mm-yyyy" is 32, so drop the century.
	sheet_name = f"Missing {from_date_str} to {to_date_str}".replace('/', '-')[:31]
	ws2 = wb.create_sheet(sheet_name)

	# Add headers for Sheet 2
	headers2 = ["STT", "Ngày", "Att ID", "Employee", "Employee Name", "Department", "Group", "Shift", "Designation",
				"Check-in", "Check-out", "Số lần chấm", "Đã xử lý", "Reason", "Other Reason"]
	for col_num, header in enumerate(headers2, 1):
		cell = ws2.cell(row=1, column=col_num)
		cell.value = header
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = header_alignment
		cell.border = border

	# Add data for Sheet 2
	incomplete_list = stats.get('incomplete_checkins', [])
	for idx, emp in enumerate(incomplete_list, 1):
		checkin_date = emp.get('checkin_date')
		checkin_date_formatted = formatdate(checkin_date, "dd/MM/yyyy") if checkin_date else ""
		checkin_count = emp.get('checkin_count') or 0

		# Determine check-in and check-out times (same logic as email)
		first_checkin = ""
		last_checkout = ""

		if checkin_count == 1:
			single_time = emp.get("first_check_in")
			if single_time:
				if isinstance(single_time, str):
					single_time = get_datetime(single_time)
				formatted_time = single_time.strftime("%H:%M:%S")
				first_checkin = formatted_time
		else:
			first_checkin_time = emp.get("first_check_in")
			if first_checkin_time:
				if isinstance(first_checkin_time, str):
					first_checkin_time = get_datetime(first_checkin_time)
				first_checkin = first_checkin_time.strftime("%H:%M:%S")

			last_checkout_time = emp.get("last_check_out")
			if last_checkout_time:
				if isinstance(last_checkout_time, str):
					last_checkout_time = get_datetime(last_checkout_time)
				last_checkout = last_checkout_time.strftime("%H:%M:%S")

		row_data = [
			idx,
			checkin_date_formatted,
			emp.get('attendance_device_id') or '',
			emp.get('employee_code') or '',
			emp.get('employee_name') or '',
			emp.get('department') or '',
			emp.get('custom_group') or '',
			emp.get('shift') or '',
			emp.get('designation') or '',
			first_checkin,
			last_checkout,
			checkin_count,
			emp.get('manual_checkins', ''),
			emp.get('reason_for_manual', ''),
			emp.get('other_reason_for_manual', '')
		]

		for col_num, value in enumerate(row_data, 1):
			cell = ws2.cell(row=idx + 1, column=col_num)
			cell.value = value
			cell.font = cell_font
			if col_num in [1, 2, 10, 11, 12, 13]:  # STT, Date, Check-in, Check-out, Count, Manual columns
				cell.alignment = center_alignment
			else:
				cell.alignment = cell_alignment
			cell.border = border

	# Adjust column widths for Sheet 2
	ws2.column_dimensions['A'].width = 6   # STT
	ws2.column_dimensions['B'].width = 12  # Date
	ws2.column_dimensions['C'].width = 10  # Att ID
	ws2.column_dimensions['D'].width = 12  # Employee
	ws2.column_dimensions['E'].width = 25  # Employee Name
	ws2.column_dimensions['F'].width = 20  # Department
	ws2.column_dimensions['G'].width = 15  # Group
	ws2.column_dimensions['H'].width = 12  # Shift
	ws2.column_dimensions['I'].width = 20  # Designation
	ws2.column_dimensions['J'].width = 12  # Check-in
	ws2.column_dimensions['K'].width = 12  # Check-out
	ws2.column_dimensions['L'].width = 12  # Số lần chấm
	ws2.column_dimensions['M'].width = 15  # Đã xử lý
	ws2.column_dimensions['N'].width = 20  # Reason
	ws2.column_dimensions['O'].width = 20  # Other Reason

	# Sheets 3 and 4 — the two anomaly lists. They used to exist only in the email
	# body; now that the detail travels as this workbook, they have to live here
	# or they would simply be lost.
	_add_anomaly_sheet(
		wb,
		"Left with check-ins",
		["STT", "Att ID", "Employee", "Employee Name", "Department", "Group",
		 "Designation", "Ngày nghỉ việc", "Check-in đầu", "Check-out cuối", "Số lần"],
		[
			(
				emp.get('attendance_device_id'),
				emp.get('employee_code'),
				emp.get('employee_name'),
				emp.get('department'),
				emp.get('custom_group'),
				emp.get('designation'),
				_fmt_date(emp.get('relieving_date')),
				_fmt_time(emp.get('first_check_in')),
				_fmt_time(emp.get('last_check_out')),
				emp.get('checkin_count') or 0,
			)
			for emp in (stats.get('left_with_checkins') or [])
		],
		header_font, header_fill, header_alignment, border,
	)

	early_date = stats.get('early_checkout_date')
	_add_anomaly_sheet(
		wb,
		"Early checkout day shift",
		["STT", "Ngày", "Att ID", "Employee", "Employee Name", "Department", "Group",
		 "Shift", "Designation", "Check-out cuối", "Giờ công"],
		[
			(
				formatdate(early_date, "dd/MM/yyyy") if early_date else "",
				emp.get('attendance_device_id'),
				emp.get('employee'),
				emp.get('employee_name'),
				emp.get('department'),
				emp.get('custom_group'),
				emp.get('shift'),
				emp.get('designation'),
				_fmt_time(emp.get('last_checkout')),
				round(emp.get('working_hours') or 0, 2),
			)
			for emp in (stats.get('early_checkout_list') or [])
		],
		header_font, header_fill, header_alignment, border,
	)

	# Save to temporary file
	temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
	wb.save(temp_file.name)
	temp_file.close()

	file_name = f"Attendance_Report_{formatted_date.replace('/', '')}.xlsx"

	return temp_file.name, file_name
