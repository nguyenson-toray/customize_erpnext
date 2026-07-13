# Copyright (c) 2025, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.utils import cint, flt


# ============================================================================
# EXPORT CONFIG & SHARED STYLES
# ============================================================================
# Sync/async threshold. Benchmark 2026-07-05 (1021 emp x 30 days = 25k rows):
# full pipeline 22.4s (query 1.5s, 2 legacy sheets 10.5s, 4 new sheets 3.7s,
# save 6.1s) -> 45k records ~ 35-40s, still safe for a sync web request.
EXPORT_ASYNC_THRESHOLD = 45000

# Shared openpyxl style objects — constructing Font/Alignment per cell inside
# the populate loops costs seconds on 25k+ row exports (openpyxl styles are
# immutable and safe to share)
from openpyxl.styles import Font as _XFont, Alignment as _XAlignment
STYLE_CELL_FONT = _XFont(name='Times New Roman', size=8)
STYLE_ALIGN_CENTER = _XAlignment(horizontal='center', vertical='center')
STYLE_ALIGN_LEFT = _XAlignment(horizontal='left', vertical='center')


def execute(filters=None):
	columns = get_columns(filters)
	data = get_data(filters)
	report_summary = get_report_summary(data, filters)
	return columns, data, None, None, report_summary


def get_columns(filters):
	"""Define columns based on Excel file structure"""
	is_summary = filters and filters.get("summary")

	if is_summary:
		# Summary mode: show aggregated data
		columns = [
			{
				"label": _("Shift"),
				"fieldname": "shift",
				"fieldtype": "Link",
				"options": "Shift Type",
				"width": 100,
			},
			{
				"label": _("Employee"),
				"fieldname": "employee",
				"fieldtype": "Link",
				"options": "Employee",
				"width": 110,
			},
			{
				"fieldname": "employee_name",
				"fieldtype": "Data",
				"label": _("Employee Name"),
				"width": 170,
				# "hidden": 1,
			},
			{
				"label": _("Group"),
				"fieldname": "custom_group",
				"fieldtype": "Link",
				"options": "Group",
				"width": 120,
			},
			{
				"label": _("Date of Joining"),
				"fieldname": "date_of_joining",
				"fieldtype": "Date",
				"width": 100,
			},
			{
				"label": _("Relieving Date"),
				"fieldname": "relieving_date",
				"fieldtype": "Date",
				"width": 100,
			},
			{
				"label": _("Total Working Hours"),
				"fieldname": "working_hours",
				"fieldtype": "Float",
				"width": 120,
			},
			{
				"label": _("Total Working Day"),
				"fieldname": "working_day",
				"fieldtype": "Float",
				"width": 120,
			},
			{
				"label": _("Total Actual OT"),
				"fieldname": "actual_overtime_duration",
				"fieldtype": "Float",
				"width": 120,
			},
			{
				"label": _("Total Approved OT"),
				"fieldname": "custom_approved_overtime_duration",
				"fieldtype": "Float",
				"width": 120,
			},
			{
				"label": _("Total Final OT"),
				"fieldname": "final_overtime_duration",
				"fieldtype": "Float",
				"width": 120,
			},
			{
				"label": _("Department"),
				"fieldname": "department",
				"fieldtype": "Link",
				"options": "Department",
				"width": 150,
			},
		]
	else:
		# Detail mode: show individual attendance records
		columns = [
			{
				"label": _("Date"),
				"fieldname": "attendance_date",
				"fieldtype": "Date",
				"width": 100,
			},
			{
				"label": _("Shift"),
				"fieldname": "shift",
				"fieldtype": "Link",
				"options": "Shift Type",
				"width": 70,
			},
			{
				"label": _("Employee"),
				"fieldname": "employee",
				"fieldtype": "Link",
				"options": "Employee",
				"width": 230,
			},
			{
				"fieldname": "employee_name",
				"fieldtype": "Data",
				"label": _("Employee Name"),
				"width": 160,
				"hidden": 1,
			},
			{
				"label": _("Group"),
				"fieldname": "custom_group",
				"fieldtype": "Link",
				"options": "Group",
				"width": 120,
			},
			{
				"label": _("Status"),
				"fieldname": "status",
				"fieldtype": "Data",
				"width": 90,
			},
			{
				"label": _("Leave Application"),
				"fieldname": "leave_application",
				"fieldtype": "Link",
				"options": "Leave Application",
				"width": 150,
			},
			{
				"label": _("Leave Type"),
				"fieldname": "leave_type",
				"fieldtype": "Link",
				"options": "Leave Type",
				"width": 120,
			},
			{
				"label": _("Status for Other Half"),
				"fieldname": "half_day_status",
				"fieldtype": "Data",
				"width": 150,
			},
			{
				"label": _("In Time"),
				"fieldname": "in_time",
				"fieldtype": "Data",
				"width": 100,
			},
			{
				"label": _("Out Time"),
				"fieldname": "out_time",
				"fieldtype": "Data",
				"width": 100,
			},
			{
				"label": _("Working Hours"),
				"fieldname": "working_hours",
				"fieldtype": "Float",
				"width": 100,
			},
			{
				"label": _("Working Day"),
				"fieldname": "working_day",
				"fieldtype": "Float",
				"width": 100,
			},
			{
				"label": _("Actual OT"),
				"fieldname": "actual_overtime_duration",
				"fieldtype": "Float",
				"width": 100,
			},
			{
				"label": _("Approved OT"),
				"fieldname": "custom_approved_overtime_duration",
				"fieldtype": "Float",
				"width": 100,
			},
			{
				"label": _("Final OT"),
				"fieldname": "final_overtime_duration",
				"fieldtype": "Float",
				"width": 100,
			},
			{
				"label": _("Maternity Benefit"),
				"fieldname": "custom_maternity_benefit",
				"fieldtype": "Data",
				"width": 100,
			},
			{
				"label": _("Late Entry"),
				"fieldname": "late_entry",
				"fieldtype": "Check",
				"width": 60,
			},
			{
				"label": _("Early Exit"),
				"fieldname": "early_exit",
				"fieldtype": "Check",
				"width": 60,
			},
			{
				"label": _("Department"),
				"fieldname": "department",
				"fieldtype": "Link",
				"options": "Department",
				"width": 150,
			},
			{
				"label": _("Attendance ID"),
				"fieldname": "name",
				"fieldtype": "Link",
				"options": "Attendance",
				"width": 120,
			},
		]

		# Add Join/Resign Date columns if filter is enabled
		if filters and filters.get("detail_join_resign_date"):
			join_resign_columns = [
				{
					"label": _("Date of Joining"),
					"fieldname": "date_of_joining",
					"fieldtype": "Date",
					"width": 100,
				},
				{
					"label": _("Relieving Date"),
					"fieldname": "relieving_date",
					"fieldtype": "Date",
					"width": 100,
				},
			]
			# Insert after Early Exit column (index 15)
			columns[15:15] = join_resign_columns

		# Hide Leave Application columns if filter is not enabled
		if not (filters and filters.get("show_leave_application")):
			columns = [col for col in columns if col.get("fieldname") not in
					   ["leave_application", "leave_type", "half_day_status"]]

	return columns


def get_data(filters):
	query = get_query(filters)
	data = query.run(as_dict=True)
	data = update_data(data)
	return data


def get_report_summary(data, filters):
	"""Generate report summary with required statistics"""
	# Hide report summary in Summary mode
	if filters and filters.get("summary"):
		return None

	if not data:
		return None 

	present_records = on_leave_records = maternity_records = absent_records = late_entries = early_exits = 0

	# Maternity count from Employee Maternity (source of truth — no Leave Application needed)
	from_date = filters.get("from_date") if filters else None
	to_date = (filters.get("to_date") if filters else None) or from_date
	if from_date:
		maternity_records = frappe.db.count("Employee Maternity", filters=[
			["maternity_from_date", "<=", to_date],
			["maternity_to_date", ">=", from_date],
		])

	for entry in data:
		# Check status (may contain HTML tags after formatting)
		status_text = entry.get("status")
		if isinstance(status_text, str):
			if "Present" in status_text:
				present_records += 1
			elif "On Leave" in status_text or "Half Day" in status_text:
				on_leave_records += 1
			elif "Absent" in status_text:
				absent_records += 1

		if entry.get("late_entry"):
			late_entries += 1
		if entry.get("early_exit"):
			early_exits += 1

	return [
		{
			"value": present_records,
			"indicator": "Green",
			"label": _("Present Records"),
			"datatype": "Int",
		},
		{
			"value": maternity_records,
			"indicator": "Pink",
			"label": _("Maternity Leave"),
			"datatype": "Int",
		},
		{
			"value": on_leave_records,
			"indicator": "Blue",
			"label": _("On Leave Records"),
			"datatype": "Int",
		},
		{
			"value": absent_records,
			"indicator": "Red",
			"label": _("Absent Records"),
			"datatype": "Int",
		},
		{
			"value": late_entries,
			"indicator": "Red",
			"label": _("Late Entries"),
			"datatype": "Int",
		},
		{
			"value": early_exits,
			"indicator": "Red",
			"label": _("Early Exits"),
			"datatype": "Int",
		},
	]


def get_query(filters):
	"""Build query to fetch attendance records - SIMPLE query from Attendance only"""
	from pypika.functions import Sum

	attendance = frappe.qb.DocType("Attendance")
	employee = frappe.qb.DocType("Employee")
	is_summary = filters and filters.get("summary")

	if is_summary:
		# Summary mode: aggregate data by employee
		query = (
			frappe.qb.from_(attendance)
			.left_join(employee)
			.on(attendance.employee == employee.name)
			.select(
				attendance.shift,
				attendance.employee,
				attendance.employee_name,
				employee.custom_group,
				employee.date_of_joining,
				employee.relieving_date,
				attendance.department,
				Sum(attendance.working_hours).as_("working_hours"),
			)
			.where(attendance.docstatus == 1)
			.groupby(
				attendance.shift,
				attendance.employee,
				attendance.employee_name,
				employee.custom_group,
				employee.date_of_joining,
				employee.relieving_date,
				attendance.department
			)
			.orderby(attendance.shift)
			.orderby(employee.custom_group)
			.orderby(attendance.employee)
		)

		# Add custom overtime fields aggregation if they exist
		try:
			query = query.select(
				Sum(attendance.actual_overtime_duration).as_("actual_overtime_duration"),
				Sum(attendance.custom_approved_overtime_duration).as_("custom_approved_overtime_duration"),
				Sum(attendance.custom_final_overtime_duration).as_("final_overtime_duration"),
			)
		except AttributeError:
			pass

	else:
		# Detail mode: show individual attendance records
		query = (
			frappe.qb.from_(attendance)
			.left_join(employee)
			.on(attendance.employee == employee.name)
			.select(
				attendance.name,
				attendance.employee,
				attendance.employee_name,
				attendance.shift,
				attendance.attendance_date,
				attendance.status,
				attendance.leave_type,
				attendance.half_day_status,
				attendance.in_time,
				attendance.out_time,
				attendance.working_hours,
				attendance.late_entry,
				attendance.early_exit,
				attendance.department,
				attendance.company,
				attendance.leave_application,
				employee.custom_group,
				employee.date_of_joining,
				employee.relieving_date,
			)
			.where(attendance.docstatus == 1)
			.orderby(attendance.attendance_date)
			.orderby(attendance.shift)
			.orderby(employee.custom_group)
			.orderby(attendance.employee)
		)

		# Add custom fields if they exist
		try:
			query = query.select(
				attendance.actual_overtime_duration.as_("actual_overtime_duration"),
				attendance.custom_approved_overtime_duration,
				attendance.custom_final_overtime_duration.as_("final_overtime_duration"),
				attendance.custom_maternity_benefit,
				attendance.custom_leave_application_abbreviation,
				attendance.overtime_type,
				# Extra fields for the Excel export sheets (Detail/Summary/Shift/Important Note)
				attendance.custom_note,
				employee.attendance_device_id,
				employee.custom_section,
			)
		except AttributeError:
			pass

	# Apply filters
	for filter_key in filters:
		if filter_key == "from_date":
			query = query.where(attendance.attendance_date >= filters['from_date'])
		elif filter_key == "to_date":
			query = query.where(attendance.attendance_date <= filters['to_date'])
		elif filter_key == "late_entry" and not is_summary:
			query = query.where(attendance.late_entry == 1)
		elif filter_key == "early_exit" and not is_summary:
			query = query.where(attendance.early_exit == 1)
		elif filter_key == "status" and not is_summary:
			query = query.where(attendance.status == filters['status'])
		elif filter_key == "group":
			query = query.where(employee.custom_group == filters['group'])
		elif filter_key == "employee":
			query = query.where(attendance.employee == filters['employee'])
		elif filter_key == "shift":
			query = query.where(attendance.shift == filters['shift'])
		elif filter_key == "department":
			query = query.where(attendance.department == filters['department'])
		# Skip summary and detail_join_resign_date filters in query
		elif filter_key in ["summary", "detail_join_resign_date"]:
			continue

	return query


def update_data(data):
	"""Update data with calculated fields and formatting"""
	for d in data:
		# Calculate Working Day = Total Working Hours / 8
		if d.get("working_hours"):
			d.working_day = flt(d.working_hours / 8.0, 2)
		else:
			d.working_day = 0

		# Format working hours
		d.working_hours = format_float_precision(d.working_hours)

		# Format actual and final overtime duration
		if d.get("actual_overtime_duration"):
			d.actual_overtime_duration = format_float_precision(d.actual_overtime_duration)
		if d.get("custom_approved_overtime_duration"):
			d.custom_approved_overtime_duration = format_float_precision(d.custom_approved_overtime_duration)
		if d.get("final_overtime_duration"):
			d.final_overtime_duration = format_float_precision(d.final_overtime_duration)

		# Format in/out time - simple string format (only in detail mode)
		if d.get("in_time") and hasattr(d.in_time, 'strftime'):
			d.in_time = d.in_time.strftime('%H:%M:%S')
		if d.get("out_time") and hasattr(d.out_time, 'strftime'):
			d.out_time = d.out_time.strftime('%H:%M:%S')

		# Add color styling for status (only in detail mode)
		if d.get("status"):
			if d.status == "Present":
				d.status = '<span style="color: green; font-weight: bold;">Present</span>'
			elif d.status == "Absent":
				d.status = '<span style="color: red; font-weight: bold;">Absent</span>'
			elif d.status == "On Leave":
				d.status = '<span style="color: blue; font-weight: bold;">On Leave</span>'
			elif d.status == "Half Day":
				d.status = '<span style="color: orange; font-weight: bold;">Half Day</span>'
			elif d.status == "Work From Home":
				d.status = '<span style="color: purple; font-weight: bold;">Work From Home</span>'

	return data


def format_float_precision(value):
	"""Format float values with proper precision"""
	if not value:
		return 0
	precision = cint(frappe.db.get_default("float_precision")) or 2
	return flt(value, precision)


# ========== Excel Export Functions ==========

@frappe.whitelist()
def export_attendance_excel(filters=None):
	"""Export attendance data to Excel format similar to the C&B template"""
	import json
	import time
	if isinstance(filters, str):
		filters = json.loads(filters)

	# Estimate operation size
	date_range = get_export_date_range(filters)
	num_days = len(date_range['dates'])

	# Get employee count estimate
	filters_temp = {
		"from_date": date_range['from_date'],
		"to_date": date_range['to_date']
	}

	emp_filter = ""
	if filters.get('department'):
		dept_list = filters['department'] if isinstance(filters['department'], list) else [filters['department']]
		dept_str = "', '".join(dept_list)
		emp_filter += f" AND emp.department IN ('{dept_str}')"

	# Only employees matching the configured ID prefix are exported
	from customize_erpnext.customize_erpnext.doctype.attendance_calculation_setting.attendance_calculation_setting import get_attendance_settings
	emp_prefix = (get_attendance_settings().employee_id_prefix or "").strip()
	if emp_prefix:
		filters_temp["emp_prefix"] = f"{emp_prefix}%"
		emp_filter += " AND emp.name LIKE %(emp_prefix)s"

	employee_count = frappe.db.sql(f"""
		SELECT COUNT(DISTINCT emp.name) as count
		FROM `tabEmployee` emp
		WHERE (
			(emp.status = 'Active'
			 AND (emp.date_of_joining IS NULL OR emp.date_of_joining <= %(to_date)s))
			OR
			(emp.status = 'Left'
			 AND (emp.date_of_joining IS NULL OR emp.date_of_joining <= %(to_date)s)
			 AND (emp.relieving_date IS NULL OR emp.relieving_date > %(from_date)s))
		)
		{emp_filter}
	""", filters_temp, as_dict=1)[0].count

	total_operations = employee_count * num_days

	# Large exports run in background (see EXPORT_ASYNC_THRESHOLD benchmark note)
	if total_operations > EXPORT_ASYNC_THRESHOLD:
		# Dedup per user+range: double-clicking Export does not start a second job
		job_id = f"export_excel_{frappe.session.user}_{date_range['from_date']}_{date_range['to_date']}"

		job = frappe.enqueue(
			'customize_erpnext.customize_erpnext.report.shift_attendance_customize.shift_attendance_customize.export_attendance_excel_job',
			queue='default',
			timeout=1800,  # 30 minutes
			job_id=job_id,
			deduplicate=True,
			filters=filters
		)

		if job is None:
			# Duplicate dropped — a job for this user+range is already queued/running
			return {
				'background_job': True,
				'job_id': job_id,
				'total_operations': total_operations,
				'message': 'This export is already running. Please wait for the notification.'
			}

		return {
			'background_job': True,
			'job_id': job_id,
			'total_operations': total_operations,
			'message': f'Large export ({employee_count} employees × {num_days} days = {total_operations} records) queued for background processing. You will receive a notification (bell icon) when ready.'
		}

	# Small operation - run synchronously
	return export_attendance_excel_sync(filters)


def export_attendance_excel_sync(filters, is_background=False):
	"""Excel export pipeline — standard-app (Flutter timesheet) 6-sheet format.

	Sheets: Important Note | Detail | Summary | Timesheet | Overtime | Shift
	(see standard_export.py). is_background=True adds realtime progress updates
	and a longer file retention (user may come back later for the download)."""
	import json
	import tempfile
	import os

	if isinstance(filters, str):
		filters = json.loads(filters)

	def _progress(percent, description):
		if is_background:
			frappe.publish_progress(percent, title="Export Excel", description=description)

	from customize_erpnext.customize_erpnext.report.shift_attendance_customize.standard_export import (
		build_standard_workbook, standard_export_filename)

	date_range = get_export_date_range(filters)
	_progress(15, "Building standard timesheet workbook...")

	wb = build_standard_workbook(
		date_range['from_date'], date_range['to_date'],
		department=filters.get('department') if filters else None)
	_progress(85, "Saving Excel file...")

	filename = standard_export_filename(date_range['from_date'], date_range['to_date'])

	# Save to temporary file
	temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
	wb.save(temp_file.name)
	temp_file.close()

	# Save file to Frappe's private files
	try:
		with open(temp_file.name, 'rb') as f:
			file_content = f.read()

		# Create a File document in Frappe
		file_doc = frappe.get_doc({
			'doctype': 'File',
			'file_name': filename,
			'is_private': 0,  # Public file so it can be downloaded
			'content': file_content
		})
		file_doc.save(ignore_permissions=True)
		frappe.db.commit()

		# Clean up temp file
		os.unlink(temp_file.name)

		# File cleanup:
		# - sync: delete after 2 minutes (user downloads immediately)
		# - background: NO per-file job — a sleeping delete job would occupy a
		#   worker for the whole retention window. The hourly scheduler task
		#   cleanup_export_files() removes export files older than 45 minutes.
		if not is_background:
			frappe.enqueue(
				'customize_erpnext.customize_erpnext.report.shift_attendance_customize.shift_attendance_customize.delete_export_file',
				queue='short',
				timeout=300,
				file_name=file_doc.name,
				enqueue_after_commit=True
			)

		_progress(100, "Export complete")
		return {
			'file_url': file_doc.file_url,
			'filename': filename
		}
	except Exception as e:
		# Clean up temp file on error
		if os.path.exists(temp_file.name):
			os.unlink(temp_file.name)
		frappe.log_error(f"Error saving Excel file: {str(e)}", "Attendance Export Error")
		frappe.throw(_("Failed to save Excel file. Please try again or contact support."))


def export_attendance_excel_job(filters):
	"""Background job for Excel export"""
	try:
		frappe.publish_progress(0, title="Starting Excel export...",
							  description="Generating attendance Excel file in background")

		result = export_attendance_excel_sync(filters, is_background=True)

		# Bell notification with download link — survives page navigation
		# (the realtime event below is lost if the user left the page)
		try:
			frappe.get_doc({
				'doctype': 'Notification Log',
				'subject': f"Excel export ready: {result.get('filename')}",
				'for_user': frappe.session.user,
				'type': 'Alert',
				'email_content': f"<a href='{result.get('file_url')}'>Download {result.get('filename')}</a> (file kept for 30 minutes)",
			}).insert(ignore_permissions=True)
			frappe.db.commit()
		except Exception:
			frappe.log_error(frappe.get_traceback(), "Export Notification Log Error")

		frappe.publish_realtime(
			event='excel_export_complete',
			message={
				"success": True,
				"file_url": result.get('file_url'),
				"filename": result.get('filename'),
				"message": f"Excel export completed! File: {result.get('filename')}"
			},
			user=frappe.session.user
		)

		return result

	except Exception as e:
		frappe.log_error(f"Background Excel export failed: {str(e)}", "Excel Export Error")
		frappe.publish_realtime(
			event='excel_export_complete',
			message={
				"success": False,
				"error": str(e),
				"message": "Excel export failed. Please try again or contact support."
			},
			user=frappe.session.user
		)
		raise e


def get_export_date_range(filters):
	"""Get date range based on current report filters"""
	from frappe.utils import getdate, add_days, formatdate

	# Use from_date and to_date from filters
	from_date = getdate(filters.get("from_date")) if filters.get("from_date") else getdate()
	to_date = getdate(filters.get("to_date")) if filters.get("to_date") else getdate()

	# Generate all dates in range
	dates = []
	current_date = from_date
	while current_date <= to_date:
		dates.append(current_date)
		current_date = add_days(current_date, 1)

	# Create display info
	period_name = f"{formatdate(from_date)} - {formatdate(to_date)}"

	return {
		'from_date': from_date,
		'to_date': to_date,
		'dates': dates,
		'period_name': period_name
	}


def clean_department_name(dept_name):
	"""Remove '- TIQN' suffix from department names"""
	if dept_name and dept_name.endswith(' - TIQN'):
		return dept_name.replace(' - TIQN', '')
	return dept_name or ''


def decimal_round(value, precision=2):
	"""Round decimal values properly"""
	from decimal import Decimal, ROUND_HALF_UP
	if value is None or value == '':
		return 0
	return float(Decimal(str(value)).quantize(Decimal(10) ** -precision, rounding=ROUND_HALF_UP))


def is_sunday_or_holiday(date_obj, holidays_set):
	"""Check if a date is Sunday or a holiday"""
	if date_obj.weekday() == 6:  # Sunday
		return True
	if date_obj in holidays_set:
		return True
	return False


def get_overtime_multipliers():
	"""Get multipliers from all Overtime Types"""
	ot_types = frappe.get_all("Overtime Type", fields=[
		"name", "standard_multiplier", "public_holiday_multiplier", "weekend_multiplier"
	])
	return {ot.name: ot for ot in ot_types}


def calculate_working_day_for_excel(abbreviation, working_hours):
	"""
	Calculate working day value based on leave abbreviation and working hours.

	Rules based on company leave regulations:
	- 1 day for: P, P/2, MC, HS, HL, HL/2 (Phép năm, Ma chay, Hỉ sự, Nghỉ hưởng lương)
	- 0 day for: KL, NB, TS, DS, O, CO, OCO/2, OK/2, COK/2 (Không lương, Nghỉ bù, Thai sản, Dưỡng sức, Ốm, Con ốm)
	- 0.5 day for: O/2, CO/2, OP/2, COP/2 (Ốm/Con ốm - Đi làm/Phép năm nửa ngày)
	- 0.4 day for: OL/2, COL/2 (Ốm/Con ốm - Đi trễ/về sớm ≤ 1 giờ)
	- Other cases: 1 - (8 - working_hours) / 8
	"""
	# Define abbreviations for each category
	full_day_abbrevs = ['P', 'P/2', 'MC', 'HS', 'HL', 'HL/2']
	zero_day_abbrevs = ['KL', 'NB', 'TS', 'DS', 'O', 'CO', 'OCO/2', 'OK/2', 'COK/2']
	half_day_abbrevs = ['O/2', 'CO/2', 'OP/2', 'COP/2']
	point_four_day_abbrevs = ['OL/2', 'COL/2']  # 0.4 day for late/early ≤ 1 hour

	if abbreviation in full_day_abbrevs:
		return 1
	elif abbreviation in zero_day_abbrevs:
		return 0
	elif abbreviation in half_day_abbrevs:
		return 0.5
	elif abbreviation in point_four_day_abbrevs:
		return 0.4
	else:
		# Other cases: formula 1 - (8 - working_hours) / 8
		working_hours = working_hours or 0
		return decimal_round(1 - (8 - working_hours) / 8, 2)


def get_excel_cell_display(abbreviation, working_hours):
	"""
	Get the value to display in Excel cell.

	Rules:
	- KL & working_hours = 0: show "KL"
	- KL & working_hours > 0: show working_hours
	- Other abbreviations: show the abbreviation
	"""
	if abbreviation == 'KL':
		if working_hours and working_hours > 0:
			return decimal_round(working_hours, 2)
		else:
			return 'KL'
	elif abbreviation:
		return abbreviation
	else:
		# No abbreviation - show working hours / 8 as working day
		if working_hours and working_hours > 0:
			return decimal_round(working_hours / 8, 2)
		return ''


def populate_single_employee_row(ws, emp, row, date_range, stt_number, sheet_type='timesheet'):
	"""Populate a single employee row"""
	from datetime import datetime
	from openpyxl.styles import Font, Alignment
	from openpyxl.utils import get_column_letter

	# Basic employee info
	ws.cell(row=row, column=1, value=stt_number)  # STT
	ws.cell(row=row, column=2, value=emp['employee_id'])
	ws.cell(row=row, column=3, value=emp['employee_name'])

	# Format dates as dd/mm/yyyy
	if emp['date_of_joining']:
		if hasattr(emp['date_of_joining'], 'strftime'):
			ws.cell(row=row, column=4, value=emp['date_of_joining'].strftime('%d/%m/%Y'))
		else:
			# Parse string date and format
			try:
				date_obj = datetime.strptime(str(emp['date_of_joining']), '%Y-%m-%d')
				ws.cell(row=row, column=4, value=date_obj.strftime('%d/%m/%Y'))
			except:
				ws.cell(row=row, column=4, value=str(emp['date_of_joining']))
	else:
		ws.cell(row=row, column=4, value='')

	if emp['relieving_date']:
		if hasattr(emp['relieving_date'], 'strftime'):
			ws.cell(row=row, column=5, value=emp['relieving_date'].strftime('%d/%m/%Y'))
		else:
			# Parse string date and format
			try:
				date_obj = datetime.strptime(str(emp['relieving_date']), '%Y-%m-%d')
				ws.cell(row=row, column=5, value=date_obj.strftime('%d/%m/%Y'))
			except:
				ws.cell(row=row, column=5, value=str(emp['relieving_date']))
	else:
		ws.cell(row=row, column=5, value='')

	ws.cell(row=row, column=6, value=emp['custom_group'])
	ws.cell(row=row, column=7, value='')  # Section - not available in Attendance
	ws.cell(row=row, column=8, value=emp['designation'])  # Position/Chức vụ

	# Set font and alignment for employee data (size 8)
	for col in range(1, 9):  # Columns 1-8
		cell = ws.cell(row=row, column=col)
		cell.font = STYLE_CELL_FONT
		# Left-align these specific columns
		cell.alignment = STYLE_ALIGN_LEFT

	# Daily data based on sheet type
	total_value = 0
	total_ot_multiplied = 0  # For overtime sheet: sum of OT x multiplier
	for j, date_obj in enumerate(date_range['dates']):
		col = 9 + j
		date_key = date_obj.strftime('%Y-%m-%d')

		if sheet_type == 'overtime':
			# For overtime sheet: use daily_overtime (raw OT) for display
			value = emp.get('daily_overtime', {}).get(date_key, 0) or 0
			# Show value if > 0, otherwise empty
			display_value = value if value > 0 else ''
			total_value += value
			# Sum up multiplied OT for the extra column
			ot_mult = emp.get('daily_overtime_multiplied', {}).get(date_key, 0) or 0
			total_ot_multiplied += ot_mult
		else:
			# For timesheet: use daily_data and daily_working_day
			display_value = emp['daily_data'].get(date_key, '')
			working_day = emp.get('daily_working_day', {}).get(date_key, 0) or 0
			total_value += working_day

		ws.cell(row=row, column=col, value=display_value)

		# Center align attendance data with font
		cell = ws.cell(row=row, column=col)
		cell.font = STYLE_CELL_FONT
		cell.alignment = STYLE_ALIGN_CENTER

	# Add total column with calculated value
	total_col = 9 + len(date_range['dates'])
	cell = ws.cell(row=row, column=total_col, value=decimal_round(total_value, 2))
	cell.font = STYLE_CELL_FONT
	cell.alignment = STYLE_ALIGN_CENTER

	if sheet_type == 'overtime':
		# Add Total OT x Multiplier column for overtime sheet
		total_ot_mult_col = total_col + 1
		cell = ws.cell(row=row, column=total_ot_mult_col, value=decimal_round(total_ot_multiplied, 2))
		cell.font = STYLE_CELL_FONT
		cell.alignment = STYLE_ALIGN_CENTER

		# Add confirmation column (empty)
		confirmation_col = total_ot_mult_col + 1
		cell = ws.cell(row=row, column=confirmation_col, value='')
		cell.font = STYLE_CELL_FONT
		cell.alignment = STYLE_ALIGN_CENTER
	else:
		# Add confirmation column (empty) for timesheet
		confirmation_col = total_col + 1
		cell = ws.cell(row=row, column=confirmation_col, value='')
		cell.font = STYLE_CELL_FONT
		cell.alignment = STYLE_ALIGN_CENTER


def delete_export_file(file_name):
	"""Delete a SYNC-exported Excel file after 2 minutes (download window).
	Background exports are cleaned by cleanup_export_files() instead — a long
	sleep here would occupy a worker for the whole retention window.
	"""
	import time

	time.sleep(120)

	try:
		if frappe.db.exists('File', file_name):
			frappe.delete_doc('File', file_name, ignore_permissions=True)
			frappe.db.commit()
	except Exception as e:
		# Log error but don't raise - file cleanup is not critical
		frappe.log_error(f"Failed to delete export file {file_name}: {str(e)}", "Export File Cleanup")


# ============================================================================
# LEGACY TIMESHEET-APP STYLE SHEETS
# Mirrors the Excel export of the old Dart timesheet app (see
# overrides/shift_type/LEGACY_APP_TIMESHEET_ALGORITHM.md §13 and the reference file
# Timesheet_260526_260625_*.xlsx): Important Note / Detail / Summary / Shift.
# Data source: the same get_data() detail rows used by the existing sheets.
# ============================================================================


def cleanup_export_files():
	"""Hourly scheduler task: delete attendance export files older than 45
	minutes (covers background exports — kept ~30+ min so the user can return
	via the bell notification — plus any leftovers from crashed jobs)."""
	old_files = frappe.get_all(
		"File",
		filters={
			# Timesheet_%.xlsx = current (standard-app) export naming;
			# Attendance_%.xlsx = legacy naming, kept for leftover files
			"file_name": ["like", "%.xlsx"],
			"is_private": 0,
			"creation": ["<", frappe.utils.add_to_date(frappe.utils.now_datetime(), minutes=-45)],
		},
		or_filters=[
			["file_name", "like", "Timesheet\\_%.xlsx"],
			["file_name", "like", "Attendance\\_%.xlsx"],
		],
		pluck="name",
	)
	for name in old_files:
		try:
			frappe.delete_doc("File", name, ignore_permissions=True)
		except Exception:
			frappe.log_error(frappe.get_traceback(), "Export File Cleanup Error")
	if old_files:
		frappe.db.commit()
